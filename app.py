#!/usr/bin/env python3
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from pathlib import Path
import json
from datetime import datetime, time as dtime
from urllib.parse import quote
import requests  # Dla Llama/Ollama
import threading
import time
import queue
import re
import unicodedata
from database import DatabaseManager
from modules import PDFReader, KWExtractor, EntityExtractor, DocumentTagger
from config import CONFIG
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
CORS(app)

# Inicjalizacja bazy
def init_app():
    """Inicjalizacja bazy + migracje (dodaje brakujące kolumny)"""
    db_path = Path(CONFIG['db_path'])
    db = DatabaseManager(CONFIG['db_path'])
    db.connect()

    # Pierwsze uruchomienie - utwórz schemat
    if not db_path.exists() or db_path.stat().st_size == 0:
        db.init_database()

    # MIGRACJE - sprawdź i dodaj brakujące kolumny
    try:
        # Sprawdź istniejące kolumny w tabeli pages
        cursor = db.connection.execute("PRAGMA table_info(pages)")
        existing_columns = {row[1] for row in cursor.fetchall()}

        # Lista wymaganych kolumn z migracji
        migrations = [
            ('fixed_text', 'ALTER TABLE pages ADD COLUMN fixed_text TEXT DEFAULT NULL'),
            ('ocr_fixed_at', 'ALTER TABLE pages ADD COLUMN ocr_fixed_at TIMESTAMP DEFAULT NULL'),
            ('ocr_confidence', 'ALTER TABLE pages ADD COLUMN ocr_confidence TEXT DEFAULT NULL'),
        ]

        for col_name, alter_sql in migrations:
            if col_name not in existing_columns:
                print(f"📦 Migracja: dodaję kolumnę '{col_name}' do tabeli pages...")
                db.connection.execute(alter_sql)
                db.connection.commit()
                print(f"   ✅ Dodano '{col_name}'")

        # Sprawdź kolumny source_files
        cursor = db.connection.execute("PRAGMA table_info(source_files)")
        source_file_columns = {row[1] for row in cursor.fetchall()}
        source_file_migrations = [
            ('section_name', "ALTER TABLE source_files ADD COLUMN section_name TEXT DEFAULT ''"),
        ]

        for col_name, alter_sql in source_file_migrations:
            if col_name not in source_file_columns:
                print(f"📦 Migracja: dodaję kolumnę '{col_name}' do tabeli source_files...")
                db.connection.execute(alter_sql)
                db.connection.commit()
                print(f"   ✅ Dodano '{col_name}'")

        # Sprawdź kolumny land_registers
        cursor = db.connection.execute("PRAGMA table_info(land_registers)")
        land_register_columns = {row[1] for row in cursor.fetchall()}
        land_register_migrations = [
            ('geoportal_url', "ALTER TABLE land_registers ADD COLUMN geoportal_url TEXT DEFAULT ''"),
        ]

        for col_name, alter_sql in land_register_migrations:
            if col_name not in land_register_columns:
                print(f"📦 Migracja: dodaję kolumnę '{col_name}' do tabeli land_registers...")
                db.connection.execute(alter_sql)
                db.connection.commit()
                print(f"   ✅ Dodano '{col_name}'")

        # Sprawdź tabelę file_ocr_stats
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='file_ocr_stats'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'file_ocr_stats'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS file_ocr_stats (
                    file_id INTEGER PRIMARY KEY,
                    total_pages INTEGER DEFAULT 0,
                    fixed_pages INTEGER DEFAULT 0,
                    last_fixed_at TIMESTAMP DEFAULT NULL,
                    analysis_complete INTEGER DEFAULT 0,
                    FOREIGN KEY (file_id) REFERENCES source_files(id)
                )
            """)
            db.connection.commit()
            print("   ✅ Utworzono 'file_ocr_stats'")

        # Sprawdź tabelę ocr_training_examples (przykłady do nauki Llamy)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='ocr_training_examples'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'ocr_training_examples'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS ocr_training_examples (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER,
                    page_id INTEGER,
                    page_number INTEGER,
                    original_text TEXT NOT NULL,
                    corrected_text TEXT NOT NULL,
                    diff_size INTEGER DEFAULT 0,
                    source TEXT DEFAULT 'manual',
                    used_count INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (file_id) REFERENCES source_files(id),
                    FOREIGN KEY (page_id) REFERENCES pages(id)
                )
            """)
            db.connection.execute(
                "CREATE INDEX IF NOT EXISTS idx_training_created ON ocr_training_examples(created_at DESC)"
            )
            db.connection.commit()
            print("   ✅ Utworzono 'ocr_training_examples'")

        # Sprawdź tabelę person_aliases (łączenie różnych form tej samej osoby)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='person_aliases'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'person_aliases'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS person_aliases (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    canonical_entity_id INTEGER NOT NULL,
                    alias_entity_id INTEGER NOT NULL,
                    confidence REAL DEFAULT 1.0,
                    match_reason TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (canonical_entity_id) REFERENCES entities(id),
                    FOREIGN KEY (alias_entity_id) REFERENCES entities(id),
                    UNIQUE(canonical_entity_id, alias_entity_id)
                )
            """)
            db.connection.execute(
                "CREATE INDEX IF NOT EXISTS idx_aliases_canonical ON person_aliases(canonical_entity_id)"
            )
            db.connection.execute(
                "CREATE INDEX IF NOT EXISTS idx_aliases_alias ON person_aliases(alias_entity_id)"
            )
            db.connection.commit()
            print("   ✅ Utworzono 'person_aliases'")

        # Sprawdź tabelę person_attributes (PESEL, adres, NIP osoby)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='person_attributes'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'person_attributes'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS person_attributes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    entity_id INTEGER NOT NULL,
                    attr_type TEXT NOT NULL,
                    attr_value TEXT NOT NULL,
                    page_id INTEGER,
                    file_id INTEGER,
                    confidence REAL DEFAULT 0.7,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (entity_id) REFERENCES entities(id),
                    UNIQUE(entity_id, attr_type, attr_value)
                )
            """)
            db.connection.execute(
                "CREATE INDEX IF NOT EXISTS idx_pattr_entity ON person_attributes(entity_id)"
            )
            db.connection.execute(
                "CREATE INDEX IF NOT EXISTS idx_pattr_value ON person_attributes(attr_type, attr_value)"
            )
            db.connection.commit()
            print("   ✅ Utworzono 'person_attributes'")

        # Sprawdź czy tabela 'entities' ma kolumny do weryfikacji
        cursor = db.connection.execute("PRAGMA table_info(entities)")
        ent_cols = {row[1] for row in cursor.fetchall()}
        if 'verified' not in ent_cols:
            print("📦 Migracja: dodaję 'verified' do entities...")
            db.connection.execute("ALTER TABLE entities ADD COLUMN verified INTEGER DEFAULT 0")
        if 'verified_at' not in ent_cols:
            print("📦 Migracja: dodaję 'verified_at' do entities...")
            db.connection.execute("ALTER TABLE entities ADD COLUMN verified_at TIMESTAMP")
        if 'verified_by' not in ent_cols:
            print("📦 Migracja: dodaję 'verified_by' do entities...")
            db.connection.execute("ALTER TABLE entities ADD COLUMN verified_by TEXT")
        db.connection.commit()

        # Sprawdź tabelę file_summaries (auto-summary przez Llamę)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='file_summaries'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'file_summaries'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS file_summaries (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER NOT NULL UNIQUE,
                    summary TEXT NOT NULL,
                    short_summary TEXT,
                    model_used TEXT,
                    generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (file_id) REFERENCES source_files(id)
                )
            """)
            db.connection.execute("CREATE INDEX IF NOT EXISTS idx_summary_file ON file_summaries(file_id)")
            db.connection.commit()
            print("   ✅ Utworzono 'file_summaries'")

        # Sprawdź tabelę tags (jeśli istnieje) - dodaj kolumny color i icon
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='document_tags'"
        )
        if cursor.fetchone():
            # Sprawdź czy kolumny istnieją
            cursor = db.connection.execute("PRAGMA table_info(document_tags)")
            existing_cols = {row[1] for row in cursor.fetchall()}
            if 'color' not in existing_cols:
                print("📦 Migracja: dodaję 'color' do document_tags...")
                db.connection.execute("ALTER TABLE document_tags ADD COLUMN color TEXT DEFAULT '#6366f1'")
            if 'icon' not in existing_cols:
                print("📦 Migracja: dodaję 'icon' do document_tags...")
                db.connection.execute("ALTER TABLE document_tags ADD COLUMN icon TEXT DEFAULT '🏷️'")
            db.connection.commit()

        # Tabela custom_tags (tagi użytkownika z kolorami i emoji)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='custom_tags'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'custom_tags'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS custom_tags (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    color TEXT DEFAULT '#6366f1',
                    icon TEXT DEFAULT '🏷️',
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS file_custom_tags (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER NOT NULL,
                    tag_id INTEGER NOT NULL,
                    added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(file_id, tag_id),
                    FOREIGN KEY (file_id) REFERENCES source_files(id),
                    FOREIGN KEY (tag_id) REFERENCES custom_tags(id)
                )
            """)
            db.connection.commit()
            # Dodaj domyślne tagi
            default_tags = [
                ('Ważne', '#ef4444', '🔴', 'Wymaga uwagi'),
                ('Zakończone', '#10b981', '🟢', 'Sprawa zamknięta'),
                ('Do przeglądu', '#3b82f6', '🔵', 'Czeka na review'),
                ('Sąd', '#a855f7', '🟣', 'Sprawa sądowa'),
                ('Pilne', '#f59e0b', '🟠', 'Termin'),
                ('Finanse', '#14b8a6', '💰', 'Faktury, płatności'),
            ]
            for name, color, icon, desc in default_tags:
                db.connection.execute(
                    "INSERT OR IGNORE INTO custom_tags (name, color, icon, description) VALUES (?, ?, ?, ?)",
                    (name, color, icon, desc)
                )
            db.connection.commit()
            print("   ✅ Utworzono 'custom_tags' + 'file_custom_tags' z domyślnymi tagami")

        # Sprawdź tabelę document_annotations (adnotacje + highlighty na PDF)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='document_annotations'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'document_annotations'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS document_annotations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER NOT NULL,
                    page_number INTEGER NOT NULL,
                    annotation_type TEXT NOT NULL DEFAULT 'note',
                    color TEXT DEFAULT 'yellow',
                    note_text TEXT,
                    selected_text TEXT,
                    position_x REAL,
                    position_y REAL,
                    width REAL,
                    height REAL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (file_id) REFERENCES source_files(id)
                )
            """)
            db.connection.execute("CREATE INDEX IF NOT EXISTS idx_ann_file_page ON document_annotations(file_id, page_number)")
            db.connection.commit()
            print("   ✅ Utworzono 'document_annotations'")

        # Sprawdź tabelę favorites (ulubione dokumenty/osoby)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='favorites'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'favorites'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS favorites (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    favorite_type TEXT NOT NULL,
                    target_id INTEGER NOT NULL,
                    note TEXT,
                    pin_order INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(favorite_type, target_id)
                )
            """)
            db.connection.commit()
            print("   ✅ Utworzono 'favorites'")

        # Sprawdź tabelę duplicates (wykryte duplikaty)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='duplicates'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'duplicates'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS duplicates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id_1 INTEGER NOT NULL,
                    file_id_2 INTEGER NOT NULL,
                    similarity REAL NOT NULL,
                    match_type TEXT NOT NULL,
                    status TEXT DEFAULT 'pending',
                    detected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(file_id_1, file_id_2),
                    FOREIGN KEY (file_id_1) REFERENCES source_files(id),
                    FOREIGN KEY (file_id_2) REFERENCES source_files(id)
                )
            """)
            db.connection.commit()
            print("   ✅ Utworzono 'duplicates'")

        # Sprawdź tabelę document_types (auto-rozpoznane typy)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='document_types'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'document_types'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS document_types (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_id INTEGER NOT NULL,
                    page_id INTEGER,
                    doc_type TEXT NOT NULL,
                    confidence REAL DEFAULT 0.5,
                    extracted_fields TEXT,
                    detected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (file_id) REFERENCES source_files(id),
                    FOREIGN KEY (page_id) REFERENCES pages(id)
                )
            """)
            db.connection.execute("CREATE INDEX IF NOT EXISTS idx_doctypes_file ON document_types(file_id)")
            db.connection.commit()
            print("   ✅ Utworzono 'document_types'")

        # Sprawdź tabelę addresses (adresy + geocoding)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='addresses'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'addresses'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS addresses (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    address_text TEXT NOT NULL UNIQUE,
                    city TEXT,
                    street TEXT,
                    postal_code TEXT,
                    lat REAL,
                    lng REAL,
                    geocoded_at TIMESTAMP,
                    geocode_source TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            db.connection.execute("CREATE INDEX IF NOT EXISTS idx_addr_city ON addresses(city)")
            db.connection.commit()
            print("   ✅ Utworzono 'addresses'")

        # Sprawdź tabelę address_occurrences (gdzie pojawia się adres)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='address_occurrences'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'address_occurrences'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS address_occurrences (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    address_id INTEGER NOT NULL,
                    file_id INTEGER NOT NULL,
                    page_id INTEGER,
                    person_entity_id INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (address_id) REFERENCES addresses(id),
                    FOREIGN KEY (file_id) REFERENCES source_files(id),
                    FOREIGN KEY (page_id) REFERENCES pages(id)
                )
            """)
            db.connection.execute("CREATE INDEX IF NOT EXISTS idx_addrocc_addr ON address_occurrences(address_id)")
            db.connection.execute("CREATE INDEX IF NOT EXISTS idx_addrocc_file ON address_occurrences(file_id)")
            db.connection.commit()
            print("   ✅ Utworzono 'address_occurrences'")

        # Sprawdź tabelę processing_queue (kolejka zadań w tle)
        cursor = db.connection.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='processing_queue'"
        )
        if not cursor.fetchone():
            print("📦 Migracja: tworzę tabelę 'processing_queue'...")
            db.connection.execute("""
                CREATE TABLE IF NOT EXISTS processing_queue (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_path TEXT,
                    file_id INTEGER,
                    filename TEXT,
                    binder_name TEXT DEFAULT 'Default',
                    section_name TEXT DEFAULT '',
                    task_type TEXT NOT NULL DEFAULT 'full_process',
                    status TEXT NOT NULL DEFAULT 'pending',
                    priority INTEGER DEFAULT 5,
                    retry_count INTEGER DEFAULT 0,
                    max_retries INTEGER DEFAULT 3,
                    error_message TEXT,
                    progress_percent INTEGER DEFAULT 0,
                    progress_text TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    started_at TIMESTAMP,
                    completed_at TIMESTAMP
                )
            """)
            db.connection.execute(
                "CREATE INDEX IF NOT EXISTS idx_queue_status ON processing_queue(status, priority, created_at)"
            )
            db.connection.commit()
            print("   ✅ Utworzono 'processing_queue'")
        else:
            cursor = db.connection.execute("PRAGMA table_info(processing_queue)")
            queue_columns = {row[1] for row in cursor.fetchall()}
            if 'section_name' not in queue_columns:
                print("📦 Migracja: dodaję kolumnę 'section_name' do tabeli processing_queue...")
                db.connection.execute("ALTER TABLE processing_queue ADD COLUMN section_name TEXT DEFAULT ''")
                db.connection.commit()
                print("   ✅ Dodano 'section_name'")

        cursor = db.connection.execute(
            "SELECT sql FROM sqlite_master WHERE type='view' AND name='view_land_registers_summary'"
        )
        view_row = cursor.fetchone()
        if view_row and ('kw_checksum' not in (view_row[0] or '') or 'geoportal_url' not in (view_row[0] or '')):
            print("📦 Migracja: odświeżam widok ksiąg wieczystych...")
            db.connection.execute("DROP VIEW IF EXISTS view_land_registers_summary")
            db.connection.execute("""
                CREATE VIEW view_land_registers_summary AS
                SELECT
                    lr.id,
                    lr.kw_full,
                    lr.kw_district,
                    lr.kw_number,
                    lr.kw_checksum,
                    COUNT(DISTINCT lro.file_id) as files_count,
                    COUNT(DISTINCT lro.page_id) as pages_count,
                    lr.property_address,
                    lr.owner_manual,
                    lr.geoportal_url
                FROM land_registers lr
                LEFT JOIN land_register_occurrences lro ON lr.id = lro.kw_id
                GROUP BY lr.id
            """)
            db.connection.commit()
            print("   ✅ Widok KW odświeżony")

    except Exception as e:
        print(f"⚠️ Błąd migracji: {e}")

    db.disconnect()

init_app()


def normalize_section_name(value):
    """Ujednolić nazwę sekcji dokumentu."""
    return (value or '').strip()


def make_safe_pdf_filename(name, default='dokument'):
    """Zamień dowolny tytuł na bezpieczną nazwę pliku PDF."""
    base = unicodedata.normalize('NFKD', (name or default)).encode('ascii', 'ignore').decode('ascii')
    base = re.sub(r'[^A-Za-z0-9._-]+', '_', base).strip('._-')
    base = base or default
    if not base.lower().endswith('.pdf'):
        base += '.pdf'
    return base


def build_geoportal_link(address=''):
    """Link do Geoportalu - jeśli jest adres, dodaj go jako query string."""
    base = 'https://mapy.geoportal.gov.pl/'
    address = (address or '').strip()
    if not address or address == '—':
        return base
    return f"{base}?q={quote(address)}"


def normalize_optional_url(value):
    """Ujednolić ręcznie wpisany URL, bez wymuszania go gdy pole jest puste."""
    value = (value or '').strip()
    if not value or value == '—':
        return ''
    if not re.match(r'^https?://', value, re.IGNORECASE):
        return f'https://{value}'
    return value


def parse_split_ranges(ranges_text, total_pages):
    """Parsuj zakresy typu 1-10,11-20,25."""
    ranges = []
    for raw_part in (ranges_text or '').split(','):
        part = raw_part.strip()
        if not part:
            continue
        if '-' in part:
            start_text, end_text = part.split('-', 1)
            start = int(start_text.strip())
            end = int(end_text.strip())
        else:
            start = end = int(part)

        if start < 1 or end < 1 or start > total_pages or end > total_pages or start > end:
            raise ValueError(f"Nieprawidłowy zakres: {part}")

        ranges.append((start, end))

    if not ranges:
        raise ValueError("Podaj przynajmniej jeden zakres stron")

    return ranges


def _delete_file_related_data(db, file_id):
    """Usuń dokument oraz wszystkie powiązane rekordy z bazy."""
    page_ids = [row['id'] for row in db.fetch_all("SELECT id FROM pages WHERE file_id = ?", (file_id,))]

    if page_ids:
        placeholders = ','.join('?' * len(page_ids))
        db.execute(f"DELETE FROM entity_occurrences WHERE page_id IN ({placeholders})", tuple(page_ids))
        db.execute(f"DELETE FROM cooccurrences WHERE page_id IN ({placeholders})", tuple(page_ids))
        db.execute(f"DELETE FROM ocr_corrections WHERE page_id IN ({placeholders})", tuple(page_ids))
        db.execute(f"DELETE FROM land_register_occurrences WHERE page_id IN ({placeholders})", tuple(page_ids))

    db.execute("DELETE FROM entity_occurrences WHERE file_id = ?", (file_id,))
    db.execute("DELETE FROM land_register_occurrences WHERE file_id = ?", (file_id,))
    db.execute("DELETE FROM document_tags WHERE file_id = ?", (file_id,))
    db.execute("DELETE FROM pages WHERE file_id = ?", (file_id,))
    db.execute("DELETE FROM source_files WHERE id = ?", (file_id,))

# API ENDPOINTS

@app.route('/')
def index():
    """Strona główna"""
    return render_template('index.html')

@app.route('/api/file/<int:file_id>')
def get_file(file_id):
    """Pobierz plik PDF"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        file_data = db.fetch_one("SELECT filepath FROM source_files WHERE id = ?", (file_id,))
        db.disconnect()

        if not file_data:
            return jsonify({'error': 'Plik nie znaleziony'}), 404

        filepath = file_data['filepath']
        return send_file(filepath, mimetype='application/pdf')

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-import', methods=['POST'])
def batch_import():
    """Batch import PDFów z pełnym auto-process pipeline"""
    try:
        imports_dir = Path(CONFIG['imports_dir'])
        pdf_files = list(imports_dir.glob('*.pdf'))

        if not pdf_files:
            return jsonify({'error': 'Brak plików PDF w folderze imports'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        total = {
            'files_imported': 0,
            'pages_processed': 0,
            'kw_found': 0,
            'persons_found': 0,
            'companies_found': 0,
            'institutions_found': 0,
            'nips_found': 0,
            'regons_found': 0,
            'krs_found': 0,
            'pesels_found': 0,
            'phones_found': 0,
            'emails_found': 0,
            'signatures_found': 0,
            'tags_added': 0,
            'cooccurrences_built': 0
        }

        errors = []
        for pdf_path in pdf_files:
            try:
                print(f"\n📦 Batch import: {pdf_path.name}")
                result = auto_process_pdf(pdf_path, pdf_path.stem, db=db)
                if result and result.get('success'):
                    total['files_imported'] += 1
                    for key, value in result['stats'].items():
                        if key in total:
                            total[key] += value
                else:
                    errors.append(f"{pdf_path.name}: brak rezultatu")
            except Exception as file_err:
                import traceback
                err_msg = f"{pdf_path.name}: {file_err}"
                errors.append(err_msg)
                print(f"❌ Błąd w {pdf_path.name}: {file_err}")
                print(traceback.format_exc())

        db.disconnect()

        return jsonify({
            'success': True,
            'errors': errors if errors else None,
            **total
        })

    except Exception as e:
        import traceback
        print(f"❌ batch_import error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

def auto_process_pdf(filepath, binder_name='Default', db=None):
    """
    Pełny auto-process pipeline dla PDF:
    1. OCR / odczyt tekstu
    2. Dodanie do bazy
    3. Ekstrakcja KW
    4. Ekstrakcja encji (osoby, firmy, NIP, REGON, KRS, PESEL, tel, email)
    5. Auto-tagowanie dokumentu
    6. Budowa cooccurrences dla tego pliku
    """
    reader = PDFReader()
    kw_extractor = KWExtractor()
    entity_extractor = EntityExtractor()
    tagger = DocumentTagger()

    own_db = False
    if db is None:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        own_db = True

    stats = {
        'pages_processed': 0,
        'kw_found': 0,
        'persons_found': 0,
        'companies_found': 0,
        'institutions_found': 0,
        'nips_found': 0,
        'regons_found': 0,
        'krs_found': 0,
        'pesels_found': 0,
        'phones_found': 0,
        'emails_found': 0,
        'signatures_found': 0,
        'tags_added': 0,
        'cooccurrences_built': 0
    }

    # Krok 1: Odczyt PDF z OCR
    pdf_data = reader.read_pdf(str(filepath))
    if not pdf_data:
        if own_db:
            db.disconnect()
        return None

    with db.transaction():
        # Krok 2: Dodanie do bazy
        binder_id = db.add_binder(binder_name)
        file_id = db.add_source_file(
            binder_id,
            pdf_data['filename'],
            pdf_data['filepath'],
            pdf_data['file_hash'],
            pdf_data['page_count']
        )

        page_ids = []
        page_texts = []

        # Krok 3-4: Przetwarzanie stron - KW + Encje
        for page_data in pdf_data['pages']:
            page_num = page_data['page_number']
            text = page_data['text']

            page_id = db.add_page(file_id, page_num, text)
            page_ids.append(page_id)
            page_texts.append(text)
            stats['pages_processed'] += 1

            if not text or len(text) < 10:
                continue

            # KW extraction
            kws = kw_extractor.extract_from_text(text, page_num)
            for kw in kws:
                kw_id = db.add_land_register(
                    kw['kw_full'],
                    kw['kw_district'],
                    kw['kw_number'],
                    kw['kw_checksum']
                )
                db.add_land_register_occurrence(
                    kw_id, page_id, file_id,
                    kw['context_before'], kw['context_after']
                )
                stats['kw_found'] += 1

            # Entity extraction
            entities = entity_extractor.extract_all(text)

            # Persons (z obsługą compound names: "Sławomir Krzysztof Feszczak")
            page_person_ids = []  # do auto-attribute matching
            for person in entities.get('persons', []):
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('person', person['full_name'], person.get('normalized_full_name', person['full_name']).upper()))
                entity_row = db.fetch_one(
                    "SELECT id FROM entities WHERE entity_type='person' AND entity_value=?",
                    (person['full_name'],)
                )
                if entity_row:
                    db.execute("""
                        INSERT INTO entity_occurrences (entity_id, page_id, file_id)
                        VALUES (?, ?, ?)
                    """, (entity_row['id'], page_id, file_id))
                    stats['persons_found'] += 1
                    page_person_ids.append(entity_row['id'])

            # AUTO-ATTRIBUTE: Przypisz PESELe/adresy do osób na tej stronie
            # Heurystyka: jeśli na stronie są PESELe i osoby - przypisz każdy PESEL do każdej osoby
            # Później manual review w UI może to skorygować
            page_pesels = [p['pesel'] for p in entities.get('pesels', [])]

            if page_person_ids and page_pesels:
                # Jeśli jedna osoba + jeden PESEL na stronie → wysokie confidence
                if len(page_person_ids) == 1 and len(page_pesels) == 1:
                    db.execute("""
                        INSERT OR IGNORE INTO person_attributes
                        (entity_id, attr_type, attr_value, page_id, file_id, confidence)
                        VALUES (?, 'pesel', ?, ?, ?, 0.9)
                    """, (page_person_ids[0], page_pesels[0], page_id, file_id))
                # Wiele osób + 1 PESEL → przypisz do każdej z niskim confidence
                elif len(page_pesels) == 1:
                    for pid in page_person_ids:
                        db.execute("""
                            INSERT OR IGNORE INTO person_attributes
                            (entity_id, attr_type, attr_value, page_id, file_id, confidence)
                            VALUES (?, 'pesel', ?, ?, ?, 0.5)
                        """, (pid, page_pesels[0], page_id, file_id))

            # Companies
            for company in entities.get('companies', []):
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('company', company['name'], company['name'].upper()))
                entity_row = db.fetch_one(
                    "SELECT id FROM entities WHERE entity_type='company' AND entity_value=?",
                    (company['name'],)
                )
                if entity_row:
                    db.execute("""
                        INSERT INTO entity_occurrences (entity_id, page_id, file_id)
                        VALUES (?, ?, ?)
                    """, (entity_row['id'], page_id, file_id))
                    stats['companies_found'] += 1

            # Institutions
            for inst in entities.get('institutions', []):
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('institution', inst['name'], inst['name'].upper()))
                entity_row = db.fetch_one(
                    "SELECT id FROM entities WHERE entity_type='institution' AND entity_value=?",
                    (inst['name'],)
                )
                if entity_row:
                    db.execute("""
                        INSERT INTO entity_occurrences (entity_id, page_id, file_id)
                        VALUES (?, ?, ?)
                    """, (entity_row['id'], page_id, file_id))
                    stats['institutions_found'] += 1

            # NIPs
            for nip in entities.get('nips', []):
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('nip', nip['formatted'], nip['nip']))
                entity_row = db.fetch_one(
                    "SELECT id FROM entities WHERE entity_type='nip' AND entity_value=?",
                    (nip['formatted'],)
                )
                if entity_row:
                    db.execute("""
                        INSERT INTO entity_occurrences (entity_id, page_id, file_id)
                        VALUES (?, ?, ?)
                    """, (entity_row['id'], page_id, file_id))
                    stats['nips_found'] += 1

            # REGONs, KRSs, PESELs, Phones, Emails, Signatures
            # Mapowanie: (entity_type, klucz_w_entities_dict, klucz_z_item, klucz_w_stats)
            for ent_type, list_key, item_key, stats_key in [
                ('regon', 'regons', 'regon', 'regons_found'),
                ('krs', 'krs', 'krs', 'krs_found'),
                ('pesel', 'pesels', 'pesel', 'pesels_found'),
                ('phone', 'phones', 'phone', 'phones_found'),
                ('email', 'emails', 'email', 'emails_found'),
            ]:
                for item in entities.get(list_key, []):
                    value = item.get(item_key, '')
                    if not value:
                        continue
                    db.execute("""
                        INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                        VALUES (?, ?, ?)
                    """, (ent_type, value, value))
                    entity_row = db.fetch_one(
                        "SELECT id FROM entities WHERE entity_type=? AND entity_value=?",
                        (ent_type, value)
                    )
                    if entity_row:
                        db.execute("""
                            INSERT INTO entity_occurrences (entity_id, page_id, file_id)
                            VALUES (?, ?, ?)
                        """, (entity_row['id'], page_id, file_id))
                        if stats_key in stats:
                            stats[stats_key] += 1

            # Signatures
            for sig in entities.get('signatures', []):
                value = sig.get('value', '')
                if not value:
                    continue
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('signature', value, value.upper()))
                entity_row = db.fetch_one(
                    "SELECT id FROM entities WHERE entity_type='signature' AND entity_value=?",
                    (value,)
                )
                if entity_row:
                    db.execute("""
                        INSERT INTO entity_occurrences (entity_id, page_id, file_id)
                        VALUES (?, ?, ?)
                    """, (entity_row['id'], page_id, file_id))
                    stats['signatures_found'] += 1

        # Krok 5: Auto-tagowanie dokumentu
        valid_texts = [t for t in page_texts if t and len(t) > 10]
        if valid_texts:
            tags = tagger.tag_pages_combined(valid_texts)
            db.execute("DELETE FROM document_tags WHERE file_id = ?", (file_id,))
            for tag in tags[:3]:
                db.execute("""
                    INSERT OR IGNORE INTO document_tags (file_id, tag, confidence)
                    VALUES (?, ?, ?)
                """, (file_id, tag['type'], tag['confidence']))
                stats['tags_added'] += 1

        # Krok 6: Build cooccurrences dla stron tego pliku
        for page_id in page_ids:
            entity_ids_row = db.fetch_one("""
                SELECT GROUP_CONCAT(entity_id) as entity_ids
                FROM entity_occurrences
                WHERE page_id = ?
                GROUP BY page_id
                HAVING COUNT(*) > 1
            """, (page_id,))

            if entity_ids_row and entity_ids_row['entity_ids']:
                entity_ids = [int(x) for x in entity_ids_row['entity_ids'].split(',')]
                for i in range(len(entity_ids)):
                    for j in range(i+1, len(entity_ids)):
                        e1, e2 = sorted([entity_ids[i], entity_ids[j]])
                        db.execute("""
                            INSERT OR IGNORE INTO cooccurrences (entity_id_1, entity_id_2, page_id)
                            VALUES (?, ?, ?)
                        """, (e1, e2, page_id))
                        stats['cooccurrences_built'] += 1

    if own_db:
        db.disconnect()

    # AUTO-LINKING: po zakończeniu importu - połącz osoby
    try:
        link_stats = auto_link_persons()
        stats['aliases_created'] = link_stats.get('by_name', 0) + link_stats.get('by_pesel', 0) + link_stats.get('by_address', 0)
        print(f"   🔗 Auto-link: {stats['aliases_created']} nowych połączeń osób")
    except Exception as e:
        print(f"⚠ Auto-link error: {e}")
        stats['aliases_created'] = 0

    return {
        'success': True,
        'file_id': file_id,
        'filename': pdf_data['filename'],
        'pages': pdf_data['page_count'],
        'stats': stats
    }


@app.route('/api/upload-pdf', methods=['POST'])
def upload_pdf():
    """Upload i pełen auto-process pipeline"""
    import traceback

    try:
        print("\n" + "="*60)
        print(f"🚀 UPLOAD PDF - {datetime.now().strftime('%H:%M:%S')}")
        print(f"Files: {list(request.files.keys())} | Form: {list(request.form.keys())}")

        if 'file' not in request.files:
            print("❌ Brak 'file' w request.files")
            return jsonify({'error': 'Brak pliku w żądaniu'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'error': 'Plik nie wybrany'}), 400

        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': f'Tylko pliki PDF (twój: {file.filename})'}), 400

        imports_dir = Path(CONFIG['imports_dir'])
        imports_dir.mkdir(parents=True, exist_ok=True)
        filepath = imports_dir / file.filename
        print(f"💾 Zapisuję: {filepath.name}")
        file.save(str(filepath))

        binder_name = request.form.get('binder_name', 'Default')
        print(f"📁 Segregator: {binder_name}")
        print(f"⚙️  Rozpoczynam auto-process (OCR + ekstrakcja)...")

        # Pełen auto-process pipeline
        result = auto_process_pdf(filepath, binder_name)

        if not result:
            return jsonify({'error': 'Błąd odczytywania PDF'}), 400

        print(f"✅ Sukces: {result['filename']} - {result['pages']} stron")
        print("="*60 + "\n")
        return jsonify(result)

    except Exception as e:
        # Loguj pełny traceback dla debugowania
        tb = traceback.format_exc()
        print(f"\n❌ BŁĄD w upload_pdf: {type(e).__name__}: {e}")
        print(f"Traceback:\n{tb}")
        print("="*60 + "\n")
        return jsonify({'error': f'{type(e).__name__}: {str(e)}', 'traceback': tb}), 500

@app.route('/api/land-registers', methods=['GET'])
def get_land_registers():
    """Pobierz wszystkie Księgi Wieczyste"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        kws = db.get_all_land_registers()
        db.disconnect()

        result = []
        for kw in kws:
            result.append({
                'id': kw['id'],
                'kw_full': kw['kw_full'],
                'district': kw['kw_district'],
                'kw_number': kw['kw_number'],
                'checksum': kw['kw_checksum'],
                'address': kw['property_address'] or '—',
                'owner': kw['owner_manual'] or '—',
                'files': kw['files_count'],
                'pages': kw['pages_count'],
                'geoportal_url': kw['geoportal_url'] or '',
                'portal_url': kw['geoportal_url'] or build_geoportal_link(kw['property_address'] or '')
            })

        return jsonify({'success': True, 'data': result})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/land-registers/<int:kw_id>', methods=['POST'])
def update_land_register(kw_id):
    """Zaktualizuj ręczne dane księgi wieczystej."""
    try:
        data = request.json or {}
        address = (data.get('address') or '').strip()
        owner = (data.get('owner') or '').strip()
        geoportal_url = normalize_optional_url(data.get('geoportal_url'))

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        existing = db.fetch_one("SELECT id FROM land_registers WHERE id = ?", (kw_id,))
        if not existing:
            db.disconnect()
            return jsonify({'error': 'Nie znaleziono księgi wieczystej'}), 404

        db.execute(
            """
            UPDATE land_registers
            SET property_address = ?, owner_manual = ?, geoportal_url = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (address, owner, geoportal_url, kw_id)
        )

        db.disconnect()

        return jsonify({
            'success': True,
            'id': kw_id,
            'address': address,
            'owner': owner,
            'geoportal_url': geoportal_url,
            'portal_url': geoportal_url or build_geoportal_link(address)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    """Pobierz statystyki"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        stats = db.get_statistics()
        db.disconnect()

        return jsonify({'success': True, 'data': stats})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/search', methods=['POST'])
def search():
    """Szukaj KW"""
    try:
        query = request.json.get('query', '').strip().upper()

        if not query:
            return jsonify({'error': 'Wpisz szukaną frazę'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Szukaj KW
        kws = db.fetch_all(
            "SELECT * FROM land_registers WHERE kw_full LIKE ?",
            (f"%{query}%",)
        )

        results = []
        for kw in kws:
            results.append({
                'kw_full': kw['kw_full'],
                'district': kw['kw_district'],
                'address': kw['property_address'] or '—',
                'owner': kw['owner_manual'] or '—'
            })

        db.disconnect()

        return jsonify({'success': True, 'data': results, 'count': len(results)})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/update-owner', methods=['POST'])
def update_owner():
    """Zaktualizuj właściciela KW"""
    try:
        kw_id = request.json.get('kw_id')
        owner = request.json.get('owner', '')

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        db.execute(
            "UPDATE land_registers SET owner_manual = ? WHERE id = ?",
            (owner, kw_id)
        )

        db.disconnect()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auto-tag-documents', methods=['POST'])
def auto_tag_documents():
    """Auto-tagowanie wszystkich dokumentów"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        tagger = DocumentTagger()

        # Pobierz wszystkie pliki
        files = db.fetch_all("SELECT id, filename FROM source_files")
        total_tagged = 0
        tags_count = {}

        for file in files:
            # Pobierz wszystkie strony pliku
            pages = db.fetch_all(
                "SELECT text_content FROM pages WHERE file_id = ? AND LENGTH(text_content) > 10",
                (file['id'],)
            )

            if not pages:
                continue

            texts = [p['text_content'] for p in pages]
            tags = tagger.tag_pages_combined(texts)

            # Usuń stare tagi
            db.execute("DELETE FROM document_tags WHERE file_id = ?", (file['id'],))

            # Zapisz nowe tagi (top 3)
            for tag in tags[:3]:
                db.execute("""
                    INSERT OR IGNORE INTO document_tags (file_id, tag, confidence)
                    VALUES (?, ?, ?)
                """, (file['id'], tag['type'], tag['confidence']))

                tags_count[tag['type']] = tags_count.get(tag['type'], 0) + 1
                total_tagged += 1

        db.disconnect()

        return jsonify({
            'success': True,
            'files_processed': len(files),
            'tags_added': total_tagged,
            'tags_distribution': tags_count
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/document-tags', methods=['GET'])
def get_document_tags():
    """Pobierz tagi dokumentów"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        result = db.fetch_all("""
            SELECT
                f.id, f.filename,
                GROUP_CONCAT(dt.tag, ', ') as tags,
                MAX(dt.confidence) as max_confidence
            FROM source_files f
            LEFT JOIN document_tags dt ON f.id = dt.file_id
            GROUP BY f.id
            ORDER BY max_confidence DESC NULLS LAST
        """)

        db.disconnect()

        return jsonify({
            'success': True,
            'data': [{
                'file_id': r['id'],
                'filename': r['filename'],
                'tags': r['tags'] or 'brak',
                'confidence': r['max_confidence'] or 0
            } for r in result]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/generate-search-report', methods=['POST'])
def generate_search_report():
    """Generuj raport z wyników wyszukiwania"""
    try:
        query = request.json.get('query', '').strip()
        if not query:
            return jsonify({'error': 'Wpisz szukaną frazę'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Szukaj stron z wynikami
        results = db.fetch_all("""
            SELECT DISTINCT
                f.id, f.filename, b.name as binder, f.page_count,
                COUNT(DISTINCT p.id) as matching_pages,
                GROUP_CONCAT(DISTINCT dt.tag, ', ') as tags
            FROM pages p
            LEFT JOIN source_files f ON p.file_id = f.id
            LEFT JOIN binders b ON f.binder_id = b.id
            LEFT JOIN document_tags dt ON f.id = dt.file_id
            WHERE p.text_content LIKE ?
            GROUP BY f.id
            ORDER BY matching_pages DESC
        """, (f"%{query}%",))

        # Szukaj encji w wynikach
        entities_found = db.fetch_all("""
            SELECT DISTINCT e.entity_type, COUNT(*) as count
            FROM pages p
            LEFT JOIN entity_occurrences eo ON p.id = eo.page_id
            LEFT JOIN entities e ON eo.entity_id = e.id
            WHERE p.text_content LIKE ? AND e.id IS NOT NULL
            GROUP BY e.entity_type
            ORDER BY count DESC
        """, (f"%{query}%",))

        # Szukaj KW w wynikach
        kw_found = db.fetch_all("""
            SELECT DISTINCT lr.kw_full, COUNT(*) as occurrences
            FROM pages p
            LEFT JOIN land_register_occurrences lro ON p.id = lro.page_id
            LEFT JOIN land_registers lr ON lro.kw_id = lr.id
            WHERE p.text_content LIKE ? AND lr.id IS NOT NULL
            GROUP BY lr.id
            ORDER BY occurrences DESC
            LIMIT 20
        """, (f"%{query}%",))

        db.disconnect()

        # Tworzenie raportu w Excel
        wb = openpyxl.Workbook()

        # Sheet 1: Podsumowanie
        ws = wb.active
        ws.title = "Podsumowanie"
        ws['A1'] = "Raport z Wyszukiwania"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:D1')

        ws['A3'] = "Fraza wyszukiwania:"
        ws['B3'] = query
        ws['A4'] = "Dokumenty znalezione:"
        ws['B4'] = len(results)
        ws['A5'] = "Encje znalezione:"
        ws['B5'] = sum([r['count'] for r in entities_found])
        ws['A6'] = "Księgi Wieczyste znalezione:"
        ws['B6'] = len(kw_found)
        ws['A7'] = "Data raportu:"
        ws['B7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Sheet 2: Dokumenty
        ws2 = wb.create_sheet("Dokumenty")
        ws2.append(['Plik', 'Segregator', 'Strony', 'Dopasowane', 'Tagi'])
        for r in results:
            ws2.append([r['filename'], r['binder'] or '', r['page_count'], r['matching_pages'], r['tags'] or ''])

        # Sheet 3: Encje
        ws3 = wb.create_sheet("Encje")
        ws3.append(['Typ encji', 'Liczba znalezionych'])
        for e in entities_found:
            ws3.append([e['entity_type'], e['count']])

        # Sheet 4: Księgi Wieczyste
        ws4 = wb.create_sheet("Księgi Wieczyste")
        ws4.append(['KW', 'Wystąpienia'])
        for kw in kw_found:
            ws4.append([kw['kw_full'], kw['occurrences']])

        # Styling
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

        exports_dir = Path(CONFIG['exports_dir'])
        exports_dir.mkdir(parents=True, exist_ok=True)
        filename = f"Raport_Szukania_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = exports_dir / filename

        wb.save(str(filepath))

        return send_file(str(filepath), as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/smart-search', methods=['POST'])
def smart_search():
    """
    Smart Search z rankingiem - implementuje priorytety ze specyfikacji:
    1. Wszystkie frazy na jednej stronie (najwyższa)
    2. Wszystkie frazy w jednym dokumencie
    3. Częściowe dopasowania
    """
    try:
        query = request.json.get('query', '').strip()
        if not query:
            return jsonify({'error': 'Wpisz szukaną frazę'}), 400

        # Rozdziel frazy
        phrases = [p.strip() for p in query.split() if p.strip()]

        if not phrases:
            return jsonify({'error': 'Brak fraz'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Szukaj stron które zawierają WSZYSTKIE frazy (priorytet 1)
        results_priority_1 = []
        results_priority_2 = []
        results_priority_3 = []

        # Pobierz wszystkie strony i sprawdź
        pages = db.fetch_all("""
            SELECT p.id, p.page_number, p.text_content, p.file_id,
                   f.filename, b.name as binder_name
            FROM pages p
            LEFT JOIN source_files f ON p.file_id = f.id
            LEFT JOIN binders b ON f.binder_id = b.id
            WHERE LENGTH(p.text_content) > 10
        """)

        # Grupowanie po plikach
        files_with_phrases = {}

        for page in pages:
            text = page['text_content'].upper() if page['text_content'] else ''

            # Sprawdź ile fraz znaleziono na stronie
            found_phrases = [p for p in phrases if p.upper() in text]

            if len(found_phrases) == len(phrases):
                # Priorytet 1 - wszystkie na jednej stronie
                # Wyciągnij kontekst z pierwszego pasującego
                idx = text.find(phrases[0].upper())
                start = max(0, idx - 50)
                end = min(len(page['text_content']), idx + 100)
                context = page['text_content'][start:end]

                results_priority_1.append({
                    'priority': 1,
                    'score': 100,
                    'filename': page['filename'],
                    'binder': page['binder_name'] or '',
                    'page': page['page_number'],
                    'context': context,
                    'phrases_found': found_phrases,
                    'why': 'Wszystkie frazy na tej stronie'
                })
            elif len(found_phrases) > 0:
                # Częściowe dopasowanie
                file_key = page['file_id']
                if file_key not in files_with_phrases:
                    files_with_phrases[file_key] = {
                        'filename': page['filename'],
                        'binder': page['binder_name'] or '',
                        'phrases': set(),
                        'pages': []
                    }
                files_with_phrases[file_key]['phrases'].update(found_phrases)
                files_with_phrases[file_key]['pages'].append({
                    'page': page['page_number'],
                    'phrases': found_phrases,
                    'page_id': page['id']
                })

        # Priorytet 2 - wszystkie frazy w jednym dokumencie
        for file_id, data in files_with_phrases.items():
            if len(data['phrases']) == len(phrases):
                # Wszystkie frazy są w tym dokumencie
                results_priority_2.append({
                    'priority': 2,
                    'score': 70,
                    'filename': data['filename'],
                    'binder': data['binder'],
                    'pages': [p['page'] for p in data['pages']],
                    'phrases_found': list(data['phrases']),
                    'why': f'Wszystkie frazy w jednym dokumencie ({len(data["pages"])} stron)'
                })

        # Priorytet 3 - częściowe dopasowanie
        for file_id, data in files_with_phrases.items():
            if 0 < len(data['phrases']) < len(phrases):
                results_priority_3.append({
                    'priority': 3,
                    'score': 30,
                    'filename': data['filename'],
                    'binder': data['binder'],
                    'pages': [p['page'] for p in data['pages'][:5]],
                    'phrases_found': list(data['phrases']),
                    'why': f'Częściowe dopasowanie ({len(data["phrases"])}/{len(phrases)} fraz)'
                })

        db.disconnect()

        # Połącz wyniki w jednej liście (limit 50)
        all_results = (
            results_priority_1[:20] +
            results_priority_2[:15] +
            results_priority_3[:15]
        )

        return jsonify({
            'success': True,
            'query': query,
            'phrases': phrases,
            'data': all_results,
            'count': len(all_results),
            'priority_1_count': len(results_priority_1),
            'priority_2_count': len(results_priority_2),
            'priority_3_count': len(results_priority_3)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/build-cooccurrences', methods=['POST'])
def build_cooccurrences():
    """Buduj indeks współwystąpień encji"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Wyczyść stare
        db.execute("DELETE FROM cooccurrences")

        # Pobierz strony z wieloma encjami
        pages_with_entities = db.fetch_all("""
            SELECT page_id, GROUP_CONCAT(entity_id) as entity_ids
            FROM entity_occurrences
            GROUP BY page_id
            HAVING COUNT(*) > 1
        """)

        cooccurrence_count = 0

        for page_data in pages_with_entities:
            page_id = page_data['page_id']
            entity_ids = [int(x) for x in page_data['entity_ids'].split(',')]

            # Wszystkie pary encji
            for i in range(len(entity_ids)):
                for j in range(i+1, len(entity_ids)):
                    e1, e2 = sorted([entity_ids[i], entity_ids[j]])
                    db.execute("""
                        INSERT OR IGNORE INTO cooccurrences (entity_id_1, entity_id_2, page_id)
                        VALUES (?, ?, ?)
                    """, (e1, e2, page_id))
                    cooccurrence_count += 1

        db.disconnect()

        return jsonify({
            'success': True,
            'cooccurrences_built': cooccurrence_count,
            'pages_analyzed': len(pages_with_entities)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/find-related', methods=['POST'])
def find_related():
    """Znajdź encje powiązane z daną encją"""
    try:
        entity_value = request.json.get('entity_value', '').strip()

        if not entity_value:
            return jsonify({'error': 'Podaj wartość'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Znajdź encję
        entity = db.fetch_one("""
            SELECT id, entity_type FROM entities
            WHERE entity_value LIKE ? LIMIT 1
        """, (f"%{entity_value}%",))

        if not entity:
            db.disconnect()
            return jsonify({'success': True, 'data': [], 'count': 0})

        # Znajdź powiązane
        related = db.fetch_all("""
            SELECT e.entity_value, e.entity_type, COUNT(*) as cnt
            FROM cooccurrences c
            JOIN entities e ON (e.id = c.entity_id_1 OR e.id = c.entity_id_2)
            WHERE (c.entity_id_1 = ? OR c.entity_id_2 = ?) AND e.id != ?
            GROUP BY e.id
            ORDER BY cnt DESC
            LIMIT 50
        """, (entity['id'], entity['id'], entity['id']))

        db.disconnect()

        return jsonify({
            'success': True,
            'entity': entity_value,
            'data': [{
                'value': r['entity_value'],
                'type': r['entity_type'],
                'cooccurrences': r['cnt']
            } for r in related],
            'count': len(related)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/extracted-names', methods=['GET'])
def get_extracted_names():
    """Pobierz wszystkie wyciagniete imiona/nazwiska z walidacja"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz wszystkie osoby z dokumentów + meta
        persons = db.fetch_all("""
            SELECT
                e.id,
                e.entity_value,
                e.validated,
                COUNT(DISTINCT eo.page_id) as occurrences,
                COUNT(DISTINCT eo.file_id) as files_count
            FROM entities e
            LEFT JOIN entity_occurrences eo ON e.id = eo.entity_id
            WHERE e.entity_type = 'person'
            GROUP BY e.id
            ORDER BY occurrences DESC, e.entity_value
            LIMIT 1000
        """)

        db.disconnect()

        # Imiona ze słownika
        extractor = EntityExtractor()
        all_names = extractor.POLISH_NAMES_MALE | extractor.POLISH_NAMES_FEMALE
        all_surnames = extractor.POLISH_SURNAMES

        result = []
        for p in persons:
            parts = p['entity_value'].split(' ', 1)
            first = parts[0] if parts else ''
            last = parts[1] if len(parts) > 1 else ''

            in_first_dict = first in all_names
            in_surname_dict = last in all_surnames or extractor._is_surname_form(last)

            # Status walidacji
            if p['validated'] == 1:
                status = 'valid'
            elif p['validated'] == -1:
                status = 'invalid'
            elif in_first_dict and in_surname_dict:
                status = 'likely_valid'
            elif in_first_dict:
                status = 'partial'
            else:
                status = 'unknown'

            result.append({
                'id': p['id'],
                'value': p['entity_value'],
                'first_name': first,
                'last_name': last,
                'occurrences': p['occurrences'],
                'files': p['files_count'],
                'in_dict': in_first_dict,
                'in_surname_dict': in_surname_dict,
                'status': status,
                'validated': p['validated'] or 0
            })

        return jsonify({'success': True, 'data': result, 'count': len(result)})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/validate-name', methods=['POST'])
def validate_name():
    """Oznacz nazwisko jako poprawne lub niepoprawne"""
    try:
        entity_id = request.json.get('entity_id')
        action = request.json.get('action', 'valid')  # valid, invalid, delete

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        if action == 'delete':
            # Usuń całkowicie
            db.execute("DELETE FROM entity_occurrences WHERE entity_id = ?", (entity_id,))
            db.execute("DELETE FROM cooccurrences WHERE entity_id_1 = ? OR entity_id_2 = ?",
                       (entity_id, entity_id))
            db.execute("DELETE FROM entities WHERE id = ?", (entity_id,))
            db.disconnect()
            return jsonify({'success': True, 'action': 'deleted'})

        # Set validated flag
        validated_val = 1 if action == 'valid' else -1
        db.execute("UPDATE entities SET validated = ? WHERE id = ?", (validated_val, entity_id))

        db.disconnect()

        return jsonify({'success': True, 'action': action})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/edit-entity', methods=['POST'])
def edit_entity():
    """Edytuj wartość encji i propaguj do OCR wszystkich stron"""
    try:
        entity_id = request.json.get('entity_id')
        new_value = request.json.get('new_value', '').strip()

        if not new_value:
            return jsonify({'error': 'Podaj nową wartość'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz starą wartość
        entity = db.fetch_one("SELECT entity_value, entity_type FROM entities WHERE id = ?", (entity_id,))
        if not entity:
            db.disconnect()
            return jsonify({'error': 'Nie znaleziono encji'}), 404

        old_value = entity['entity_value']

        # Aktualizuj encje
        db.execute(
            "UPDATE entities SET entity_value = ?, normalized_value = ?, validated = 1 WHERE id = ?",
            (new_value, new_value.upper(), entity_id)
        )

        # Propaguj do OCR - zaktualizuj wszystkie strony gdzie ta encja wystepowala
        pages_updated = 0
        page_ids = db.fetch_all(
            "SELECT DISTINCT page_id FROM entity_occurrences WHERE entity_id = ?",
            (entity_id,)
        )

        for pid_row in page_ids:
            pid = pid_row['page_id']
            page = db.fetch_one("SELECT text_content FROM pages WHERE id = ?", (pid,))
            if page and page['text_content'] and old_value in page['text_content']:
                new_text = page['text_content'].replace(old_value, new_value)
                db.execute("UPDATE pages SET text_content = ? WHERE id = ?", (new_text, pid))
                # Zapisz korekte
                db.execute("""
                    INSERT INTO ocr_corrections (page_id, original_text, corrected_text)
                    VALUES (?, ?, ?)
                """, (pid, old_value, new_value))
                pages_updated += 1

        # Zapamietaj korekte w user_dictionary aby nie pytac ponownie
        db.execute("""
            INSERT OR IGNORE INTO user_dictionaries (original, replacement)
            VALUES (?, ?)
        """, (old_value, new_value))

        db.disconnect()

        return jsonify({
            'success': True,
            'old_value': old_value,
            'new_value': new_value,
            'pages_updated': pages_updated
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/delete-file', methods=['POST'])
def delete_file():
    """Usun dokument z bazy (i opcjonalnie plik z dysku)"""
    try:
        file_id = request.json.get('file_id')
        hard_delete = request.json.get('hard_delete', False)

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        file_info = db.fetch_one("SELECT filepath, filename FROM source_files WHERE id = ?", (file_id,))
        if not file_info:
            db.disconnect()
            return jsonify({'error': 'Nie znaleziono pliku'}), 404

        _delete_file_related_data(db, file_id)

        db.disconnect()

        # Hard delete - usun plik z dysku
        if hard_delete and file_info['filepath']:
            try:
                Path(file_info['filepath']).unlink(missing_ok=True)
            except:
                pass

        return jsonify({'success': True, 'filename': file_info['filename']})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/person-documents/<int:entity_id>', methods=['GET'])
def get_person_documents(entity_id):
    """Pobierz wszystkie dokumenty gdzie osoba wystepuje"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        entity = db.fetch_one("SELECT entity_value, entity_type FROM entities WHERE id = ?", (entity_id,))
        if not entity:
            db.disconnect()
            return jsonify({'error': 'Nie znaleziono'}), 404

        documents = db.fetch_all("""
            SELECT DISTINCT
                f.id, f.filename, b.name as binder,
                COUNT(DISTINCT eo.page_id) as pages_count,
                GROUP_CONCAT(DISTINCT eo.page_id) as page_ids
            FROM entity_occurrences eo
            LEFT JOIN source_files f ON eo.file_id = f.id
            LEFT JOIN binders b ON f.binder_id = b.id
            WHERE eo.entity_id = ?
            GROUP BY f.id
            ORDER BY pages_count DESC
        """, (entity_id,))

        db.disconnect()

        return jsonify({
            'success': True,
            'entity': entity['entity_value'],
            'data': [{
                'file_id': d['id'],
                'filename': d['filename'],
                'binder': d['binder'],
                'pages_count': d['pages_count']
            } for d in documents],
            'count': len(documents)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/entity-relations/<entity_type>/<entity_value>', methods=['GET'])
def get_entity_relations(entity_type, entity_value):
    """Pobierz szczegółowe relacje encji z metadata"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Znajdź encję
        entity = db.fetch_one("""
            SELECT id, entity_value, entity_type FROM entities
            WHERE entity_type = ? AND entity_value LIKE ? LIMIT 1
        """, (entity_type, f"%{entity_value}%"))

        if not entity:
            db.disconnect()
            return jsonify({'success': True, 'entity': None, 'data': [], 'count': 0})

        # Znajdź powiązane z metadata
        related = db.fetch_all("""
            SELECT
                e.id,
                e.entity_value,
                e.entity_type,
                COUNT(*) as strength,
                COUNT(DISTINCT p.id) as pages_count,
                COUNT(DISTINCT p.file_id) as files_count,
                GROUP_CONCAT(DISTINCT p.page_number) as pages
            FROM cooccurrences c
            JOIN entities e ON (e.id = c.entity_id_1 OR e.id = c.entity_id_2)
            LEFT JOIN pages p ON p.id = c.page_id
            WHERE (c.entity_id_1 = ? OR c.entity_id_2 = ?) AND e.id != ?
            GROUP BY e.id
            ORDER BY strength DESC
            LIMIT 100
        """, (entity['id'], entity['id'], entity['id']))

        db.disconnect()

        result = [{
            'id': r['id'],
            'value': r['entity_value'],
            'type': r['entity_type'],
            'strength': r['strength'],
            'pages': r['pages_count'],
            'files': r['files_count'],
            'page_numbers': r['pages'].split(',') if r['pages'] else []
        } for r in related]

        return jsonify({
            'success': True,
            'entity': {
                'value': entity['entity_value'],
                'type': entity['entity_type']
            },
            'data': result,
            'count': len(result)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/extract-entities', methods=['POST'])
def extract_entities():
    """Rozpoznaj encje we wszystkich dokumentach"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        extractor = EntityExtractor()

        # Pobierz wszystkie strony z tekstem
        pages = db.fetch_all("""
            SELECT id, file_id, page_number, text_content
            FROM pages
            WHERE LENGTH(text_content) > 10
        """)

        total_entities = {
            'persons': 0, 'companies': 0, 'institutions': 0,
            'nips': 0, 'regons': 0, 'krs': 0, 'pesels': 0,
            'phones': 0, 'emails': 0, 'dates': 0,
            'signatures': 0, 'postal_codes': 0, 'amounts': 0
        }

        for page in pages:
            text = page['text_content']
            if not text:
                continue

            entities = extractor.extract_all(text)

            # Zapisz osoby
            for person in entities['persons']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('person', person['full_name'], person['full_name'].upper()))

                entity_row = db.fetch_one("SELECT id FROM entities WHERE entity_type='person' AND entity_value=?", (person['full_name'],))
                if entity_row:
                    db.execute("""
                        INSERT INTO entity_occurrences (entity_id, page_id, file_id)
                        VALUES (?, ?, ?)
                    """, (entity_row['id'], page['id'], page['file_id']))
                    total_entities['persons'] += 1

            # Zapisz firmy
            for company in entities['companies']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('company', company['name'], company['name'].upper()))
                total_entities['companies'] += 1

            # Zapisz instytucje
            for inst in entities['institutions']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('institution', inst['name'], inst['name'].upper()))
                total_entities['institutions'] += 1

            # NIP
            for nip in entities['nips']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('nip', nip['formatted'], nip['nip']))
                total_entities['nips'] += 1

            # REGON
            for regon in entities['regons']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('regon', regon['regon'], regon['regon']))
                total_entities['regons'] += 1

            # KRS
            for krs in entities['krs']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('krs', krs['krs'], krs['krs']))
                total_entities['krs'] += 1

            # PESEL
            for pesel in entities['pesels']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('pesel', pesel['pesel'], pesel['pesel']))
                total_entities['pesels'] += 1

            # Phones
            for phone in entities['phones']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('phone', phone['phone'], phone['phone']))
                total_entities['phones'] += 1

            # Emails
            for email in entities['emails']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('email', email['email'], email['email']))
                total_entities['emails'] += 1

            # Signatures
            for sig in entities['signatures']:
                db.execute("""
                    INSERT OR IGNORE INTO entities (entity_type, entity_value, normalized_value)
                    VALUES (?, ?, ?)
                """, ('signature', sig['value'], sig['value'].upper()))
                total_entities['signatures'] += 1

        db.disconnect()

        return jsonify({
            'success': True,
            'data': total_entities,
            'pages_processed': len(pages)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entities/<entity_type>', methods=['GET'])
def get_entities(entity_type):
    """Pobierz encje danego typu"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        entities = db.fetch_all("""
            SELECT
                e.id,
                e.entity_value,
                COUNT(DISTINCT eo.page_id) as occurrences,
                COUNT(DISTINCT eo.file_id) as files_count
            FROM entities e
            LEFT JOIN entity_occurrences eo ON e.id = eo.entity_id
            WHERE e.entity_type = ?
            GROUP BY e.id
            ORDER BY occurrences DESC, e.entity_value
            LIMIT 500
        """, (entity_type,))

        db.disconnect()

        result = [{
            'id': e['id'],
            'value': e['entity_value'],
            'occurrences': e['occurrences'] or 0,
            'files': e['files_count'] or 0
        } for e in entities]

        return jsonify({'success': True, 'data': result, 'count': len(result)})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/entity-stats', methods=['GET'])
def get_entity_stats():
    """Statystyki encji"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        stats = db.fetch_all("""
            SELECT entity_type, COUNT(*) as count
            FROM entities
            GROUP BY entity_type
            ORDER BY count DESC
        """)

        db.disconnect()

        result = {row['entity_type']: row['count'] for row in stats}
        return jsonify({'success': True, 'data': result})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/user-dictionary', methods=['GET'])
def get_user_dictionary():
    """Pobierz słownik użytkownika"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        words = db.fetch_all("SELECT id, original, replacement FROM user_dictionaries ORDER BY original")
        db.disconnect()

        result = [{'id': w['id'], 'original': w['original'], 'replacement': w['replacement']} for w in words]

        return jsonify({'success': True, 'data': result})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/add-dictionary-word', methods=['POST'])
def add_dictionary_word():
    """Dodaj słowo do słownika"""
    try:
        original = request.json.get('original', '').strip()
        replacement = request.json.get('replacement', '').strip()

        if not original or not replacement:
            return jsonify({'error': 'Wypełnij oba pola'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        db.execute(
            "INSERT OR IGNORE INTO user_dictionaries (original, replacement) VALUES (?, ?)",
            (original, replacement)
        )

        db.disconnect()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete-dictionary-word', methods=['POST'])
def delete_dictionary_word():
    """Usuń słowo ze słownika"""
    try:
        word_id = request.json.get('word_id')

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        db.execute("DELETE FROM user_dictionaries WHERE id = ?", (word_id,))

        db.disconnect()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/report-kw', methods=['GET'])
def report_kw():
    """Raport Ksiąg Wieczystych"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        kws = db.fetch_all("""
            SELECT
                lr.kw_full,
                lr.kw_district,
                lr.property_address,
                lr.owner_manual,
                COUNT(DISTINCT lro.file_id) as files_count,
                COUNT(DISTINCT lro.page_id) as pages_count
            FROM land_registers lr
            LEFT JOIN land_register_occurrences lro ON lr.id = lro.kw_id
            GROUP BY lr.id
            ORDER BY lr.kw_district, lr.kw_number
        """)

        db.disconnect()

        result = []
        for kw in kws:
            result.append({
                'kw': kw['kw_full'],
                'district': kw['kw_district'],
                'address': kw['property_address'] or '—',
                'owner': kw['owner_manual'] or '—',
                'files': kw['files_count'],
                'pages': kw['pages_count']
            })

        return jsonify({'success': True, 'data': result})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/report-coverage', methods=['GET'])
def report_coverage():
    """Raport Coverage - ile stron przeskanowano"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        total_pages = db.fetch_one("SELECT COUNT(*) as count FROM pages")['count']
        indexed_pages = db.fetch_one("SELECT COUNT(*) as count FROM pages WHERE LENGTH(text_content) > 50")['count']
        total_files = db.fetch_one("SELECT COUNT(*) as count FROM source_files")['count']
        total_kw = db.fetch_one("SELECT COUNT(*) as count FROM land_registers")['count']

        db.disconnect()

        coverage = round((indexed_pages / total_pages * 100), 2) if total_pages > 0 else 0

        return jsonify({
            'success': True,
            'data': {
                'total_pages': total_pages,
                'indexed_pages': indexed_pages,
                'coverage_percent': coverage,
                'total_files': total_files,
                'total_kw': total_kw
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Eksportuj do Excel"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        kws = db.get_all_land_registers()
        db.disconnect()

        # Utwórz workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Księgi Wieczyste"

        # Nagłówki
        headers = ['Lp.', 'Księga Wieczysta', 'Okręg', 'Adres', 'Właściciel', 'W plikach', 'Na stronach']
        ws.append(headers)

        # Style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dane
        for idx, kw in enumerate(kws, 1):
            ws.append([
                idx,
                kw['kw_full'],
                kw['kw_district'],
                kw['property_address'] or '—',
                kw['owner_manual'] or '—',
                kw['files_count'],
                kw['pages_count']
            ])

        # Szerokość kolumn
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12

        # Zapisz
        exports_dir = Path(CONFIG['exports_dir'])
        exports_dir.mkdir(parents=True, exist_ok=True)

        filename = f"Ksiegi_Wieczyste_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = exports_dir / filename

        wb.save(str(filepath))

        return send_file(str(filepath), as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete-kw', methods=['POST'])
def delete_kw():
    """Usuń KW z bazy"""
    try:
        kw_id = request.json.get('kw_id')

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        db.execute("DELETE FROM land_register_occurrences WHERE kw_id = ?", (kw_id,))
        db.execute("DELETE FROM land_registers WHERE id = ?", (kw_id,))

        db.disconnect()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/files', methods=['GET'])
def get_files():
    """Pobierz listę wszystkich plików PDF"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        files = db.fetch_all("""
            SELECT
                f.id,
                f.filename,
                f.page_count,
                f.section_name,
                b.name as binder_name,
                COUNT(DISTINCT p.id) as pages_in_db,
                GROUP_CONCAT(DISTINCT dt.tag) as tags
            FROM source_files f
            LEFT JOIN binders b ON f.binder_id = b.id
            LEFT JOIN pages p ON f.id = p.file_id
            LEFT JOIN document_tags dt ON f.id = dt.file_id
            GROUP BY f.id
            ORDER BY f.id DESC
        """)
        binders = db.fetch_all("""
            SELECT b.id, b.name, b.description, COUNT(sf.id) as files_count
            FROM binders b
            LEFT JOIN source_files sf ON sf.binder_id = b.id
            GROUP BY b.id
            ORDER BY b.name COLLATE NOCASE
        """)

        result = []
        for file in files:
            result.append({
                'id': file['id'],
                'filename': file['filename'],
                'pages': file['page_count'],
                'binder': file['binder_name'] or 'Unknown',
                'section_name': file['section_name'] or '',
                'pages_indexed': file['pages_in_db'],
                'tags': [tag for tag in (file['tags'] or '').split(',') if tag]
            })

        db.disconnect()
        return jsonify({
            'success': True,
            'data': result,
            'binders': [
                {
                    'id': binder['id'],
                    'name': binder['name'],
                    'description': binder['description'] or '',
                    'files_count': binder['files_count']
                }
                for binder in binders
            ]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/binders', methods=['GET'])
def list_binders():
    """Lista segregatorów z liczbą plików"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT b.id, b.name, b.description, b.created_at,
                (SELECT COUNT(*) FROM source_files WHERE binder_id = b.id) as files_count,
                (SELECT COUNT(*) FROM source_files sf JOIN pages p ON p.file_id = sf.id WHERE sf.binder_id = b.id) as pages_count
            FROM binders b
            ORDER BY b.id ASC
        """)
        db.disconnect()
        return jsonify({
            'success': True,
            'binders': [dict(r) for r in rows]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/binders/<int:binder_id>', methods=['DELETE'])
def delete_binder(binder_id):
    """
    Usuń segregator.
    Tryby (przez query param mode):
      - mode=move&target_binder_id=X - przenieś pliki do innego segregatora
      - mode=delete_files - usuń segregator wraz z wszystkimi plikami
      - mode=keep_files - przenieś pliki do "Default" (utworzy jeśli nie istnieje)
    """
    try:
        mode = request.args.get('mode', 'keep_files')
        target_binder_id = request.args.get('target_binder_id')

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Sprawdź czy segregator istnieje
        binder = db.fetch_one("SELECT * FROM binders WHERE id = ?", (binder_id,))
        if not binder:
            db.disconnect()
            return jsonify({'error': 'Segregator nie znaleziony'}), 404

        # Liczba plików w segregatorze
        files_in = db.fetch_all("SELECT id, filepath, filename FROM source_files WHERE binder_id = ?", (binder_id,))
        files_count = len(files_in)

        if mode == 'delete_files':
            # USUŃ wszystkie pliki tego segregatora
            file_ids = [f['id'] for f in files_in]
            for f in files_in:
                # Usuń plik fizyczny (jeśli istnieje)
                try:
                    if f['filepath'] and Path(f['filepath']).exists():
                        Path(f['filepath']).unlink()
                except Exception as ex:
                    print(f"⚠ Nie można usunąć pliku {f['filepath']}: {ex}")

            if file_ids:
                placeholders = ','.join('?' * len(file_ids))
                # Usuń powiązania w innych tabelach
                db.execute(f"DELETE FROM entity_occurrences WHERE file_id IN ({placeholders})", file_ids)
                db.execute(f"DELETE FROM land_register_occurrences WHERE file_id IN ({placeholders})", file_ids)
                db.execute(f"DELETE FROM pages WHERE file_id IN ({placeholders})", file_ids)
                db.execute(f"DELETE FROM document_annotations WHERE file_id IN ({placeholders})", file_ids)
                db.execute(f"DELETE FROM file_summaries WHERE file_id IN ({placeholders})", file_ids)
                db.execute(f"DELETE FROM file_custom_tags WHERE file_id IN ({placeholders})", file_ids)
                db.execute(f"DELETE FROM document_types WHERE file_id IN ({placeholders})", file_ids)
                db.execute(f"DELETE FROM source_files WHERE id IN ({placeholders})", file_ids)

            db.execute("DELETE FROM binders WHERE id = ?", (binder_id,))
            message = f"Usunięto segregator '{binder['name']}' wraz z {files_count} plikami"

        elif mode == 'move' and target_binder_id:
            # Przenieś do innego segregatora
            target_id = int(target_binder_id)
            target = db.fetch_one("SELECT * FROM binders WHERE id = ?", (target_id,))
            if not target:
                db.disconnect()
                return jsonify({'error': 'Docelowy segregator nie znaleziony'}), 404

            db.execute("UPDATE source_files SET binder_id = ? WHERE binder_id = ?", (target_id, binder_id))
            db.execute("DELETE FROM binders WHERE id = ?", (binder_id,))
            message = f"Usunięto segregator '{binder['name']}', przeniesiono {files_count} plików do '{target['name']}'"

        else:
            # Domyślnie: przenieś do "Default"
            default_id = db.add_binder('Default')
            if default_id == binder_id:
                db.disconnect()
                return jsonify({'error': 'Nie można usunąć segregatora "Default"'}), 400

            db.execute("UPDATE source_files SET binder_id = ? WHERE binder_id = ?", (default_id, binder_id))
            db.execute("DELETE FROM binders WHERE id = ?", (binder_id,))
            message = f"Usunięto segregator '{binder['name']}', przeniesiono {files_count} plików do 'Default'"

        db.disconnect()
        return jsonify({
            'success': True,
            'message': message,
            'deleted_binder': binder['name'],
            'files_affected': files_count
        })

    except Exception as e:
        import traceback
        print(f"❌ delete_binder: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/binders/<int:binder_id>', methods=['PUT'])
def rename_binder(binder_id):
    """Zmień nazwę segregatora"""
    try:
        data = request.json or {}
        new_name = (data.get('name') or '').strip()
        if not new_name:
            return jsonify({'error': 'Pusta nazwa'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Sprawdź konflikt
        existing = db.fetch_one("SELECT id FROM binders WHERE name = ? AND id != ?", (new_name, binder_id))
        if existing:
            db.disconnect()
            return jsonify({'error': f'Segregator o nazwie "{new_name}" już istnieje'}), 400

        db.execute("UPDATE binders SET name = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                   (new_name, binder_id))
        db.disconnect()
        return jsonify({'success': True, 'name': new_name})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/binders', methods=['POST'])
def create_binder():
    """Utwórz pusty segregator."""
    try:
        data = request.json or {}
        name = (data.get('name') or '').strip()
        description = (data.get('description') or '').strip()

        if not name:
            return jsonify({'error': 'Podaj nazwę segregatora'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        binder_id = db.add_binder(name, description)
        db.disconnect()

        return jsonify({
            'success': True,
            'id': binder_id,
            'name': name,
            'description': description
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/move-binder', methods=['POST'])
def move_files_to_binder():
    """Przenieś zaznaczone pliki do wskazanego segregatora."""
    try:
        data = request.json or {}
        file_ids = data.get('file_ids') or []
        binder_name = (data.get('binder_name') or '').strip()

        if not file_ids:
            return jsonify({'error': 'Brak wybranych plików'}), 400
        if not binder_name:
            return jsonify({'error': 'Podaj nazwę segregatora'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        binder_id = db.add_binder(binder_name)

        placeholders = ','.join('?' * len(file_ids))
        db.execute(
            f"UPDATE source_files SET binder_id = ? WHERE id IN ({placeholders})",
            (binder_id, *file_ids)
        )
        db.disconnect()

        return jsonify({
            'success': True,
            'binder_id': binder_id,
            'binder_name': binder_name,
            'updated_count': len(file_ids)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/assign-section', methods=['POST'])
def assign_files_section():
    """Przypisz sekcję do zaznaczonych plików."""
    try:
        data = request.json or {}
        file_ids = data.get('file_ids') or []
        section_name = normalize_section_name(data.get('section_name'))

        if not file_ids:
            return jsonify({'error': 'Brak wybranych plików'}), 400

        placeholders = ','.join('?' * len(file_ids))
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute(
            f"UPDATE source_files SET section_name = ? WHERE id IN ({placeholders})",
            (section_name, *file_ids)
        )
        db.disconnect()

        return jsonify({
            'success': True,
            'updated_count': len(file_ids),
            'section_name': section_name
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/merge', methods=['POST'])
def merge_files():
    """Scal kilka PDF-ów w jeden nowy dokument i dodaj do kolejki."""
    try:
        data = request.json or {}
        file_ids = data.get('file_ids') or []
        output_filename = make_safe_pdf_filename(data.get('output_filename') or f"scalone_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        binder_name = (data.get('binder_name') or 'Default').strip() or 'Default'
        section_name = normalize_section_name(data.get('section_name'))

        if len(file_ids) < 2:
            return jsonify({'error': 'Zaznacz co najmniej 2 pliki do scalenia'}), 400

        try:
            from PyPDF2 import PdfMerger
        except ImportError:
            return jsonify({'error': 'Brak PyPDF2 - nie mogę scalać PDF-ów'}), 500

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        placeholders = ','.join('?' * len(file_ids))
        rows = db.fetch_all(
            f"""SELECT sf.filepath, sf.filename, b.name as binder_name
                FROM source_files sf
                LEFT JOIN binders b ON sf.binder_id = b.id
                WHERE sf.id IN ({placeholders})
                ORDER BY sf.id""",
            tuple(file_ids)
        )
        db.disconnect()

        if len(rows) != len(file_ids):
            return jsonify({'error': 'Nie znaleziono wszystkich wybranych plików'}), 404

        imports_dir = Path(CONFIG['imports_dir'])
        imports_dir.mkdir(parents=True, exist_ok=True)
        output_path = imports_dir / output_filename

        merger = PdfMerger()
        for row in rows:
            merger.append(row['filepath'])
        with output_path.open('wb') as target:
            merger.write(target)
        merger.close()

        background_processor.add_to_queue(
            output_path,
            binder_name=binder_name,
            task_type='full_process',
            priority=4,
            section_name=section_name
        )

        if not background_processor.running:
            background_processor.start()

        return jsonify({
            'success': True,
            'filename': output_filename,
            'file_count': len(file_ids),
            'binder_name': binder_name,
            'section_name': section_name,
            'message': f'Scalono {len(file_ids)} pliki i dodano nowy PDF do kolejki'
        })
    except Exception as e:
        import traceback
        print(f"❌ merge_files error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/split', methods=['POST'])
def split_file():
    """Podziel jeden PDF na mniejsze pliki wg zakresów stron."""
    try:
        data = request.json or {}
        file_id = data.get('file_id')
        ranges_text = data.get('ranges', '')
        section_name = normalize_section_name(data.get('section_name'))

        if not file_id:
            return jsonify({'error': 'Brak file_id'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        file_row = db.fetch_one("""
            SELECT sf.filename, sf.filepath, sf.page_count, COALESCE(b.name, 'Default') as binder_name
            FROM source_files sf
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE sf.id = ?
        """, (file_id,))
        db.disconnect()

        if not file_row:
            return jsonify({'error': 'Plik nie znaleziony'}), 404

        page_ranges = parse_split_ranges(ranges_text, file_row['page_count'])

        try:
            from PyPDF2 import PdfReader, PdfWriter
        except ImportError:
            return jsonify({'error': 'Brak PyPDF2 - nie mogę dzielić PDF-ów'}), 500

        imports_dir = Path(CONFIG['imports_dir'])
        imports_dir.mkdir(parents=True, exist_ok=True)

        reader = PdfReader(file_row['filepath'])
        base_name = Path(file_row['filename']).stem
        created_files = []

        for index, (start, end) in enumerate(page_ranges, start=1):
            writer = PdfWriter()
            for page_idx in range(start - 1, end):
                writer.add_page(reader.pages[page_idx])

            part_filename = make_safe_pdf_filename(f"{base_name}_czesc_{index:02d}_{start}-{end}")
            part_path = imports_dir / part_filename
            with part_path.open('wb') as handle:
                writer.write(handle)

            background_processor.add_to_queue(
                part_path,
                binder_name=file_row['binder_name'],
                task_type='full_process',
                priority=4,
                section_name=section_name
            )
            created_files.append({
                'filename': part_filename,
                'range': f"{start}-{end}"
            })

        if not background_processor.running:
            background_processor.start()

        return jsonify({
            'success': True,
            'created_count': len(created_files),
            'created_files': created_files,
            'section_name': section_name,
            'message': f'Utworzono {len(created_files)} nowych PDF-ów'
        })
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        print(f"❌ split_file error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/cleanup-test-data', methods=['POST'])
def cleanup_test_data():
    """Usuń testowe dane i osierocone wpisy KW."""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        test_file_rows = db.fetch_all("""
            SELECT sf.id
            FROM source_files sf
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE LOWER(COALESCE(b.name, '')) LIKE 'test%%'
               OR LOWER(COALESCE(sf.filename, '')) LIKE 'test%%'
               OR LOWER(COALESCE(sf.filepath, '')) LIKE '/tmp/test%%'
        """)

        removed_files = 0
        with db.transaction():
            for row in test_file_rows:
                _delete_file_related_data(db, row['id'])
                removed_files += 1

            db.execute("DELETE FROM ocr_training_examples")
            db.execute("DELETE FROM binders WHERE LOWER(name) LIKE 'test%%'")
            db.execute("""
                DELETE FROM land_registers
                WHERE id NOT IN (
                    SELECT DISTINCT kw_id FROM land_register_occurrences
                )
            """)
            db.execute("""
                DELETE FROM entities
                WHERE id NOT IN (
                    SELECT DISTINCT entity_id FROM entity_occurrences
                )
            """)

        remaining_kw = db.fetch_one("SELECT COUNT(*) AS cnt FROM land_registers")
        db.disconnect()

        return jsonify({
            'success': True,
            'removed_files': removed_files,
            'remaining_land_registers': remaining_kw['cnt'] if remaining_kw else 0
        })
    except Exception as e:
        import traceback
        print(f"❌ cleanup_test_data error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/search-all', methods=['POST'])
def search_all():
    """Szukaj we wszystkich dokumentach"""
    try:
        query = request.json.get('query', '').strip()

        if not query:
            return jsonify({'error': 'Wpisz szukaną frazę'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Szukaj w tekście stron
        results = db.fetch_all("""
            SELECT
                p.id as page_id,
                p.page_number,
                f.filename,
                b.name as binder_name,
                p.text_content,
                f.id as file_id
            FROM pages p
            LEFT JOIN source_files f ON p.file_id = f.id
            LEFT JOIN binders b ON f.binder_id = b.id
            WHERE p.text_content LIKE ?
            ORDER BY f.id, p.page_number
            LIMIT 100
        """, (f"%{query}%",))

        result_list = []
        for row in results:
            # Wyciągnij kontekst (50 znaków przed i po)
            text = row['text_content'] or ''
            idx = text.upper().find(query.upper())
            if idx != -1:
                start = max(0, idx - 50)
                end = min(len(text), idx + len(query) + 50)
                context = text[start:end]
            else:
                context = text[:100]

            result_list.append({
                'file_id': row['file_id'],
                'page_id': row['page_id'],
                'filename': row['filename'],
                'page': row['page_number'],
                'binder': row['binder_name'] or 'Unknown',
                'context': context.strip()
            })

        db.disconnect()

        return jsonify({
            'success': True,
            'data': result_list,
            'count': len(result_list)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/rescan-ocr', methods=['POST'])
def rescan_ocr():
    """Pełny rescan OCR wszystkich dokumentów"""
    try:
        from modules import PDFReader, KWExtractor

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz wszystkie pliki
        files = db.fetch_all("SELECT id, filepath FROM source_files")

        if not files:
            db.disconnect()
            return jsonify({'error': 'Brak plików do skanowania'}), 400

        reader = PDFReader()
        extractor = KWExtractor()

        total_pages = 0
        total_kw = 0

        for file in files:
            file_id = file['id']
            filepath = file['filepath']

            # Czytaj PDF z OCR
            pdf_data = reader.read_pdf(filepath)
            if not pdf_data:
                continue

            # Usuń stare strony
            db.execute("DELETE FROM pages WHERE file_id = ?", (file_id,))

            # Dodaj nowe strony z OCR
            kw_found = 0
            for page_data in pdf_data['pages']:
                page_num = page_data['page_number']
                text = page_data['text']

                page_id = db.add_page(file_id, page_num, text)
                total_pages += 1

                # Ekstrahuj KW
                kws = extractor.extract_from_text(text, page_num)
                for kw in kws:
                    kw_full = kw['kw_full']
                    kw_id = db.add_land_register(
                        kw_full,
                        kw['kw_district'],
                        kw['kw_number'],
                        kw['kw_checksum']
                    )
                    db.add_land_register_occurrence(
                        kw_id,
                        page_id,
                        file_id,
                        kw['context_before'],
                        kw['context_after']
                    )
                    kw_found += 1
                    total_kw += 1

        db.disconnect()

        return jsonify({
            'success': True,
            'pages_scanned': total_pages,
            'kw_found': total_kw,
            'files': len(files)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/page-by-file-page/<int:file_id>/<int:page_number>', methods=['GET'])
def get_page_by_file_page(file_id, page_number):
    """Pobierz page_id + tekst po file_id i page_number"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        page = db.fetch_one(
            "SELECT id, page_number, text_content FROM pages WHERE file_id = ? AND page_number = ?",
            (file_id, page_number)
        )
        db.disconnect()

        if not page:
            return jsonify({'success': False, 'error': 'Strona nie znaleziona'}), 404

        return jsonify({
            'success': True,
            'page_id': page['id'],
            'page_number': page['page_number'],
            'text': page['text_content'] or ''
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/page-text/<int:page_id>', methods=['GET'])
def get_page_text(page_id):
    """Pobierz tekst strony"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        page = db.fetch_one("SELECT id, page_number, text_content FROM pages WHERE id = ?", (page_id,))
        db.disconnect()

        if not page:
            return jsonify({'error': 'Strona nie znaleziona'}), 404

        return jsonify({
            'success': True,
            'page_id': page['id'],
            'page_number': page['page_number'],
            'text': page['text_content']
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/correct-page-text', methods=['POST'])
def correct_page_text():
    """Zapisz poprawioną wersję tekstu strony"""
    try:
        page_id = request.json.get('page_id')
        corrected_text = request.json.get('text', '')

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Aktualizuj tekst strony
        db.execute(
            "UPDATE pages SET text_content = ? WHERE id = ?",
            (corrected_text, page_id)
        )

        # Dodaj zapis do historii
        db.execute(
            "INSERT INTO ocr_corrections (page_id, original_text, corrected_text) VALUES (?, ?, ?)",
            (page_id, "", corrected_text)
        )

        db.disconnect()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/add-kw', methods=['POST'])
def add_kw():
    """Dodaj KW ręcznie"""
    try:
        kw_full = request.json.get('kw_full', '').upper().strip()
        address = request.json.get('address', '').strip()
        owner = request.json.get('owner', '').strip()

        if not kw_full:
            return jsonify({'error': 'Podaj numer KW'}), 400

        # Validacja formatu KW
        import re
        kw_pattern = r'([A-Z]{2}\d{1,2}[A-Z]{1,2})[/\s\-]?(\d{8})[/\s\-]?(\d)'
        match = re.match(kw_pattern, kw_full.replace(' ', '').replace('-', '/'))

        if not match:
            return jsonify({'error': 'Nieprawidłowy format KW (np. SZ1S/00012345/6)'}), 400

        kw_district = match.group(1)
        kw_number = match.group(2)
        kw_checksum = match.group(3)
        kw_full = f"{kw_district}/{kw_number}/{kw_checksum}"

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Dodaj KW
        db.execute("""
            INSERT OR IGNORE INTO land_registers
            (kw_full, kw_district, kw_number, kw_checksum, property_address, owner_manual)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (kw_full, kw_district, kw_number, kw_checksum, address, owner))

        db.disconnect()

        return jsonify({'success': True, 'kw_full': kw_full})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/backup', methods=['POST'])
def backup_database():
    """Tworzenie backupu bazy danych"""
    try:
        import shutil
        from datetime import datetime

        backup_dir = Path(CONFIG['exports_dir']) / 'backups'
        backup_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = backup_dir / f"backup_{timestamp}.db"

        shutil.copy2(CONFIG['db_path'], str(backup_path))

        return jsonify({
            'success': True,
            'backup_file': str(backup_path),
            'size_mb': round(backup_path.stat().st_size / 1024 / 1024, 2)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/advanced-export', methods=['POST'])
def advanced_export():
    """Zaawansowany eksport z filtrami"""
    try:
        request_data = request.json or {}
        sheets = request_data.get('sheets', ['entities', 'documents'])
        entity_types = request_data.get('entity_types', None)
        date_from = request_data.get('date_from', None)

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Sheet 1: Entities (if requested)
        if 'entities' in sheets:
            ws = wb.create_sheet("Encje")
            ws.append(['ID', 'Wartość', 'Typ', 'Wystąpienia', 'Dokumenty'])

            query = "SELECT e.id, e.entity_value, e.entity_type, COUNT(DISTINCT eo.page_id) as pages, COUNT(DISTINCT eo.file_id) as files FROM entities e LEFT JOIN entity_occurrences eo ON e.id = eo.entity_id"
            conditions = []
            params = []

            if entity_types:
                placeholders = ','.join(['?' for _ in entity_types])
                conditions.append(f"e.entity_type IN ({placeholders})")
                params.extend(entity_types)

            if conditions:
                query += " WHERE " + " AND ".join(conditions)

            query += " GROUP BY e.id ORDER BY pages DESC"

            entities = db.fetch_all(query, tuple(params))
            for e in entities:
                ws.append([e['id'], e['entity_value'], e['entity_type'], e['pages'] or 0, e['files'] or 0])

        # Sheet 2: Documents (if requested)
        if 'documents' in sheets:
            ws = wb.create_sheet("Dokumenty")
            ws.append(['Plik', 'Segregator', 'Strony', 'Tagi', 'Data importu'])

            files = db.fetch_all("""
                SELECT f.id, f.filename, b.name as binder, f.page_count,
                       GROUP_CONCAT(dt.tag, ', ') as tags, f.created_at
                FROM source_files f
                LEFT JOIN binders b ON f.binder_id = b.id
                LEFT JOIN document_tags dt ON f.id = dt.file_id
                GROUP BY f.id
                ORDER BY f.created_at DESC
            """)

            for f in files:
                ws.append([f['filename'], f['binder'] or '', f['page_count'], f['tags'] or '', str(f['created_at'])[:10] if f['created_at'] else ''])

        # Sheet 3: Land Registers (if requested)
        if 'kw' in sheets:
            ws = wb.create_sheet("Księgi Wieczyste")
            ws.append(['KW', 'Okręg', 'Adres', 'Właściciel', 'Pliki', 'Strony'])

            kws = db.get_all_land_registers()
            for kw in kws:
                ws.append([kw['kw_full'], kw['kw_district'], kw['property_address'] or '', kw['owner_manual'] or '', kw['files_count'], kw['pages_count']])

        # Styling all sheets
        header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

        db.disconnect()

        exports_dir = Path(CONFIG['exports_dir'])
        exports_dir.mkdir(parents=True, exist_ok=True)
        filename = f"Zaawansowany_Eksport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = exports_dir / filename

        wb.save(str(filepath))

        return send_file(str(filepath), as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-all-excel', methods=['GET'])
def export_all_excel():
    """Eksport wszystkiego do Excel - wiele zakładek"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        wb = openpyxl.Workbook()

        # Sheet 1: Księgi Wieczyste
        ws_kw = wb.active
        ws_kw.title = "Księgi Wieczyste"
        ws_kw.append(['KW', 'Okręg', 'Adres', 'Właściciel', 'Pliki', 'Strony'])
        kws = db.get_all_land_registers()
        for kw in kws:
            ws_kw.append([
                kw['kw_full'], kw['kw_district'],
                kw['property_address'] or '', kw['owner_manual'] or '',
                kw['files_count'], kw['pages_count']
            ])

        # Sheet 2: Pliki
        ws_files = wb.create_sheet("Pliki PDF")
        ws_files.append(['Plik', 'Segregator', 'Strony', 'Tagi'])
        files = db.fetch_all("""
            SELECT f.filename, b.name as binder, f.page_count,
                   GROUP_CONCAT(dt.tag, ', ') as tags
            FROM source_files f
            LEFT JOIN binders b ON f.binder_id = b.id
            LEFT JOIN document_tags dt ON f.id = dt.file_id
            GROUP BY f.id
        """)
        for f in files:
            ws_files.append([f['filename'], f['binder'] or '', f['page_count'], f['tags'] or ''])

        # Sheet 3: Osoby
        ws_persons = wb.create_sheet("Osoby")
        ws_persons.append(['Imię i nazwisko', 'Wystąpienia', 'Pliki'])
        persons = db.fetch_all("""
            SELECT e.entity_value,
                   COUNT(DISTINCT eo.page_id) as occ,
                   COUNT(DISTINCT eo.file_id) as files
            FROM entities e
            LEFT JOIN entity_occurrences eo ON e.id = eo.entity_id
            WHERE e.entity_type='person'
            GROUP BY e.id ORDER BY occ DESC
        """)
        for p in persons:
            ws_persons.append([p['entity_value'], p['occ'] or 0, p['files'] or 0])

        # Sheet 4: Firmy
        ws_companies = wb.create_sheet("Firmy")
        ws_companies.append(['Firma', 'Wystąpienia'])
        companies = db.fetch_all("""
            SELECT entity_value, COUNT(*) as cnt FROM entities
            WHERE entity_type='company' GROUP BY entity_value
        """)
        for c in companies:
            ws_companies.append([c['entity_value'], c['cnt']])

        # Sheet 5: NIPy
        ws_nips = wb.create_sheet("NIPy")
        ws_nips.append(['NIP'])
        nips = db.fetch_all("SELECT entity_value FROM entities WHERE entity_type='nip'")
        for n in nips:
            ws_nips.append([n['entity_value']])

        # Sheet 6: Telefony
        ws_phones = wb.create_sheet("Telefony")
        ws_phones.append(['Numer'])
        phones = db.fetch_all("SELECT entity_value FROM entities WHERE entity_type='phone'")
        for p in phones:
            ws_phones.append([p['entity_value']])

        # Styling
        header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        db.disconnect()

        # Zapisz
        exports_dir = Path(CONFIG['exports_dir'])
        exports_dir.mkdir(parents=True, exist_ok=True)
        from datetime import datetime
        filename = f"Pelny_Raport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = exports_dir / filename
        wb.save(str(filepath))

        return send_file(str(filepath), as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# PAGE VIEWER ENDPOINTS
@app.route('/api/file/<int:file_id>/pages-summary')
def get_pages_summary(file_id):
    """Pobierz statystyki stron w dokumencie"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        file_data = db.fetch_one(
            "SELECT filename, page_count FROM source_files WHERE id = ?",
            (file_id,)
        )

        if not file_data:
            db.disconnect()
            return jsonify({'error': 'Dokument nie znaleziony'}), 404

        # Sprawdź ile stron ma rzeczywiście tekst
        indexed = db.fetch_one(
            """SELECT COUNT(*) as cnt FROM pages
               WHERE file_id = ? AND text_content IS NOT NULL AND TRIM(text_content) != ''""",
            (file_id,)
        )

        # Sprawdź ile stron ma poprawiony OCR
        fixed = db.fetch_one(
            """SELECT COUNT(*) as cnt FROM pages
               WHERE file_id = ? AND fixed_text IS NOT NULL""",
            (file_id,)
        )

        db.disconnect()

        return jsonify({
            'success': True,
            'filename': file_data['filename'],
            'total_pages': file_data['page_count'],
            'indexed_pages': indexed['cnt'] if indexed else 0,
            'fixed_pages': fixed['cnt'] if fixed else 0
        })
    except Exception as e:
        import traceback
        print(f"❌ pages-summary error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/file/<int:file_id>/page/<int:page_number>')
def get_page_data(file_id, page_number):
    """Pobierz tekst i metadane konkretnej strony"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz informacje o pliku
        file_data = db.fetch_one(
            "SELECT page_count FROM source_files WHERE id = ?",
            (file_id,)
        )

        if not file_data:
            db.disconnect()
            return jsonify({'error': 'Dokument nie znaleziony'}), 404

        if page_number < 1 or page_number > file_data['page_count']:
            db.disconnect()
            return jsonify({'error': 'Strona poza zakresem'}), 400

        # Pobierz tekst strony (preferuj fixed_text jeśli dostępny)
        page_data = db.fetch_one(
            """SELECT id, text_content, fixed_text, ocr_fixed_at
               FROM pages WHERE file_id = ? AND page_number = ?""",
            (file_id, page_number)
        )

        if not page_data:
            db.disconnect()
            return jsonify({'error': 'Strona nie znaleziona'}), 404

        page_id = page_data['id']
        # Użyj fixed_text (poprawione OCR) jeśli istnieje, inaczej oryginalny tekst
        is_fixed = bool(page_data['fixed_text'])
        text_content = page_data['fixed_text'] if is_fixed else page_data['text_content']
        original_text = page_data['text_content']

        if not text_content or not text_content.strip():
            text_content = "Brak tekstu na tej stronie"

        # Pobierz KW na tej stronie
        kw_count = db.fetch_one(
            """SELECT COUNT(DISTINCT lr_id) as cnt FROM land_register_occurrences
               WHERE page_id = ?""",
            (page_id,)
        )

        # Pobierz encje na tej stronie
        entity_count = db.fetch_one(
            """SELECT COUNT(DISTINCT entity_id) as cnt FROM entity_occurrences
               WHERE page_id = ?""",
            (page_id,)
        )

        db.disconnect()

        # Przygotuj metadata
        nav = {
            'prev_page': page_number - 1 if page_number > 1 else None,
            'next_page': page_number + 1 if page_number < file_data['page_count'] else None
        }

        return jsonify({
            'success': True,
            'file_id': file_id,
            'page_id': page_id,
            'page_number': page_number,
            'total_pages': file_data['page_count'],
            'text_content': text_content,
            'original_text': original_text,
            'is_ocr_fixed': is_fixed,
            'ocr_fixed_at': page_data['ocr_fixed_at'] if is_fixed else None,
            'kw_count': kw_count['cnt'] if kw_count else 0,
            'entity_count': entity_count['cnt'] if entity_count else 0,
            'navigation': nav
        })
    except Exception as e:
        import traceback
        print(f"❌ get_page_data error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/file/<int:file_id>/ocr-analysis-status')
def get_ocr_analysis_status(file_id):
    """Pobierz status analizy OCR dokumentu"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz info o pliku
        file_data = db.fetch_one(
            "SELECT page_count FROM source_files WHERE id = ?",
            (file_id,)
        )

        if not file_data:
            db.disconnect()
            return jsonify({'error': 'Dokument nie znaleziony'}), 404

        # Policz poprawione strony
        stats = db.fetch_one("""
            SELECT
                COUNT(*) as total_pages,
                SUM(CASE WHEN fixed_text IS NOT NULL THEN 1 ELSE 0 END) as fixed_pages,
                SUM(CASE WHEN fixed_text IS NULL THEN 1 ELSE 0 END) as remaining_pages,
                MAX(ocr_fixed_at) as last_fixed_at
            FROM pages
            WHERE file_id = ?
        """, (file_id,))

        db.disconnect()

        total = stats['total_pages'] or file_data['page_count']
        fixed = stats['fixed_pages'] or 0
        remaining = stats['remaining_pages'] or 0
        progress = round((fixed / total * 100) if total > 0 else 0, 1)

        return jsonify({
            'success': True,
            'file_id': file_id,
            'total_pages': total,
            'fixed_pages': fixed,
            'remaining_pages': remaining,
            'progress_percent': progress,
            'analysis_complete': fixed == total,
            'last_fixed_at': stats['last_fixed_at']
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# LLAMA / OLLAMA HELPER FUNCTIONS
# ============================================

# Konfiguracja Llama - można zmieniać
LLAMA_CONFIG = {
    'url': 'http://localhost:11434/api/generate',
    'model': 'llama3.2:3b',  # Mniejszy = szybszy. Alternatives: 'llama3.2:1b', 'mistral', 'neural-chat'
    'timeout': 300,  # 5 minut na chunk
    'chunk_size': 2500,  # znaków per chunk
    'temperature': 0.2,
    'top_p': 0.9,
    'max_retries': 2,
    'few_shot_count': 3,  # Ile przykładów nauki użyć (0 = wyłącz)
}

LLAMA_SYSTEM_PROMPT = """Jesteś ekspertem w poprawianiu błędów OCR dla polskich dokumentów.
Pracujesz w systemie wyszukiwania dokumentów (umowy, KW, akty notarialne, faktury, pisma sądowe).

Najczęstsze błędy OCR do poprawienia:
- Pomylenia liter: l/I/1, O/0, rn/m, cl/d, vv/w
- Znaki diakrytyczne: ą,ć,ę,ł,ń,ó,ś,ź,ż (często pomijane przez OCR)
- Dziury w słowach: brakujące litery wewnątrz wyrazów
- Pomieszane spacje: brakujące lub nadmiarowe

WAŻNE:
1. ZACHOWAJ oryginalny sens i strukturę zdań
2. POPRAW błędy OCR do poprawnego polskiego
3. NIE dodawaj nowych informacji
4. NIE komentuj - zwróć TYLKO poprawiony tekst
5. Jeśli słowo jest niezrozumiałe, zostaw je bez zmian"""


def _llama_check_available():
    """Sprawdź czy Ollama jest dostępna"""
    try:
        r = requests.get('http://localhost:11434/api/tags', timeout=3)
        return r.status_code == 200
    except Exception:
        return False


def _llama_get_models():
    """Pobierz listę dostępnych modeli z Ollama"""
    try:
        r = requests.get('http://localhost:11434/api/tags', timeout=5)
        if r.status_code == 200:
            return [m['name'] for m in r.json().get('models', [])]
    except Exception:
        pass
    return []


def _resolve_llama_model(requested_model=None):
    """Wybierz działający model: preferowany lub pierwszy sensowny zainstalowany."""
    requested_model = requested_model or LLAMA_CONFIG['model']
    installed = _llama_get_models()

    if not installed:
        return None, installed

    if requested_model in installed:
        return requested_model, installed

    preferred_order = [
        'llama3.2:3b',
        'llama3.2:1b',
        'phi3:mini',
        'mistral',
        'neural-chat',
    ]

    for candidate in preferred_order:
        if candidate in installed:
            return candidate, installed

    return installed[0], installed


def _get_training_examples(limit=5):
    """Pobierz ostatnie przykłady poprawek do few-shot learning"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        examples = db.fetch_all("""
            SELECT original_text, corrected_text
            FROM ocr_training_examples
            WHERE LENGTH(original_text) < 800 AND LENGTH(corrected_text) < 800
            ORDER BY used_count ASC, created_at DESC
            LIMIT ?
        """, (limit,))

        # Inkrementuj used_count dla wybranych
        if examples:
            db.execute(f"""
                UPDATE ocr_training_examples
                SET used_count = used_count + 1
                WHERE id IN (
                    SELECT id FROM ocr_training_examples
                    WHERE LENGTH(original_text) < 800 AND LENGTH(corrected_text) < 800
                    ORDER BY used_count ASC, created_at DESC
                    LIMIT {int(limit)}
                )
            """)

        db.disconnect()
        return [dict(row) for row in examples]
    except Exception as e:
        print(f"⚠ Błąd pobierania training examples: {e}")
        return []


def _build_few_shot_prompt(examples):
    """Buduj sekcję few-shot z przykładami nauki"""
    if not examples:
        return ""

    parts = ["\n\n=== PRZYKŁADY POPRAWNYCH KOREKCJI (ucz się od nich!) ==="]
    for i, ex in enumerate(examples, 1):
        parts.append(f"""
PRZYKŁAD {i}:
BŁĘDNY OCR: {ex['original_text']}
POPRAWNY TEKST: {ex['corrected_text']}""")
    parts.append("\n=== KONIEC PRZYKŁADÓW ===\n")
    return '\n'.join(parts)


def _llama_correct_chunk(text_chunk, model=None, context="", use_training=True):
    """
    Popraw jeden chunk tekstu przez Llama.
    Zwraca: (success, corrected_text_or_error_msg)
    """
    model, installed_models = _resolve_llama_model(model)
    if not model:
        return False, "Brak zainstalowanych modeli Ollama. Zainstaluj np. `ollama pull llama3.2:1b`"

    # Pobierz few-shot examples z bazy (nauka z poprzednich korekt usera)
    few_shot = ""
    if use_training:
        examples = _get_training_examples(limit=LLAMA_CONFIG.get('few_shot_count', 3))
        if examples:
            few_shot = _build_few_shot_prompt(examples)
            print(f"   🧠 Few-shot: używam {len(examples)} przykładów nauki")

    prompt = f"""{LLAMA_SYSTEM_PROMPT}
{few_shot}
{context}
ORYGINALNY TEKST OCR:
{text_chunk}

POPRAWIONY TEKST (tylko tekst, bez komentarzy):"""

    last_error = None
    for attempt in range(LLAMA_CONFIG['max_retries'] + 1):
        try:
            response = requests.post(
                LLAMA_CONFIG['url'],
                json={
                    'model': model,
                    'prompt': prompt,
                    'stream': False,
                    'options': {
                        'temperature': LLAMA_CONFIG['temperature'],
                        'top_p': LLAMA_CONFIG['top_p'],
                        'num_predict': len(text_chunk) + 500,  # Zostaw miejsce na dłuższy output
                    }
                },
                timeout=LLAMA_CONFIG['timeout']
            )

            if response.status_code == 200:
                result = response.json()
                fixed = result.get('response', text_chunk).strip()
                # Usuń wielokrotne spacje
                fixed = ' '.join(fixed.split())
                return True, fixed
            elif response.status_code == 404:
                if installed_models:
                    return False, f"Model '{model}' jest niedostępny. Dostępne lokalnie: {', '.join(installed_models[:5])}"
                return False, f"Model '{model}' nie zainstalowany. Uruchom: ollama pull {model}"
            else:
                last_error = f"HTTP {response.status_code}: {response.text[:200]}"

        except requests.exceptions.ConnectionError:
            return False, "Ollama nie uruchomiona - uruchom: brew services start ollama"
        except requests.exceptions.Timeout:
            last_error = f"Timeout po {LLAMA_CONFIG['timeout']}s (próba {attempt + 1})"
            if attempt < LLAMA_CONFIG['max_retries']:
                continue
        except Exception as e:
            last_error = f"Błąd: {type(e).__name__}: {str(e)[:200]}"

    return False, last_error or "Nieznany błąd"


def _llama_correct_text(text, context="", model=None):
    """
    Popraw długi tekst dzieląc go na chunki gdy potrzeba.
    Zwraca: (success, corrected_text_or_error)
    """
    if not text or len(text.strip()) < 20:
        return True, text

    # Jeśli tekst jest krótki, jeden chunk
    if len(text) <= LLAMA_CONFIG['chunk_size']:
        return _llama_correct_chunk(text, model=model, context=context)

    # Dla długich tekstów - dzielenie na chunki
    chunks = []
    chunk_size = LLAMA_CONFIG['chunk_size']

    # Dzielenie po zdaniach żeby nie ciąć w środku
    sentences = text.replace('. ', '.|').replace('? ', '?|').replace('! ', '!|').split('|')
    current_chunk = ""

    for sentence in sentences:
        if len(current_chunk) + len(sentence) + 1 < chunk_size:
            current_chunk += " " + sentence if current_chunk else sentence
        else:
            if current_chunk:
                chunks.append(current_chunk.strip())
            current_chunk = sentence

    if current_chunk:
        chunks.append(current_chunk.strip())

    print(f"📦 Tekst podzielony na {len(chunks)} chunków")

    # Popraw każdy chunk
    corrected_chunks = []
    for i, chunk in enumerate(chunks):
        print(f"   🔄 Chunk {i + 1}/{len(chunks)} ({len(chunk)} znaków)...")
        success, result = _llama_correct_chunk(chunk, model=model, context=context)
        if success:
            corrected_chunks.append(result)
        else:
            print(f"   ❌ Chunk {i + 1} failed: {result}")
            corrected_chunks.append(chunk)  # Fallback: oryginalny tekst

    return True, ' '.join(corrected_chunks)


@app.route('/api/llama/status', methods=['GET'])
def llama_status():
    """Sprawdź status Ollama i dostępne modele"""
    available = _llama_check_available()
    models = _llama_get_models() if available else []
    resolved_model, _ = _resolve_llama_model(LLAMA_CONFIG['model']) if available else (None, [])
    return jsonify({
        'available': available,
        'url': LLAMA_CONFIG['url'],
        'configured_model': LLAMA_CONFIG['model'],
        'resolved_model': resolved_model,
        'installed_models': models,
        'current_model_installed': LLAMA_CONFIG['model'] in models,
        'config': {
            'timeout': LLAMA_CONFIG['timeout'],
            'chunk_size': LLAMA_CONFIG['chunk_size']
        }
    })


@app.route('/api/llama/config', methods=['POST'])
def llama_set_config():
    """Zmień konfigurację Llama (model, timeout, few_shot_count)"""
    try:
        data = request.json
        if 'model' in data:
            LLAMA_CONFIG['model'] = data['model']
        if 'timeout' in data:
            LLAMA_CONFIG['timeout'] = int(data['timeout'])
        if 'chunk_size' in data:
            LLAMA_CONFIG['chunk_size'] = int(data['chunk_size'])
        if 'few_shot_count' in data:
            LLAMA_CONFIG['few_shot_count'] = int(data['few_shot_count'])
        return jsonify({'success': True, 'config': LLAMA_CONFIG})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# OCR TRAINING - nauka Llamy z user corrections
# ============================================

@app.route('/api/save-ocr-correction', methods=['POST'])
def save_ocr_correction():
    """
    Zapisz ręczną korekcję OCR jako przykład treningowy dla Llamy.
    User edytuje tekst → llama uczy się typowych błędów OCR
    """
    try:
        data = request.json
        file_id = data.get('file_id')
        page_number = data.get('page_number')
        original_text = data.get('original_text', '').strip()
        corrected_text = data.get('corrected_text', '').strip()

        if not original_text or not corrected_text:
            return jsonify({'error': 'Brak tekstów do zapisu'}), 400

        if original_text == corrected_text:
            return jsonify({
                'success': True,
                'message': 'Tekst niezmieniony - nic do nauki'
            })

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Znajdź page_id
        page_id = None
        if file_id and page_number:
            row = db.fetch_one(
                "SELECT id FROM pages WHERE file_id = ? AND page_number = ?",
                (file_id, page_number)
            )
            if row:
                page_id = row['id']

        # Oblicz różnicę rozmiaru (jak bardzo tekst się różni)
        diff_size = abs(len(original_text) - len(corrected_text))

        # Zapisz jako training example
        db.execute("""
            INSERT INTO ocr_training_examples
            (file_id, page_id, page_number, original_text, corrected_text, diff_size, source)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (file_id, page_id, page_number, original_text, corrected_text, diff_size, 'manual'))

        # Aktualizuj fixed_text na stronie (zapisz aktualną korekcję)
        if page_id:
            db.execute("""
                UPDATE pages
                SET fixed_text = ?,
                    ocr_fixed_at = CURRENT_TIMESTAMP,
                    ocr_confidence = 'manual'
                WHERE id = ?
            """, (corrected_text, page_id))

        # Policz total examples
        count_row = db.fetch_one("SELECT COUNT(*) as cnt FROM ocr_training_examples")
        total = count_row['cnt'] if count_row else 0

        db.disconnect()

        return jsonify({
            'success': True,
            'message': f'Zapisano korekcję jako przykład nauki ({total} w bazie)',
            'total_examples': total,
            'diff_size': diff_size
        })

    except Exception as e:
        import traceback
        print(f"❌ save_ocr_correction error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/training-examples', methods=['GET'])
def get_training_examples():
    """Pobierz listę przykładów treningowych"""
    try:
        limit = int(request.args.get('limit', 50))

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        examples = db.fetch_all("""
            SELECT te.*, sf.filename
            FROM ocr_training_examples te
            LEFT JOIN source_files sf ON te.file_id = sf.id
            ORDER BY te.created_at DESC
            LIMIT ?
        """, (limit,))

        total = db.fetch_one("SELECT COUNT(*) as cnt FROM ocr_training_examples")

        db.disconnect()

        return jsonify({
            'success': True,
            'total': total['cnt'] if total else 0,
            'examples': [{
                'id': e['id'],
                'filename': e['filename'] or 'unknown',
                'page_number': e['page_number'],
                'original_text': e['original_text'][:200] + '...' if len(e['original_text']) > 200 else e['original_text'],
                'corrected_text': e['corrected_text'][:200] + '...' if len(e['corrected_text']) > 200 else e['corrected_text'],
                'diff_size': e['diff_size'],
                'used_count': e['used_count'],
                'source': e['source'],
                'created_at': e['created_at']
            } for e in examples]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/training-example/<int:example_id>', methods=['DELETE'])
def delete_training_example(example_id):
    """Usuń przykład treningowy"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("DELETE FROM ocr_training_examples WHERE id = ?", (example_id,))
        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/fix-all-ocr', methods=['POST'])
def fix_all_ocr():
    """Popraw wszystkie strony dokumentu za pomocą Llama (z chunking i retry)"""
    try:
        # Sprawdź czy Ollama dostępna
        if not _llama_check_available():
            return jsonify({
                'success': False,
                'error': 'Ollama nie uruchomiona. Uruchom: brew services start ollama'
            }), 503

        data = request.json
        file_id = data.get('file_id')

        if not file_id:
            return jsonify({'error': 'Brak file_id'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz wszystkie strony dokumentu które nie mają fixed_text
        pages = db.fetch_all("""
            SELECT p.id, p.page_number, p.text_content, sf.filename
            FROM pages p
            JOIN source_files sf ON p.file_id = sf.id
            WHERE p.file_id = ? AND p.fixed_text IS NULL
            ORDER BY p.page_number
        """, (file_id,))

        db.disconnect()

        if not pages:
            return jsonify({
                'success': True,
                'message': 'Wszystkie strony już są poprawione!',
                'fixed_pages': 0,
                'total_pages': 0
            })

        fixed_count = 0
        failed_count = 0
        skipped_count = 0
        errors = []
        total = len(pages)

        print(f"\n🤖 Llama OCR Fix: {total} stron do poprawienia")
        print(f"   Model: {LLAMA_CONFIG['model']}, Timeout: {LLAMA_CONFIG['timeout']}s")

        for i, page in enumerate(pages):
            text = page['text_content']
            page_num = page['page_number']

            if not text or len(text.strip()) < 20:
                skipped_count += 1
                continue

            print(f"\n📄 [{i + 1}/{total}] Strona {page_num} ({len(text)} znaków)...")

            try:
                context = f"DOKUMENT: {page['filename']}, strona {page_num}"
                success, result_text = _llama_correct_text(text, context=context)

                if success and result_text:
                    db2 = DatabaseManager(CONFIG['db_path'])
                    db2.connect()
                    db2.execute("""
                        UPDATE pages
                        SET fixed_text = ?, ocr_fixed_at = CURRENT_TIMESTAMP, ocr_confidence = ?
                        WHERE id = ?
                    """, (result_text, 'high', page['id']))
                    db2.disconnect()

                    fixed_count += 1
                    print(f"   ✅ Poprawiona ({len(result_text)} znaków)")
                else:
                    failed_count += 1
                    error_msg = f"Strona {page_num}: {result_text}"
                    errors.append(error_msg)
                    print(f"   ❌ {error_msg}")

            except Exception as e:
                failed_count += 1
                error_msg = f"Strona {page_num}: {type(e).__name__}: {str(e)[:100]}"
                errors.append(error_msg)
                print(f"   ❌ {error_msg}")
                continue

        print(f"\n✅ Llama Fix zakończone: {fixed_count} OK, {failed_count} fail, {skipped_count} skip")

        return jsonify({
            'success': True,
            'fixed_pages': fixed_count,
            'failed_pages': failed_count,
            'skipped_pages': skipped_count,
            'total_pages': total,
            'errors': errors[:10] if errors else None,  # Pierwsze 10 błędów
            'message': f'Poprawiono {fixed_count}/{total} stron (fail: {failed_count}, skip: {skipped_count})'
        })

    except Exception as e:
        import traceback
        print(f"❌ fix_all_ocr error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/fix-ocr', methods=['POST'])
def fix_ocr():
    """Popraw OCR tekst za pomocą Llama (z chunking + retry)"""
    try:
        data = request.json
        text = data.get('text', '').strip()
        page_num = data.get('page_number', '?')
        file_id = data.get('file_id')
        filename = data.get('filename', 'dokument')

        if not text:
            return jsonify({'error': 'Pusty tekst'}), 400

        if len(text) < 20:
            return jsonify({'success': True, 'original': text, 'fixed': text})

        # Sprawdź czy Ollama dostępna
        if not _llama_check_available():
            return jsonify({
                'success': False,
                'error': 'Ollama nie uruchomiona. Uruchom: brew services start ollama',
                'original': text
            }), 503

        # Użyj helper function z chunkingiem
        context = f"DOKUMENT: {filename}, strona {page_num}"
        success, result = _llama_correct_text(text, context=context)

        if not success:
            return jsonify({
                'success': False,
                'error': result,
                'original': text
            }), 500

        fixed_text = result

        # Zapisz w bazie
        try:
            if page_num != '?' and file_id:
                db = DatabaseManager(CONFIG['db_path'])
                db.connect()
                db.execute("""
                    UPDATE pages
                    SET fixed_text = ?, ocr_fixed_at = CURRENT_TIMESTAMP, ocr_confidence = ?
                    WHERE file_id = ? AND page_number = ?
                """, (fixed_text, 'high', file_id, page_num))
                db.disconnect()
        except Exception as save_error:
            print(f"⚠ Błąd zapisu do bazy: {save_error}")

        return jsonify({
            'success': True,
            'original': text,
            'fixed': fixed_text,
            'confidence': 'high'
        })

    except Exception as e:
        import traceback
        print(f"❌ fix_ocr error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/search-text', methods=['POST'])
def search_text():
    """Wyszukaj tekst we wszystkich dokumentach (case-insensitive, polskie znaki)"""
    try:
        data = request.json
        original_query = data.get('query', '').strip()
        search_query = original_query.lower()

        if not search_query:
            return jsonify({'error': 'Puste zapytanie'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Wyszukaj w tabelce pages - tekst zawierający query
        # Używamy fixed_text gdy jest dostępny (poprawione OCR), inaczej text_content
        results = db.fetch_all("""
            SELECT
                p.id as page_id,
                p.page_number,
                COALESCE(p.fixed_text, p.text_content) as text_content,
                sf.id as file_id,
                sf.filename,
                b.id as binder_id,
                b.name as binder_name
            FROM pages p
            JOIN source_files sf ON p.file_id = sf.id
            JOIN binders b ON sf.binder_id = b.id
            WHERE LOWER(COALESCE(p.fixed_text, p.text_content)) LIKE ?
            ORDER BY sf.id, p.page_number
        """, (f'%{search_query}%',))

        db.disconnect()

        # Formatuj wyniki
        formatted_results = []
        for row in results:
            text = row['text_content']
            if not text:
                continue

            text_lower = text.lower()
            pos = text_lower.find(search_query)
            if pos == -1:
                continue

            # Context: 100 znaków przed i po dla lepszego zrozumienia
            context_size = 100
            start = max(0, pos - context_size)
            end = min(len(text), pos + len(search_query) + context_size)
            context = text[start:end].strip()

            # Liczba wystąpień (case-insensitive)
            match_count = text_lower.count(search_query)

            formatted_results.append({
                'file_id': row['file_id'],
                'filename': row['filename'],
                'binder_name': row['binder_name'] or 'Default',
                'page_number': row['page_number'],
                'context': context,
                'match_count': match_count
            })

        # Sortuj wyniki - dokumenty z największą liczbą wystąpień najpierw
        # (grupowanie zachowuje kolejność wewnątrz dokumentu)
        file_match_counts = {}
        for r in formatted_results:
            file_match_counts[r['file_id']] = file_match_counts.get(r['file_id'], 0) + r['match_count']

        formatted_results.sort(key=lambda r: (-file_match_counts[r['file_id']], r['file_id'], r['page_number']))

        return jsonify({
            'success': True,
            'query': original_query,
            'total_matches': sum(r['match_count'] for r in formatted_results),
            'total_pages': len(formatted_results),
            'total_documents': len(file_match_counts),
            'results': formatted_results
        })

    except Exception as e:
        import traceback
        print(f"❌ search_text error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/file/<int:file_id>/page/<int:page_number>/pdf')
def get_page_pdf(file_id, page_number):
    """Pobierz PDF konkretnej strony"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        file_data = db.fetch_one(
            "SELECT filepath, page_count FROM source_files WHERE id = ?",
            (file_id,)
        )
        db.disconnect()

        if not file_data:
            return jsonify({'error': 'Dokument nie znaleziony'}), 404

        if page_number < 1 or page_number > file_data['page_count']:
            return jsonify({'error': 'Strona poza zakresem'}), 400

        filepath = file_data['filepath']

        # Import PyPDF2 dla ekstrakcji strony
        try:
            from PyPDF2 import PdfReader, PdfWriter
            from io import BytesIO

            reader = PdfReader(filepath)
            writer = PdfWriter()
            writer.add_page(reader.pages[page_number - 1])

            output = BytesIO()
            writer.write(output)
            output.seek(0)

            return send_file(
                output,
                mimetype='application/pdf',
                as_attachment=False,
                download_name=f'page_{page_number}.pdf'
            )
        except ImportError:
            # Fallback: zwróć całą PDF-ę jeśli PyPDF2 nie zainstalowany
            return send_file(filepath, mimetype='application/pdf')

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# PERSON LINKING - łączenie osób
# ============================================

def _parse_person_name(name):
    """
    Rozdziel imię/imiona/nazwisko.
    "Sławomir Krzysztof Feszczak" → {first: 'Sławomir', middle: ['Krzysztof'], last: 'Feszczak'}
    """
    if not name:
        return None
    parts = name.strip().split()
    if len(parts) < 2:
        return None
    return {
        'first': parts[0],
        'middle': parts[1:-1] if len(parts) > 2 else [],
        'last': parts[-1],
        'full': name,
        'base': f"{parts[0]} {parts[-1]}"  # imię + nazwisko bez middle
    }


def _normalize_polish_name(name):
    """
    Normalizuj polską formę imienia/nazwiska do mianownika.
    "Sławomira" → "Sławomir", "Feszczaka" → "Feszczak", "Kowalskiemu" → "Kowalski"
    """
    if not name:
        return name
    n = name.strip()
    nl = n.lower()

    # Końcówki deklinacji - dłuższe pierwsze!
    endings = [
        ('skiego', 'ski'), ('skiemu', 'ski'), ('skimi', 'scy'), ('skim', 'ski'),
        ('skiej', 'ska'), ('skie', 'ska'), ('ską', 'ska'),
        ('ckiego', 'cki'), ('ckiemu', 'cki'), ('ckimi', 'ccy'), ('ckim', 'cki'),
        ('ckiej', 'cka'), ('ckie', 'cka'), ('cką', 'cka'),
        ('zkiego', 'zki'), ('zkiemu', 'zki'), ('zkim', 'zki'),
        ('zkiej', 'zka'),
        ('iem', ''),  # Marcinem → Marcin
        ('ach', ''),
        ('ami', ''),
    ]

    for ending, replacement in endings:
        if nl.endswith(ending):
            base_lower = nl[:-len(ending)] + replacement
            if len(base_lower) >= 3:
                return name[0].upper() + base_lower[1:]

    # Krótsze końcówki - tylko dla dłuższych słów (>= 5)
    if len(n) >= 5:
        if nl.endswith('owi') and len(n) >= 6:
            # Janowi → Jan
            return name[:-3]
        if nl.endswith('em') and len(n) >= 6:
            return name[:-2]
        if nl.endswith('ie') and not nl.endswith('owie'):
            # Annie → Anna (zachowuje końcowe -a)
            return name[:-2] + 'a'
        if nl.endswith('y'):
            # Anny → Anna
            return name[:-1] + 'a'
        if nl.endswith('ą'):
            return name[:-1] + 'a'
        # "a" na końcu - dla dłuższych słów (5+ liter)
        if nl.endswith('a') and not nl.endswith('ka') and not nl.endswith('na') and not nl.endswith('sa'):
            # Sławomira → Sławomir, Feszczaka → Feszczak
            # ALE: zachowaj Anna, Maria, Joanna (męskie imiona z -a)
            # Heurystyka: jeśli przedostatnia to spółgłoska a samogłoska - usuń -a
            if len(n) >= 5 and n[-2].lower() not in 'aeiouyąęó':
                return name[:-1]

    return name


def _normalize_person_canonical(full_name):
    """
    Znormalizowana forma pełnej nazwy osoby (lower).
    "Sławomira Feszczaka" → "sławomir feszczak"
    """
    parsed = _parse_person_name(full_name)
    if not parsed:
        return None
    first = _normalize_polish_name(parsed['first']) or parsed['first']
    last = _normalize_polish_name(parsed['last']) or parsed['last']
    middles = [_normalize_polish_name(m) or m for m in parsed.get('middle', [])]
    parts = [first.lower()] + [m.lower() for m in middles] + [last.lower()]
    return ' '.join(parts)


def _can_be_same_person(p1, p2):
    """
    Czy dwie osoby mogą być tą samą osobą?
    Porównanie po znormalizowanych formach (mianownik):
    - "Sławomir Feszczak" + "Sławomira Feszczaka" → TAK
    - "Sławomir Feszczak" + "Sławomir Krzysztof Feszczak" → TAK
    """
    if not p1 or not p2:
        return False

    # Normalizuj nazwiska
    last1 = (_normalize_polish_name(p1['last']) or p1['last']).lower()
    last2 = (_normalize_polish_name(p2['last']) or p2['last']).lower()
    if last1 != last2:
        return False

    # Normalizuj wszystkie imiona (pierwsze + middle)
    p1_names = {(_normalize_polish_name(p1['first']) or p1['first']).lower()}
    p1_names |= {(_normalize_polish_name(m) or m).lower() for m in p1.get('middle', [])}

    p2_names = {(_normalize_polish_name(p2['first']) or p2['first']).lower()}
    p2_names |= {(_normalize_polish_name(m) or m).lower() for m in p2.get('middle', [])}

    # Czy mają choć jedno wspólne (znormalizowane) imię?
    if p1_names & p2_names:
        return True

    return False


def auto_link_persons():
    """
    Główna funkcja łącząca osoby:
    1. Łączenie po ZNORMALIZOWANEJ canonical form (uwzględnia deklinacje PL)
       "Sławomir Feszczak" + "Sławomira Feszczaka" → te same osoby
    2. Łączenie po compound names (1+ imion + nazwisko)
    3. Łączenie po PESEL (100% match)
    4. Łączenie po nazwisku + adresie
    """
    db = DatabaseManager(CONFIG['db_path'])
    db.connect()

    stats = {
        'by_name': 0,
        'by_normalized': 0,
        'by_pesel': 0,
        'by_address': 0,
        'total_aliases': 0
    }

    try:
        # Pobierz wszystkie osoby
        persons = db.fetch_all("""
            SELECT id, entity_value, normalized_value
            FROM entities
            WHERE entity_type = 'person'
            ORDER BY LENGTH(entity_value) DESC
        """)

        # ETAP 1: Łączenie po ZNORMALIZOWANEJ canonical form
        # "Sławomir Feszczak" → "sławomir feszczak"
        # "Sławomira Feszczaka" → "sławomir feszczak"  (po normalizacji deklinacji)
        # Te same canonical → te same osoby
        normalized_groups = {}
        for p in persons:
            canonical_form = _normalize_person_canonical(p['entity_value'])
            if not canonical_form:
                continue
            parsed = _parse_person_name(p['entity_value'])
            normalized_groups.setdefault(canonical_form, []).append({
                'id': p['id'],
                'value': p['entity_value'],
                'parsed': parsed,
                'normalized_canonical': canonical_form
            })

        # W każdej grupie - wybierz canonical jako formę mianownika
        # (najbliższa znormalizowanej formie = mianownik = oryginalna forma)
        for canonical_form, members in normalized_groups.items():
            if len(members) < 2:
                continue

            def score_canonical(m):
                # Wybierz formę która jest najbliższa znormalizowanej (mianownik)
                # 1. najwięcej członów (compound names)
                # 2. odległość od canonical_form (im bliżej = lepszy mianownik)
                value_lower = m['value'].lower()
                parts_count = len(value_lower.split())
                # Jak blisko mianownika? Im więcej liter wspólnych z canonical_form, tym lepiej
                norm_parts = canonical_form.split()
                value_parts = value_lower.split()
                similarity = 0
                for vp, np in zip(value_parts, norm_parts):
                    if vp == np:
                        similarity += 100  # Identyczne (mianownik)
                    elif vp.startswith(np):
                        similarity += 50   # Prawie mianownik
                # Preferuj: więcej członów + bliżej mianownika
                return (parts_count, similarity, -len(value_lower))

            members.sort(key=score_canonical, reverse=True)
            canonical = members[0]

            for alias in members[1:]:
                if alias['id'] == canonical['id']:
                    continue
                try:
                    db.execute("""
                        INSERT OR IGNORE INTO person_aliases
                        (canonical_entity_id, alias_entity_id, confidence, match_reason)
                        VALUES (?, ?, ?, ?)
                    """, (canonical['id'], alias['id'], 0.95, f"normalized_form: '{alias['value']}' → '{canonical['value']}'"))
                    stats['by_normalized'] += 1
                except Exception:
                    pass

        # ETAP 2: Łączenie po compound names (base = imię + nazwisko)
        groups = {}  # base_name (lower) → list of {id, parsed}

        for p in persons:
            parsed = _parse_person_name(p['entity_value'])
            if not parsed:
                continue

            # Użyj znormalizowanej formy base
            first_norm = (_normalize_polish_name(parsed['first']) or parsed['first']).lower()
            last_norm = (_normalize_polish_name(parsed['last']) or parsed['last']).lower()
            base = f"{first_norm} {last_norm}"

            if base not in groups:
                groups[base] = []
            groups[base].append({
                'id': p['id'],
                'value': p['entity_value'],
                'parsed': parsed
            })

        # Dla każdej grupy znajdź canonical (najdłuższe imię)
        for base, members in groups.items():
            if len(members) < 2:
                continue

            members.sort(key=lambda m: -len(m['parsed']['full']))
            canonical = members[0]

            for alias in members[1:]:
                if alias['id'] == canonical['id']:
                    continue
                try:
                    # Sprawdź czy już połączone w którąkolwiek stronę
                    existing = db.fetch_one("""
                        SELECT id FROM person_aliases
                        WHERE (canonical_entity_id = ? AND alias_entity_id = ?)
                           OR (canonical_entity_id = ? AND alias_entity_id = ?)
                    """, (canonical['id'], alias['id'], alias['id'], canonical['id']))
                    if not existing:
                        db.execute("""
                            INSERT INTO person_aliases
                            (canonical_entity_id, alias_entity_id, confidence, match_reason)
                            VALUES (?, ?, ?, ?)
                        """, (canonical['id'], alias['id'], 0.9, f"compound_name: '{alias['value']}' → '{canonical['value']}'"))
                        stats['by_name'] += 1
                except Exception:
                    pass

        # Cross-group linking: sprawdź pary z różnych baz ale tym samym nazwiskiem
        # "Sławomir Feszczak" (base: slawomir feszczak)
        # "Krzysztof Feszczak" (base: krzysztof feszczak)
        # NIE łącz - to różne osoby
        # ALE: "Sławomir Krzysztof Feszczak" (base: slawomir feszczak)
        # ↔ "Krzysztof Sławomir Feszczak" (base: krzysztof feszczak)
        # → mają wspólne imiona, ta sama osoba

        all_parsed = []
        for p in persons:
            parsed = _parse_person_name(p['entity_value'])
            if parsed:
                all_parsed.append({
                    'id': p['id'],
                    'value': p['entity_value'],
                    'parsed': parsed
                })

        # O(n^2) - dla dużych zbiorów może być wolne, ale na początek OK
        for i, p1 in enumerate(all_parsed):
            for p2 in all_parsed[i+1:]:
                if p1['id'] == p2['id']:
                    continue
                if _can_be_same_person(p1['parsed'], p2['parsed']):
                    # Canonical = osoba z dłuższym full_name
                    if len(p1['parsed']['full']) >= len(p2['parsed']['full']):
                        canonical_id, alias_id = p1['id'], p2['id']
                        reason = f"compound_match: '{p2['value']}' → '{p1['value']}'"
                    else:
                        canonical_id, alias_id = p2['id'], p1['id']
                        reason = f"compound_match: '{p1['value']}' → '{p2['value']}'"

                    try:
                        # Sprawdź czy już istnieje (w którąś stronę)
                        existing = db.fetch_one("""
                            SELECT id FROM person_aliases
                            WHERE (canonical_entity_id = ? AND alias_entity_id = ?)
                               OR (canonical_entity_id = ? AND alias_entity_id = ?)
                        """, (canonical_id, alias_id, alias_id, canonical_id))

                        if not existing:
                            db.execute("""
                                INSERT INTO person_aliases
                                (canonical_entity_id, alias_entity_id, confidence, match_reason)
                                VALUES (?, ?, ?, ?)
                            """, (canonical_id, alias_id, 0.85, reason))
                            stats['by_name'] += 1
                    except Exception:
                        pass

        # ETAP 3: Łączenie po PESEL - TYLKO wysokie confidence (>= 0.8)
        # Niskie confidence (0.5) = wiele osób na 1 stronie z 1 PESEL = niewiarygodne
        # Też: tylko jeśli osoby mają TO SAMO NAZWISKO (additional check)
        pesel_groups = db.fetch_all("""
            SELECT attr_value, GROUP_CONCAT(entity_id) as entity_ids
            FROM person_attributes
            WHERE attr_type = 'pesel' AND confidence >= 0.8
            GROUP BY attr_value
            HAVING COUNT(DISTINCT entity_id) > 1
        """)

        for grp in pesel_groups:
            ids = [int(x) for x in grp['entity_ids'].split(',')]
            # Pobierz nazwiska
            placeholders_ids = ','.join('?' * len(ids))
            entities_in_grp = db.fetch_all(
                f"SELECT id, entity_value FROM entities WHERE id IN ({placeholders_ids})",
                ids
            )
            # Grupuj po nazwisku (znormalizowane)
            by_surname = {}
            for e in entities_in_grp:
                parsed = _parse_person_name(e['entity_value'])
                if parsed:
                    last_norm = (_normalize_polish_name(parsed['last']) or parsed['last']).lower()
                    by_surname.setdefault(last_norm, []).append({'id': e['id'], 'value': e['entity_value']})

            # Łącz TYLKO jeśli to samo nazwisko + ten sam PESEL
            for surname, members in by_surname.items():
                if len(members) < 2:
                    continue
                canonical_id = members[0]['id']
                for member in members[1:]:
                    try:
                        existing = db.fetch_one("""
                            SELECT id FROM person_aliases
                            WHERE (canonical_entity_id = ? AND alias_entity_id = ?)
                               OR (canonical_entity_id = ? AND alias_entity_id = ?)
                        """, (canonical_id, member['id'], member['id'], canonical_id))
                        if not existing:
                            db.execute("""
                                INSERT INTO person_aliases
                                (canonical_entity_id, alias_entity_id, confidence, match_reason)
                                VALUES (?, ?, ?, ?)
                            """, (canonical_id, member['id'], 1.0, f"pesel_match: {grp['attr_value']} (same surname)"))
                            stats['by_pesel'] += 1
                    except Exception:
                        pass

        # ETAP 3: Łączenie po nazwisku + adresie
        # Osoby z tym samym nazwiskiem które dzielą adres
        addr_groups = db.fetch_all("""
            SELECT attr_value, GROUP_CONCAT(entity_id) as entity_ids
            FROM person_attributes
            WHERE attr_type = 'address'
            GROUP BY attr_value
            HAVING COUNT(DISTINCT entity_id) > 1
        """)

        for grp in addr_groups:
            ids = [int(x) for x in grp['entity_ids'].split(',')]
            # Pobierz nazwiska tych osób
            entities = db.fetch_all(
                f"SELECT id, entity_value FROM entities WHERE id IN ({','.join('?' * len(ids))})",
                ids
            )
            # Grupuj po nazwisku
            by_surname = {}
            for e in entities:
                parsed = _parse_person_name(e['entity_value'])
                if parsed:
                    sn = parsed['last'].lower()
                    by_surname.setdefault(sn, []).append({'id': e['id'], 'value': e['entity_value']})

            for surname, members in by_surname.items():
                if len(members) < 2:
                    continue
                canonical = members[0]
                for alias in members[1:]:
                    try:
                        existing = db.fetch_one("""
                            SELECT id FROM person_aliases
                            WHERE (canonical_entity_id = ? AND alias_entity_id = ?)
                               OR (canonical_entity_id = ? AND alias_entity_id = ?)
                        """, (canonical['id'], alias['id'], alias['id'], canonical['id']))
                        if not existing:
                            db.execute("""
                                INSERT INTO person_aliases
                                (canonical_entity_id, alias_entity_id, confidence, match_reason)
                                VALUES (?, ?, ?, ?)
                            """, (canonical['id'], alias['id'], 0.75, f"address+surname: '{grp['attr_value'][:50]}'"))
                            stats['by_address'] += 1
                    except Exception:
                        pass

        # Policz total aliases
        total = db.fetch_one("SELECT COUNT(*) as cnt FROM person_aliases")
        stats['total_aliases'] = total['cnt'] if total else 0

    finally:
        db.disconnect()

    return stats


# ============================================
# PERSON VERIFICATION (akceptacja / odrzucenie)
# ============================================

@app.route('/api/persons/<int:entity_id>/verify', methods=['POST'])
def person_verify(entity_id):
    """
    Oznacz osobę jako zweryfikowaną.
    Po weryfikacji znika z głównej listy i ląduje w sekcji "Zweryfikowane".
    Propaguje do wszystkich aliasów.
    """
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Znajdź canonical
        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id

        # Wszystkie powiązane ID (canonical + aliasy)
        aliases = db.fetch_all("SELECT alias_entity_id FROM person_aliases WHERE canonical_entity_id = ?", (canonical_id,))
        all_ids = [canonical_id] + [a['alias_entity_id'] for a in aliases]

        placeholders = ','.join('?' * len(all_ids))
        db.execute(f"""
            UPDATE entities
            SET verified = 1, verified_at = CURRENT_TIMESTAMP, verified_by = 'manual'
            WHERE id IN ({placeholders})
        """, all_ids)

        db.disconnect()
        return jsonify({'success': True, 'verified_count': len(all_ids)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/<int:entity_id>/unverify', methods=['POST'])
def person_unverify(entity_id):
    """Cofnij weryfikację - osoba wraca do głównej listy"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id

        aliases = db.fetch_all("SELECT alias_entity_id FROM person_aliases WHERE canonical_entity_id = ?", (canonical_id,))
        all_ids = [canonical_id] + [a['alias_entity_id'] for a in aliases]
        placeholders = ','.join('?' * len(all_ids))

        db.execute(f"""
            UPDATE entities
            SET verified = 0, verified_at = NULL, verified_by = NULL
            WHERE id IN ({placeholders})
        """, all_ids)

        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/verification', methods=['GET'])
def persons_verification_list():
    """
    Lista osób pogrupowana wg statusu weryfikacji.
    Zwraca: pending (do weryfikacji) + verified (zaakceptowane)
    Dla pending dodaje sugestie podobnych osób
    """
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz wszystkie osoby z liczbą wystąpień
        all_persons = db.fetch_all("""
            SELECT
                e.id, e.entity_value, e.verified, e.verified_at,
                (SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = e.id) as alias_of,
                (SELECT COUNT(*) FROM entity_occurrences WHERE entity_id = e.id) as occurrences
            FROM entities e
            WHERE e.entity_type = 'person'
            ORDER BY occurrences DESC
        """)

        # Wykryj duplikaty/aliasy które mogą być scalone
        # Dla każdej canonical/standalone osoby - znajdź podobne (potencjalne aliasy do scalenia)
        pending = []
        verified = []
        all_parsed = []
        for p in all_persons:
            parsed = _parse_person_name(p['entity_value'])
            all_parsed.append({
                'id': p['id'],
                'value': p['entity_value'],
                'verified': bool(p['verified']),
                'verified_at': p['verified_at'],
                'alias_of': p['alias_of'],
                'occurrences': p['occurrences'],
                'parsed': parsed,
                'canonical_form': _normalize_person_canonical(p['entity_value']) if parsed else None
            })

        # Pokaż tylko canonical (nie aliasy) - aliasy już są pod canonical
        canonical_persons = [p for p in all_parsed if not p['alias_of']]

        for p in canonical_persons:
            # Pobierz aliasy
            aliases = db.fetch_all("""
                SELECT pa.alias_entity_id as id, e.entity_value, pa.confidence, pa.match_reason
                FROM person_aliases pa
                JOIN entities e ON pa.alias_entity_id = e.id
                WHERE pa.canonical_entity_id = ?
            """, (p['id'],))

            # Znajdź sugestie podobnych osób (nie w aliasach jeszcze)
            existing_alias_ids = {p['id']} | {a['id'] for a in aliases}
            suggestions = []
            for other in canonical_persons:
                if other['id'] == p['id'] or other['id'] in existing_alias_ids:
                    continue
                if other['parsed'] and p['parsed']:
                    if _can_be_same_person(p['parsed'], other['parsed']):
                        suggestions.append({
                            'id': other['id'],
                            'value': other['value'],
                            'occurrences': other['occurrences']
                        })
                    elif p['canonical_form'] and other['canonical_form'] and p['canonical_form'] == other['canonical_form']:
                        suggestions.append({
                            'id': other['id'],
                            'value': other['value'],
                            'occurrences': other['occurrences']
                        })

            item = {
                'id': p['id'],
                'name': p['value'],
                'occurrences': p['occurrences'],
                'verified': p['verified'],
                'verified_at': p['verified_at'],
                'aliases': [dict(a) for a in aliases],
                'aliases_count': len(aliases),
                'suggestions': suggestions[:5]  # max 5 sugestii
            }

            if p['verified']:
                verified.append(item)
            else:
                pending.append(item)

        db.disconnect()

        return jsonify({
            'success': True,
            'pending': pending,
            'verified': verified,
            'stats': {
                'total': len(canonical_persons),
                'pending': len(pending),
                'verified': len(verified)
            }
        })
    except Exception as e:
        import traceback
        print(f"❌ persons_verification_list: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/merge-bulk', methods=['POST'])
def persons_merge_bulk():
    """
    Scal wiele osób w jedną canonical.
    Body: {canonical_id: X, alias_ids: [Y, Z, ...]}
    """
    try:
        data = request.json or {}
        canonical_id = data.get('canonical_id')
        alias_ids = data.get('alias_ids', [])

        if not canonical_id or not alias_ids:
            return jsonify({'error': 'Brak parametrów'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        merged = 0
        for alias_id in alias_ids:
            if alias_id == canonical_id:
                continue

            # Jeśli alias jest canonical innej grupy - przenieś jego aliasy
            sub_aliases = db.fetch_all(
                "SELECT alias_entity_id FROM person_aliases WHERE canonical_entity_id = ?",
                (alias_id,)
            )
            for sub in sub_aliases:
                try:
                    db.execute("""
                        UPDATE person_aliases
                        SET canonical_entity_id = ?
                        WHERE alias_entity_id = ?
                    """, (canonical_id, sub['alias_entity_id']))
                except Exception:
                    pass

            # Usuń stare aliasy gdzie alias_id był canonical
            db.execute("DELETE FROM person_aliases WHERE canonical_entity_id = ?", (alias_id,))

            # Dodaj jako alias
            try:
                db.execute("""
                    INSERT OR REPLACE INTO person_aliases
                    (canonical_entity_id, alias_entity_id, confidence, match_reason)
                    VALUES (?, ?, ?, ?)
                """, (canonical_id, alias_id, 1.0, 'manual_merge'))
                merged += 1
            except Exception:
                pass

        db.disconnect()
        return jsonify({'success': True, 'merged': merged})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/auto-link', methods=['POST'])
def persons_auto_link():
    """Automatycznie połącz osoby (compound names, PESEL, adresy)"""
    try:
        stats = auto_link_persons()
        return jsonify({
            'success': True,
            'stats': stats,
            'message': (f"Połączono: {stats.get('by_normalized', 0)} po formie znormalizowanej (deklinacje), "
                       f"{stats['by_name']} po imionach, "
                       f"{stats['by_pesel']} po PESEL, "
                       f"{stats['by_address']} po adresie. "
                       f"Łącznie {stats['total_aliases']} aliasów.")
        })
    except Exception as e:
        import traceback
        print(f"❌ persons_auto_link error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/groups', methods=['GET'])
def persons_groups():
    """Pobierz grupy połączonych osób"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz wszystkie aliasy z canonical entities
        aliases = db.fetch_all("""
            SELECT
                pa.canonical_entity_id,
                pa.alias_entity_id,
                pa.confidence,
                pa.match_reason,
                ce.entity_value as canonical_value,
                ae.entity_value as alias_value,
                (SELECT COUNT(*) FROM entity_occurrences WHERE entity_id = pa.canonical_entity_id) as canonical_occurrences,
                (SELECT COUNT(*) FROM entity_occurrences WHERE entity_id = pa.alias_entity_id) as alias_occurrences
            FROM person_aliases pa
            JOIN entities ce ON pa.canonical_entity_id = ce.id
            JOIN entities ae ON pa.alias_entity_id = ae.id
            ORDER BY pa.canonical_entity_id
        """)

        # Grupuj po canonical
        groups = {}
        for a in aliases:
            cid = a['canonical_entity_id']
            if cid not in groups:
                groups[cid] = {
                    'canonical_id': cid,
                    'canonical_name': a['canonical_value'],
                    'canonical_occurrences': a['canonical_occurrences'],
                    'aliases': []
                }
            groups[cid]['aliases'].append({
                'id': a['alias_entity_id'],
                'name': a['alias_value'],
                'occurrences': a['alias_occurrences'],
                'confidence': a['confidence'],
                'reason': a['match_reason']
            })

        # Pobierz attrybuty (PESEL, adres) dla każdego canonical
        for grp in groups.values():
            all_ids = [grp['canonical_id']] + [a['id'] for a in grp['aliases']]
            placeholders = ','.join('?' * len(all_ids))
            attrs = db.fetch_all(
                f"SELECT DISTINCT attr_type, attr_value FROM person_attributes WHERE entity_id IN ({placeholders})",
                all_ids
            )
            grp['attributes'] = [{'type': a['attr_type'], 'value': a['attr_value']} for a in attrs]

        db.disconnect()
        return jsonify({
            'success': True,
            'groups': list(groups.values()),
            'count': len(groups)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/link', methods=['POST'])
def persons_link():
    """Ręczne połączenie dwóch osób"""
    try:
        data = request.json
        canonical_id = data.get('canonical_id')
        alias_id = data.get('alias_id')
        reason = data.get('reason', 'manual')

        if not canonical_id or not alias_id:
            return jsonify({'error': 'Brak ID'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("""
            INSERT OR IGNORE INTO person_aliases
            (canonical_entity_id, alias_entity_id, confidence, match_reason)
            VALUES (?, ?, ?, ?)
        """, (canonical_id, alias_id, 1.0, f"manual: {reason}"))
        db.disconnect()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/unlink/<int:alias_id>', methods=['DELETE'])
def persons_unlink(alias_id):
    """Usuń alias (rozłącz osoby)"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("DELETE FROM person_aliases WHERE alias_entity_id = ?", (alias_id,))
        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/<int:entity_id>/attribute', methods=['POST'])
def person_add_attribute(entity_id):
    """Dodaj atrybut do osoby (PESEL, adres, NIP)"""
    try:
        data = request.json
        attr_type = data.get('attr_type', '').strip()
        attr_value = data.get('attr_value', '').strip()

        if attr_type not in ('pesel', 'address', 'nip', 'phone', 'email', 'birthdate'):
            return jsonify({'error': 'Nieprawidłowy typ atrybutu'}), 400

        if not attr_value:
            return jsonify({'error': 'Pusta wartość'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("""
            INSERT OR IGNORE INTO person_attributes
            (entity_id, attr_type, attr_value, confidence)
            VALUES (?, ?, ?, 1.0)
        """, (entity_id, attr_type, attr_value))
        db.disconnect()

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/persons/<int:entity_id>/details', methods=['GET'])
def person_get_details(entity_id):
    """Pobierz szczegóły osoby - atrybuty, aliasy, wystąpienia"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        person = db.fetch_one("SELECT * FROM entities WHERE id = ? AND entity_type = 'person'", (entity_id,))
        if not person:
            db.disconnect()
            return jsonify({'error': 'Osoba nie znaleziona'}), 404

        # Sprawdź czy to alias - jeśli tak, znajdź canonical
        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id

        canonical = db.fetch_one("SELECT * FROM entities WHERE id = ?", (canonical_id,))

        # Pobierz wszystkie aliasy (canonical + jego aliasy)
        aliases_rows = db.fetch_all(
            "SELECT * FROM person_aliases WHERE canonical_entity_id = ?",
            (canonical_id,)
        )

        all_entity_ids = [canonical_id] + [a['alias_entity_id'] for a in aliases_rows]
        placeholders = ','.join('?' * len(all_entity_ids))

        # Atrybuty
        attrs = db.fetch_all(
            f"SELECT DISTINCT attr_type, attr_value, confidence FROM person_attributes WHERE entity_id IN ({placeholders})",
            all_entity_ids
        )

        # Pliki w których pojawia się osoba (suma)
        files = db.fetch_all(f"""
            SELECT DISTINCT sf.id, sf.filename, COUNT(eo.id) as occurrence_count
            FROM entity_occurrences eo
            JOIN source_files sf ON eo.file_id = sf.id
            WHERE eo.entity_id IN ({placeholders})
            GROUP BY sf.id
            ORDER BY occurrence_count DESC
        """, all_entity_ids)

        db.disconnect()

        return jsonify({
            'success': True,
            'canonical': dict(canonical),
            'is_alias_of': canonical_id if canonical_id != entity_id else None,
            'aliases': [{
                'id': a['alias_entity_id'],
                'confidence': a['confidence'],
                'reason': a['match_reason']
            } for a in aliases_rows],
            'attributes': [{'type': a['attr_type'], 'value': a['attr_value'], 'confidence': a['confidence']} for a in attrs],
            'files': [{'id': f['id'], 'filename': f['filename'], 'occurrences': f['occurrence_count']} for f in files]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# PERSON PROFILE + TIMELINE
# ============================================

def _extract_date_from_text(text):
    """Wyciągnij datę z tekstu (różne formaty PL)"""
    if not text:
        return None
    import re as _re
    # YYYY-MM-DD, DD.MM.YYYY, DD/MM/YYYY
    patterns = [
        (r'\b(20\d{2})-(\d{1,2})-(\d{1,2})\b', lambda m: f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"),
        (r'\b(\d{1,2})\.(\d{1,2})\.(20\d{2})\b', lambda m: f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"),
        (r'\b(\d{1,2})/(\d{1,2})/(20\d{2})\b', lambda m: f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"),
    ]
    for pattern, fmt in patterns:
        m = _re.search(pattern, text)
        if m:
            try:
                return fmt(m)
            except Exception:
                continue
    return None


@app.route('/api/persons/<int:entity_id>/profile', methods=['GET'])
def person_profile(entity_id):
    """Pełny profil osoby z wszystkimi danymi"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        person = db.fetch_one("SELECT * FROM entities WHERE id = ? AND entity_type = 'person'", (entity_id,))
        if not person:
            db.disconnect()
            return jsonify({'error': 'Osoba nie znaleziona'}), 404

        # Sprawdź alias chain - znajdź canonical
        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id
        canonical = db.fetch_one("SELECT * FROM entities WHERE id = ?", (canonical_id,))

        # Wszystkie aliasy
        aliases = db.fetch_all(
            "SELECT * FROM person_aliases WHERE canonical_entity_id = ?",
            (canonical_id,)
        )
        all_entity_ids = [canonical_id] + [a['alias_entity_id'] for a in aliases]
        placeholders = ','.join('?' * len(all_entity_ids))

        # Pobierz wszystkie nazwy (formy)
        all_names = db.fetch_all(
            f"SELECT id, entity_value FROM entities WHERE id IN ({placeholders})",
            all_entity_ids
        )

        # Atrybuty (PESEL, adres, NIP, tel, email)
        attrs = db.fetch_all(
            f"SELECT attr_type, attr_value, confidence FROM person_attributes WHERE entity_id IN ({placeholders})",
            all_entity_ids
        )
        attrs_by_type = {}
        for a in attrs:
            attrs_by_type.setdefault(a['attr_type'], []).append({
                'value': a['attr_value'],
                'confidence': a['confidence']
            })

        # Wszystkie dokumenty (z liczbą wystąpień)
        files = db.fetch_all(f"""
            SELECT
                sf.id, sf.filename, sf.page_count,
                b.name as binder_name,
                COUNT(eo.id) as occurrences,
                MIN(eo.page_id) as first_page_id,
                sf.created_at
            FROM entity_occurrences eo
            JOIN source_files sf ON eo.file_id = sf.id
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE eo.entity_id IN ({placeholders})
            GROUP BY sf.id
            ORDER BY sf.created_at DESC
        """, all_entity_ids)

        # Timeline - dokumenty z wykrytymi datami
        timeline = []
        for f in files:
            # Spróbuj wyciągnąć datę z pierwszej strony
            first_page = db.fetch_one(
                "SELECT text_content, fixed_text FROM pages WHERE id = ?",
                (f['first_page_id'],)
            )
            doc_date = None
            doc_type = None
            if first_page:
                text = first_page['fixed_text'] or first_page['text_content'] or ''
                doc_date = _extract_date_from_text(text[:1000])

            # Pobierz typ dokumentu (jeśli wykryty)
            doc_type_row = db.fetch_one(
                "SELECT doc_type FROM document_types WHERE file_id = ? ORDER BY confidence DESC LIMIT 1",
                (f['id'],)
            )
            if doc_type_row:
                doc_type = doc_type_row['doc_type']

            timeline.append({
                'file_id': f['id'],
                'filename': f['filename'],
                'binder_name': f['binder_name'],
                'page_count': f['page_count'],
                'occurrences': f['occurrences'],
                'date': doc_date,
                'doc_type': doc_type,
                'created_at': f['created_at']
            })

        # Sortuj timeline po dacie (najnowsze pierwsze)
        timeline.sort(key=lambda t: t['date'] or t['created_at'] or '', reverse=True)

        # KW związane z osobą (osoby pojawiają się na tych samych stronach co KW)
        kws = db.fetch_all(f"""
            SELECT DISTINCT lr.kw_full, lr.kw_district, lr.id as kw_id
            FROM entity_occurrences eo
            JOIN land_register_occurrences lro ON eo.page_id = lro.page_id
            JOIN land_registers lr ON lro.kw_id = lr.id
            WHERE eo.entity_id IN ({placeholders})
        """, all_entity_ids)

        # Powiązane osoby (pojawiają się na tych samych stronach)
        related_persons = db.fetch_all(f"""
            SELECT DISTINCT e2.id, e2.entity_value, COUNT(DISTINCT eo2.page_id) as shared_pages
            FROM entity_occurrences eo1
            JOIN entity_occurrences eo2 ON eo1.page_id = eo2.page_id
            JOIN entities e2 ON eo2.entity_id = e2.id
            WHERE eo1.entity_id IN ({placeholders})
              AND e2.entity_type = 'person'
              AND e2.id NOT IN ({placeholders})
            GROUP BY e2.id
            ORDER BY shared_pages DESC
            LIMIT 10
        """, all_entity_ids + all_entity_ids)

        # Czy jest ulubiona?
        fav = db.fetch_one(
            "SELECT id FROM favorites WHERE favorite_type = 'person' AND target_id = ?",
            (canonical_id,)
        )

        db.disconnect()

        return jsonify({
            'success': True,
            'canonical': dict(canonical),
            'is_alias_of': canonical_id if canonical_id != entity_id else None,
            'all_names': [dict(n) for n in all_names],
            'aliases_count': len(aliases),
            'attributes': attrs_by_type,
            'files': [dict(f) for f in files],
            'timeline': timeline,
            'kws': [dict(k) for k in kws],
            'related_persons': [dict(rp) for rp in related_persons],
            'is_favorite': bool(fav),
            'stats': {
                'total_files': len(files),
                'total_occurrences': sum(f['occurrences'] for f in files),
                'total_kws': len(kws),
                'total_related': len(related_persons)
            }
        })

    except Exception as e:
        import traceback
        print(f"❌ person_profile error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ============================================
# DUPLICATES DETECTION
# ============================================

def _compute_text_hash(text, length=500):
    """Hash pierwszych N znaków tekstu (do porównania)"""
    if not text:
        return None
    import hashlib
    normalized = ''.join(text.lower().split())[:length]
    if len(normalized) < 100:
        return None
    return hashlib.md5(normalized.encode()).hexdigest()


@app.route('/api/duplicates/scan', methods=['POST'])
def duplicates_scan():
    """Wykryj duplikaty w bazie"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        stats = {'by_file_hash': 0, 'by_text_hash': 0, 'by_page_count': 0}

        # 1. Duplikaty po file_hash (identyczne pliki)
        hash_groups = db.fetch_all("""
            SELECT file_hash, GROUP_CONCAT(id) as ids, COUNT(*) as cnt
            FROM source_files
            WHERE file_hash IS NOT NULL AND file_hash != ''
            GROUP BY file_hash
            HAVING cnt > 1
        """)

        for grp in hash_groups:
            ids = [int(x) for x in grp['ids'].split(',')]
            for i in range(len(ids)):
                for j in range(i + 1, len(ids)):
                    try:
                        db.execute("""
                            INSERT OR IGNORE INTO duplicates
                            (file_id_1, file_id_2, similarity, match_type, status)
                            VALUES (?, ?, 1.0, 'file_hash', 'pending')
                        """, (min(ids[i], ids[j]), max(ids[i], ids[j])))
                        stats['by_file_hash'] += 1
                    except Exception:
                        pass

        # 2. Duplikaty po liczbie stron + treści pierwszej strony
        files = db.fetch_all("""
            SELECT sf.id, sf.filename, sf.page_count,
                   (SELECT COALESCE(p.fixed_text, p.text_content)
                    FROM pages p WHERE p.file_id = sf.id AND p.page_number = 1) as first_text
            FROM source_files sf
            WHERE sf.page_count > 0
        """)

        text_hash_map = {}
        for f in files:
            th = _compute_text_hash(f['first_text'])
            if th:
                key = (f['page_count'], th)
                text_hash_map.setdefault(key, []).append(f['id'])

        for ids in text_hash_map.values():
            if len(ids) < 2:
                continue
            for i in range(len(ids)):
                for j in range(i + 1, len(ids)):
                    try:
                        db.execute("""
                            INSERT OR IGNORE INTO duplicates
                            (file_id_1, file_id_2, similarity, match_type, status)
                            VALUES (?, ?, 0.95, 'text_hash', 'pending')
                        """, (min(ids[i], ids[j]), max(ids[i], ids[j])))
                        stats['by_text_hash'] += 1
                    except Exception:
                        pass

        total = db.fetch_one("SELECT COUNT(*) as cnt FROM duplicates WHERE status = 'pending'")
        db.disconnect()

        return jsonify({
            'success': True,
            'stats': stats,
            'total_pending': total['cnt'] if total else 0,
            'message': f"Znaleziono {sum(stats.values())} potencjalnych duplikatów"
        })
    except Exception as e:
        import traceback
        print(f"❌ duplicates_scan: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/duplicates/list', methods=['GET'])
def duplicates_list():
    """Lista wykrytych duplikatów"""
    try:
        status_filter = request.args.get('status', 'pending')
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        if status_filter == 'all':
            rows = db.fetch_all("""
                SELECT d.*,
                    f1.filename as filename_1, f1.page_count as pages_1,
                    f2.filename as filename_2, f2.page_count as pages_2
                FROM duplicates d
                JOIN source_files f1 ON d.file_id_1 = f1.id
                JOIN source_files f2 ON d.file_id_2 = f2.id
                ORDER BY d.detected_at DESC
            """)
        else:
            rows = db.fetch_all("""
                SELECT d.*,
                    f1.filename as filename_1, f1.page_count as pages_1,
                    f2.filename as filename_2, f2.page_count as pages_2
                FROM duplicates d
                JOIN source_files f1 ON d.file_id_1 = f1.id
                JOIN source_files f2 ON d.file_id_2 = f2.id
                WHERE d.status = ?
                ORDER BY d.detected_at DESC
            """, (status_filter,))

        db.disconnect()
        return jsonify({
            'success': True,
            'duplicates': [dict(r) for r in rows]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/duplicates/<int:dup_id>/resolve', methods=['POST'])
def duplicates_resolve(dup_id):
    """Rozstrzygnij duplikat: keep / delete / ignore"""
    try:
        data = request.json or {}
        action = data.get('action', 'ignore')

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        dup = db.fetch_one("SELECT * FROM duplicates WHERE id = ?", (dup_id,))
        if not dup:
            db.disconnect()
            return jsonify({'error': 'Duplikat nie znaleziony'}), 404

        if action == 'delete_second':
            # Usuń file_id_2
            file_id_to_delete = dup['file_id_2']
            db.execute("DELETE FROM entity_occurrences WHERE file_id = ?", (file_id_to_delete,))
            db.execute("DELETE FROM land_register_occurrences WHERE file_id = ?", (file_id_to_delete,))
            db.execute("DELETE FROM pages WHERE file_id = ?", (file_id_to_delete,))
            db.execute("DELETE FROM source_files WHERE id = ?", (file_id_to_delete,))
            db.execute("UPDATE duplicates SET status = 'resolved' WHERE id = ?", (dup_id,))
        elif action == 'delete_first':
            file_id_to_delete = dup['file_id_1']
            db.execute("DELETE FROM entity_occurrences WHERE file_id = ?", (file_id_to_delete,))
            db.execute("DELETE FROM land_register_occurrences WHERE file_id = ?", (file_id_to_delete,))
            db.execute("DELETE FROM pages WHERE file_id = ?", (file_id_to_delete,))
            db.execute("DELETE FROM source_files WHERE id = ?", (file_id_to_delete,))
            db.execute("UPDATE duplicates SET status = 'resolved' WHERE id = ?", (dup_id,))
        else:
            # ignore
            db.execute("UPDATE duplicates SET status = 'ignored' WHERE id = ?", (dup_id,))

        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# ANNOTATIONS - adnotacje + highlighty na PDF
# ============================================

@app.route('/api/annotations/file/<int:file_id>/page/<int:page_number>', methods=['GET'])
def annotations_get(file_id, page_number):
    """Pobierz adnotacje dla strony"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT * FROM document_annotations
            WHERE file_id = ? AND page_number = ?
            ORDER BY created_at ASC
        """, (file_id, page_number))
        db.disconnect()
        return jsonify({'success': True, 'annotations': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/annotations', methods=['POST'])
def annotations_create():
    """Dodaj adnotację"""
    try:
        data = request.json or {}
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("""
            INSERT INTO document_annotations
            (file_id, page_number, annotation_type, color, note_text, selected_text,
             position_x, position_y, width, height)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get('file_id'),
            data.get('page_number'),
            data.get('annotation_type', 'note'),
            data.get('color', 'yellow'),
            data.get('note_text', ''),
            data.get('selected_text', ''),
            data.get('position_x', 0),
            data.get('position_y', 0),
            data.get('width', 0),
            data.get('height', 0),
        ))
        new_id = db.fetch_one("SELECT last_insert_rowid() as id")['id']
        db.disconnect()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/annotations/<int:ann_id>', methods=['PUT'])
def annotations_update(ann_id):
    """Update adnotacji (np. zmień notatkę)"""
    try:
        data = request.json or {}
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        updates = []
        params = []
        for field in ['note_text', 'color', 'annotation_type']:
            if field in data:
                updates.append(f"{field} = ?")
                params.append(data[field])

        if updates:
            updates.append("updated_at = CURRENT_TIMESTAMP")
            params.append(ann_id)
            db.execute(f"UPDATE document_annotations SET {', '.join(updates)} WHERE id = ?", params)

        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/annotations/<int:ann_id>', methods=['DELETE'])
def annotations_delete(ann_id):
    """Usuń adnotację"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("DELETE FROM document_annotations WHERE id = ?", (ann_id,))
        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/annotations/file/<int:file_id>', methods=['GET'])
def annotations_for_file(file_id):
    """Wszystkie adnotacje w pliku"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT * FROM document_annotations
            WHERE file_id = ?
            ORDER BY page_number ASC, created_at ASC
        """, (file_id,))
        db.disconnect()
        return jsonify({'success': True, 'annotations': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# FAVORITES - ulubione
# ============================================

@app.route('/api/favorites', methods=['GET'])
def favorites_list():
    """Pobierz wszystkie ulubione"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("SELECT * FROM favorites ORDER BY pin_order DESC, created_at DESC")

        # Dodaj nazwy targetów
        result = []
        for r in rows:
            r = dict(r)
            if r['favorite_type'] == 'person':
                ent = db.fetch_one("SELECT entity_value FROM entities WHERE id = ?", (r['target_id'],))
                r['target_name'] = ent['entity_value'] if ent else f'Osoba #{r["target_id"]}'
            elif r['favorite_type'] == 'file':
                f = db.fetch_one("SELECT filename FROM source_files WHERE id = ?", (r['target_id'],))
                r['target_name'] = f['filename'] if f else f'Plik #{r["target_id"]}'
            elif r['favorite_type'] == 'kw':
                kw = db.fetch_one("SELECT kw_full FROM land_registers WHERE id = ?", (r['target_id'],))
                r['target_name'] = kw['kw_full'] if kw else f'KW #{r["target_id"]}'
            else:
                r['target_name'] = f'#{r["target_id"]}'
            result.append(r)

        db.disconnect()
        return jsonify({'success': True, 'favorites': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/favorites/toggle', methods=['POST'])
def favorites_toggle():
    """Dodaj/usuń z ulubionych"""
    try:
        data = request.json or {}
        fav_type = data.get('favorite_type')
        target_id = data.get('target_id')

        if not fav_type or not target_id:
            return jsonify({'error': 'Brak parametrów'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        existing = db.fetch_one(
            "SELECT id FROM favorites WHERE favorite_type = ? AND target_id = ?",
            (fav_type, target_id)
        )

        if existing:
            db.execute("DELETE FROM favorites WHERE id = ?", (existing['id'],))
            action = 'removed'
        else:
            db.execute("""
                INSERT INTO favorites (favorite_type, target_id, note)
                VALUES (?, ?, ?)
            """, (fav_type, target_id, data.get('note', '')))
            action = 'added'

        db.disconnect()
        return jsonify({'success': True, 'action': action})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# DOCUMENT TEMPLATES - auto-detect type + fields
# ============================================

DOCUMENT_TEMPLATES = {
    'faktura': {
        'keywords': ['faktura', 'faktura vat', 'nabywca', 'sprzedawca', 'numer faktury', 'termin płatności', 'do zapłaty'],
        'fields': {
            'nr_faktury': r'(?:faktura\s+(?:vat\s+)?(?:nr|numer)\.?\s*[:\.\s]?\s*)([A-Z0-9/\-]+)',
            'data': r'(?:data\s+(?:wystawienia|sprzedaży)\s*[:\.\s]?\s*)(\d{1,2}[\.\-/]\d{1,2}[\.\-/]20\d{2})',
            'kwota': r'(?:do\s+zapłaty|razem|łącznie|suma)\s*[:\.\s]*([0-9\s]+[,\.]\d{2})\s*(?:zł|pln)',
            'nip_nabywcy': r'(?:nip\s+nabywcy\s*[:\.\s]?\s*)((?:\d{3}[\-\s]?){2}\d{2}[\-\s]?\d{2}|\d{10})',
        }
    },
    'umowa': {
        'keywords': ['umowa', 'strony umowy', 'zawarta w', 'paragraf', '§', 'warunki umowy'],
        'fields': {
            'data_zawarcia': r'(?:zawarta\s+w\s+[^,]+,?\s+)?(?:dnia\s+)?(\d{1,2}[\.\-/]\d{1,2}[\.\-/]20\d{2})',
            'kwota': r'(?:kwot[ay]|cena|wartość)\s*[:\.\s]*([0-9\s]+[,\.]\d{2})\s*(?:zł|pln)',
        }
    },
    'umowa_najmu': {
        'keywords': ['umowa najmu', 'najemca', 'wynajmujący', 'czynsz'],
        'fields': {
            'czynsz': r'czynsz\s*[:\.\s]*([0-9\s]+[,\.]\d{2})\s*(?:zł|pln)',
            'data': r'dnia\s+(\d{1,2}[\.\-/]\d{1,2}[\.\-/]20\d{2})',
        }
    },
    'umowa_sprzedazy': {
        'keywords': ['umowa sprzedaży', 'kupujący', 'sprzedający', 'cena sprzedaży'],
        'fields': {
            'cena': r'cena\s+sprzedaży\s*[:\.\s]*([0-9\s]+[,\.]\d{2})',
        }
    },
    'akt_notarialny': {
        'keywords': ['akt notarialny', 'notariusz', 'rep. a', 'repertorium', 'kancelarii notarialnej'],
        'fields': {
            'rep_a': r'rep(?:ertorium|\.)\s*a\s*[:\.\s]*(\d+/\d{4})',
            'data': r'dnia\s+(\d{1,2}[\.\-/]\d{1,2}[\.\-/]20\d{2})',
        }
    },
    'decyzja': {
        'keywords': ['decyzja', 'decyzja nr', 'orzeka', 'na podstawie art', 'odmawia', 'uchyla'],
        'fields': {
            'nr_decyzji': r'decyzja\s+(?:nr\.?\s*)?([A-Z0-9/\-\.]+/(?:19|20)\d{2})',
            'organ': r'(?:starosta|wójt|burmistrz|prezydent|sąd|naczelnik)[\w\s]+',
            'data': r'dnia\s+(\d{1,2}[\.\-/]\d{1,2}[\.\-/]20\d{2})',
        }
    },
    'ksiega_wieczysta': {
        'keywords': ['księga wieczysta', 'księga wieczystej', 'dział i', 'dział ii', 'hipoteka', 'kw'],
        'fields': {
            'kw_nr': r'([A-Z]{2}\d{1,2}[A-Z]{1,2}/\d{8}/\d)',
        }
    },
    'wniosek': {
        'keywords': ['wniosek', 'wnoszę', 'zwracam się z prośbą', 'proszę o'],
        'fields': {
            'data': r'dnia\s+(\d{1,2}[\.\-/]\d{1,2}[\.\-/]20\d{2})',
        }
    },
}


def _detect_document_type(text):
    """Wykryj typ dokumentu na podstawie słów kluczowych"""
    if not text:
        return None, 0.0, {}

    text_lower = text.lower()
    best_type = None
    best_score = 0
    best_fields = {}

    for doc_type, config in DOCUMENT_TEMPLATES.items():
        score = sum(1 for kw in config['keywords'] if kw in text_lower)
        if score > best_score:
            best_score = score
            best_type = doc_type
            # Spróbuj wyciągnąć pola
            import re as _re
            fields = {}
            for field_name, pattern in config.get('fields', {}).items():
                m = _re.search(pattern, text, _re.IGNORECASE)
                if m:
                    fields[field_name] = m.group(1).strip()
            best_fields = fields

    if best_type and best_score >= 2:
        confidence = min(best_score / 5.0, 1.0)  # Normalizuj
        return best_type, confidence, best_fields

    return None, 0.0, {}


@app.route('/api/templates/detect/<int:file_id>', methods=['POST'])
def templates_detect(file_id):
    """Wykryj typ dokumentu + wyciągnij pola"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz tekst pierwszych 3 stron
        pages = db.fetch_all("""
            SELECT id, page_number, COALESCE(fixed_text, text_content) as text
            FROM pages
            WHERE file_id = ?
            ORDER BY page_number
            LIMIT 3
        """, (file_id,))

        if not pages:
            db.disconnect()
            return jsonify({'error': 'Brak stron'}), 404

        # Połącz tekst
        combined_text = ' '.join(p['text'] or '' for p in pages)
        doc_type, confidence, fields = _detect_document_type(combined_text)

        # Zapisz wynik
        if doc_type:
            import json as _json
            db.execute("DELETE FROM document_types WHERE file_id = ?", (file_id,))
            db.execute("""
                INSERT INTO document_types (file_id, page_id, doc_type, confidence, extracted_fields)
                VALUES (?, ?, ?, ?, ?)
            """, (file_id, pages[0]['id'], doc_type, confidence, _json.dumps(fields, ensure_ascii=False)))

        db.disconnect()

        return jsonify({
            'success': True,
            'doc_type': doc_type,
            'confidence': round(confidence, 2),
            'fields': fields
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/detect-all', methods=['POST'])
def templates_detect_all():
    """Wykryj typy dla wszystkich dokumentów"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        files = db.fetch_all("SELECT id FROM source_files")
        db.disconnect()

        results = {'detected': 0, 'no_type': 0, 'failed': 0}
        for f in files:
            try:
                db = DatabaseManager(CONFIG['db_path'])
                db.connect()
                pages = db.fetch_all("""
                    SELECT id, page_number, COALESCE(fixed_text, text_content) as text
                    FROM pages WHERE file_id = ? ORDER BY page_number LIMIT 3
                """, (f['id'],))
                if not pages:
                    db.disconnect()
                    results['no_type'] += 1
                    continue

                combined = ' '.join(p['text'] or '' for p in pages)
                doc_type, confidence, fields = _detect_document_type(combined)

                if doc_type:
                    import json as _json
                    db.execute("DELETE FROM document_types WHERE file_id = ?", (f['id'],))
                    db.execute("""
                        INSERT INTO document_types (file_id, page_id, doc_type, confidence, extracted_fields)
                        VALUES (?, ?, ?, ?, ?)
                    """, (f['id'], pages[0]['id'], doc_type, confidence, _json.dumps(fields, ensure_ascii=False)))
                    results['detected'] += 1
                else:
                    results['no_type'] += 1
                db.disconnect()
            except Exception:
                results['failed'] += 1

        return jsonify({'success': True, 'stats': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/by-type/<doc_type>', methods=['GET'])
def templates_by_type(doc_type):
    """Lista dokumentów określonego typu"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT dt.*, sf.filename, sf.page_count
            FROM document_types dt
            JOIN source_files sf ON dt.file_id = sf.id
            WHERE dt.doc_type = ?
            ORDER BY dt.confidence DESC
        """, (doc_type,))
        db.disconnect()
        import json as _json
        result = []
        for r in rows:
            r = dict(r)
            try:
                r['extracted_fields'] = _json.loads(r['extracted_fields']) if r['extracted_fields'] else {}
            except Exception:
                r['extracted_fields'] = {}
            result.append(r)
        return jsonify({'success': True, 'documents': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/summary', methods=['GET'])
def templates_summary():
    """Podsumowanie typów dokumentów"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT doc_type, COUNT(*) as count, AVG(confidence) as avg_confidence
            FROM document_types
            GROUP BY doc_type
            ORDER BY count DESC
        """)
        db.disconnect()
        return jsonify({
            'success': True,
            'types': [dict(r) for r in rows]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# ADDRESSES + GEOCODING + MAP
# ============================================

def _parse_address(text):
    """Wyciągnij adres z tekstu"""
    if not text:
        return None
    import re as _re
    # Pattern: ul. <nazwa> <numer>, <kod> <miasto>
    patterns = [
        r'ul\.?\s+([A-ZŁŻŚĆŃÓĄĘ][a-zżółćńąęś]+(?:\s+[A-Za-zżółćńąęś]+)?)\s+(\d+[a-zA-Z]?(?:/\d+)?),?\s+(\d{2}-\d{3})\s+([A-ZŁŻŚĆŃÓĄĘ][a-zżółćńąęś]+)',
        r'([A-ZŁŻŚĆŃÓĄĘ][a-zżółćńąęś]+)\s+(\d+[a-zA-Z]?(?:/\d+)?),?\s+(\d{2}-\d{3})\s+([A-ZŁŻŚĆŃÓĄĘ][a-zżółćńąęś]+)',
    ]
    for pattern in patterns:
        m = _re.search(pattern, text)
        if m:
            return {
                'street': m.group(1).strip(),
                'number': m.group(2).strip(),
                'postal_code': m.group(3).strip(),
                'city': m.group(4).strip(),
                'full': f"{m.group(1)} {m.group(2)}, {m.group(3)} {m.group(4)}"
            }
    # Fallback: szukaj samego kodu pocztowego + miasta
    m = _re.search(r'(\d{2}-\d{3})\s+([A-ZŁŻŚĆŃÓĄĘ][a-zżółćńąęś]+)', text)
    if m:
        return {
            'street': '',
            'number': '',
            'postal_code': m.group(1),
            'city': m.group(2),
            'full': f"{m.group(1)} {m.group(2)}"
        }
    return None


@app.route('/api/addresses/extract', methods=['POST'])
def addresses_extract():
    """Wyekstrahuj adresy z wszystkich dokumentów"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        pages = db.fetch_all("""
            SELECT p.id as page_id, p.file_id, COALESCE(p.fixed_text, p.text_content) as text
            FROM pages p
            WHERE p.text_content IS NOT NULL
        """)

        addresses_found = 0
        for page in pages:
            text = page['text']
            if not text or len(text) < 50:
                continue

            addr = _parse_address(text[:3000])
            if not addr:
                continue

            # Zapisz adres
            try:
                db.execute("""
                    INSERT OR IGNORE INTO addresses (address_text, street, postal_code, city)
                    VALUES (?, ?, ?, ?)
                """, (addr['full'], addr['street'], addr['postal_code'], addr['city']))

                addr_row = db.fetch_one("SELECT id FROM addresses WHERE address_text = ?", (addr['full'],))
                if addr_row:
                    db.execute("""
                        INSERT INTO address_occurrences (address_id, file_id, page_id)
                        VALUES (?, ?, ?)
                    """, (addr_row['id'], page['file_id'], page['page_id']))
                    addresses_found += 1
            except Exception:
                pass

        db.disconnect()
        return jsonify({'success': True, 'found': addresses_found})
    except Exception as e:
        import traceback
        print(f"❌ addresses_extract: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/addresses/geocode', methods=['POST'])
def addresses_geocode():
    """Geocode adresy używając Nominatim (OpenStreetMap)"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Pobierz adresy bez koordynatów
        addrs = db.fetch_all("""
            SELECT id, address_text, city, postal_code, street
            FROM addresses
            WHERE lat IS NULL OR lng IS NULL
            LIMIT 50
        """)

        geocoded = 0
        failed = 0
        for addr in addrs:
            try:
                query = f"{addr['street']} {addr['postal_code']} {addr['city']}, Poland".strip()
                if not query or len(query) < 5:
                    continue

                resp = requests.get(
                    'https://nominatim.openstreetmap.org/search',
                    params={'q': query, 'format': 'json', 'limit': 1, 'countrycodes': 'pl'},
                    headers={'User-Agent': 'VirtualSegregatory/1.0'},
                    timeout=10
                )
                data = resp.json()
                if data and len(data) > 0:
                    lat = float(data[0]['lat'])
                    lng = float(data[0]['lon'])
                    db.execute("""
                        UPDATE addresses
                        SET lat = ?, lng = ?, geocoded_at = CURRENT_TIMESTAMP, geocode_source = 'nominatim'
                        WHERE id = ?
                    """, (lat, lng, addr['id']))
                    geocoded += 1
                else:
                    failed += 1
                time.sleep(1.1)  # Nominatim: max 1 req/sec
            except Exception as e:
                print(f"⚠ Geocoding {addr['address_text']}: {e}")
                failed += 1

        db.disconnect()
        return jsonify({'success': True, 'geocoded': geocoded, 'failed': failed})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/addresses/map-data', methods=['GET'])
def addresses_map_data():
    """Dane dla mapy: adresy z koordynatami + liczba wystąpień"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT
                a.id, a.address_text, a.city, a.street, a.lat, a.lng,
                COUNT(DISTINCT ao.file_id) as file_count,
                COUNT(DISTINCT ao.id) as occurrence_count
            FROM addresses a
            LEFT JOIN address_occurrences ao ON a.id = ao.address_id
            WHERE a.lat IS NOT NULL AND a.lng IS NOT NULL
            GROUP BY a.id
        """)
        db.disconnect()
        return jsonify({
            'success': True,
            'addresses': [dict(r) for r in rows]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/addresses/<int:address_id>/files', methods=['GET'])
def addresses_files(address_id):
    """Pliki związane z adresem"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT DISTINCT sf.id, sf.filename, sf.page_count, b.name as binder_name
            FROM address_occurrences ao
            JOIN source_files sf ON ao.file_id = sf.id
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE ao.address_id = ?
        """, (address_id,))
        db.disconnect()
        return jsonify({'success': True, 'files': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# EXPORT / RAPORTY
# ============================================

@app.route('/api/export/person/<int:entity_id>/excel', methods=['GET'])
def export_person_excel(entity_id):
    """Eksport profilu osoby do Excel"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        person = db.fetch_one("SELECT * FROM entities WHERE id = ?", (entity_id,))
        if not person:
            db.disconnect()
            return jsonify({'error': 'Osoba nie znaleziona'}), 404

        # Canonical
        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id

        aliases = db.fetch_all("SELECT * FROM person_aliases WHERE canonical_entity_id = ?", (canonical_id,))
        all_ids = [canonical_id] + [a['alias_entity_id'] for a in aliases]
        placeholders = ','.join('?' * len(all_ids))

        # Files
        files = db.fetch_all(f"""
            SELECT DISTINCT sf.id, sf.filename, sf.page_count, b.name as binder_name, sf.created_at
            FROM entity_occurrences eo
            JOIN source_files sf ON eo.file_id = sf.id
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE eo.entity_id IN ({placeholders})
            ORDER BY sf.created_at DESC
        """, all_ids)

        # Attrs
        attrs = db.fetch_all(f"""
            SELECT DISTINCT attr_type, attr_value
            FROM person_attributes
            WHERE entity_id IN ({placeholders})
        """, all_ids)

        db.disconnect()

        # Stwórz Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profil"

        # Header
        ws['A1'] = "PROFIL OSOBY"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = person['entity_value']
        ws['A2'].font = Font(bold=True, size=14)

        # Atrybuty
        ws['A4'] = "Atrybuty"
        ws['A4'].font = Font(bold=True)
        row = 5
        for a in attrs:
            ws[f'A{row}'] = a['attr_type'].upper()
            ws[f'B{row}'] = a['attr_value']
            row += 1

        # Files sheet
        ws_files = wb.create_sheet("Dokumenty")
        ws_files['A1'] = "Nazwa pliku"
        ws_files['B1'] = "Stron"
        ws_files['C1'] = "Segregator"
        ws_files['D1'] = "Data importu"
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_files[col].font = Font(bold=True)
            ws_files[col].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        for i, f in enumerate(files, start=2):
            ws_files[f'A{i}'] = f['filename']
            ws_files[f'B{i}'] = f['page_count']
            ws_files[f'C{i}'] = f['binder_name'] or 'Default'
            ws_files[f'D{i}'] = f['created_at']

        # Auto-width
        for ws_name in [ws, ws_files]:
            for col in ws_name.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_name.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = re.sub(r'[^\w\s-]', '', person['entity_value']).strip().replace(' ', '_')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'profil_{safe_name}.xlsx'
        )
    except Exception as e:
        import traceback
        print(f"❌ export_person_excel: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/person/<int:entity_id>/zip', methods=['GET'])
def export_person_zip(entity_id):
    """ZIP wszystkich PDF-ów osoby"""
    try:
        import zipfile
        from io import BytesIO

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id

        aliases = db.fetch_all("SELECT * FROM person_aliases WHERE canonical_entity_id = ?", (canonical_id,))
        all_ids = [canonical_id] + [a['alias_entity_id'] for a in aliases]
        placeholders = ','.join('?' * len(all_ids))

        files = db.fetch_all(f"""
            SELECT DISTINCT sf.filename, sf.filepath
            FROM entity_occurrences eo
            JOIN source_files sf ON eo.file_id = sf.id
            WHERE eo.entity_id IN ({placeholders})
        """, all_ids)

        person = db.fetch_one("SELECT entity_value FROM entities WHERE id = ?", (canonical_id,))
        db.disconnect()

        # Stwórz ZIP w pamięci
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in files:
                if f['filepath'] and Path(f['filepath']).exists():
                    zf.write(f['filepath'], arcname=f['filename'])

        zip_buffer.seek(0)
        safe_name = re.sub(r'[^\w\s-]', '', person['entity_value']).strip().replace(' ', '_')
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'dokumenty_{safe_name}.zip'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# AUTO-SUMMARY - Llama generuje streszczenia plików
# ============================================

def _generate_file_summary(file_id):
    """Wygeneruj streszczenie pliku przez Llamę"""
    db = DatabaseManager(CONFIG['db_path'])
    db.connect()

    # Pobierz pierwsze 3 strony tekstu
    pages = db.fetch_all("""
        SELECT page_number, COALESCE(fixed_text, text_content) as text
        FROM pages WHERE file_id = ? ORDER BY page_number LIMIT 3
    """, (file_id,))

    file_info = db.fetch_one("SELECT filename FROM source_files WHERE id = ?", (file_id,))
    db.disconnect()

    if not pages or not file_info:
        return False, "Brak treści do streszczenia"

    text = ' '.join((p['text'] or '') for p in pages)[:3000]
    if len(text) < 50:
        return False, "Tekst zbyt krótki"

    if not _llama_check_available():
        return False, "Ollama nie dostępna"

    prompt = f"""Streszcz polski dokument w 1-2 krótkich zdaniach (max 200 znaków).
ZWRÓĆ TYLKO STRESZCZENIE bez wstępu, komentarzy, prefiksów. Zachowaj kluczowe fakty: nazwiska, kwoty, daty, sygnatury.

DOKUMENT: {file_info['filename']}
TREŚĆ:
{text}

STRESZCZENIE (1-2 zdania, max 200 znaków):"""

    try:
        resp = requests.post(
            LLAMA_CONFIG['url'],
            json={
                'model': LLAMA_CONFIG['model'],
                'prompt': prompt,
                'stream': False,
                'options': {'temperature': 0.3, 'num_predict': 150}
            },
            timeout=120
        )
        if resp.status_code == 200:
            summary = resp.json().get('response', '').strip()
            summary = ' '.join(summary.split())[:500]
            short = summary[:150] + ('...' if len(summary) > 150 else '')

            db = DatabaseManager(CONFIG['db_path'])
            db.connect()
            db.execute("DELETE FROM file_summaries WHERE file_id = ?", (file_id,))
            db.execute("""
                INSERT INTO file_summaries (file_id, summary, short_summary, model_used)
                VALUES (?, ?, ?, ?)
            """, (file_id, summary, short, LLAMA_CONFIG['model']))
            db.disconnect()
            return True, summary
        return False, f"HTTP {resp.status_code}"
    except Exception as e:
        return False, str(e)


@app.route('/api/files/<int:file_id>/summary', methods=['GET'])
def file_summary_get(file_id):
    """Pobierz istniejące streszczenie"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        row = db.fetch_one("SELECT * FROM file_summaries WHERE file_id = ?", (file_id,))
        db.disconnect()
        if not row:
            return jsonify({'success': True, 'summary': None})
        return jsonify({'success': True, 'summary': dict(row)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/<int:file_id>/summary/generate', methods=['POST'])
def file_summary_generate(file_id):
    """Wygeneruj streszczenie przez Llamę"""
    try:
        success, result = _generate_file_summary(file_id)
        if success:
            return jsonify({'success': True, 'summary': result})
        return jsonify({'success': False, 'error': result}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/summaries/generate-all', methods=['POST'])
def file_summaries_generate_all():
    """Wygeneruj streszczenia dla wszystkich plików bez summary - dodaje do queue"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        files = db.fetch_all("""
            SELECT sf.id, sf.filename, sf.filepath
            FROM source_files sf
            LEFT JOIN file_summaries fs ON sf.id = fs.file_id
            WHERE fs.id IS NULL
        """)
        db.disconnect()

        for f in files:
            background_processor.add_to_queue(
                f['filepath'] or f['filename'],
                task_type='summary',
                priority=8
            )

        return jsonify({'success': True, 'queued': len(files)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# QUICK RENAME OSOBY (propagacja do aliasów)
# ============================================

@app.route('/api/persons/<int:entity_id>/rename', methods=['POST'])
def person_rename(entity_id):
    """Zmień nazwę osoby - propaguje do wszystkich aliasów"""
    try:
        data = request.json or {}
        new_name = (data.get('new_name') or '').strip()
        propagate = data.get('propagate', True)

        if not new_name:
            return jsonify({'error': 'Pusta nazwa'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # Znajdź canonical
        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id

        # Zmień nazwę canonical
        old_canonical = db.fetch_one("SELECT entity_value FROM entities WHERE id = ?", (canonical_id,))
        old_name = old_canonical['entity_value'] if old_canonical else ''

        db.execute(
            "UPDATE entities SET entity_value = ?, normalized_value = ? WHERE id = ?",
            (new_name, new_name.upper(), canonical_id)
        )

        return_data = {
            'success': True,
            'old_name': old_name,
            'new_name': new_name,
            'updated_count': 1
        }

        # Propaguj do aliasów (opcjonalne)
        if propagate:
            aliases = db.fetch_all(
                "SELECT alias_entity_id FROM person_aliases WHERE canonical_entity_id = ?",
                (canonical_id,)
            )
            return_data['aliases_count'] = len(aliases)

        db.disconnect()
        return jsonify(return_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# CUSTOM TAGS - tagi z kolorami i emoji
# ============================================

@app.route('/api/tags/custom', methods=['GET'])
def tags_custom_list():
    """Lista własnych tagów z kolorami"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT ct.*,
                (SELECT COUNT(*) FROM file_custom_tags WHERE tag_id = ct.id) as usage_count
            FROM custom_tags ct
            ORDER BY usage_count DESC, name ASC
        """)
        db.disconnect()
        return jsonify({'success': True, 'tags': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tags/custom', methods=['POST'])
def tags_custom_create():
    """Utwórz nowy tag"""
    try:
        data = request.json or {}
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'Pusta nazwa'}), 400

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("""
            INSERT OR IGNORE INTO custom_tags (name, color, icon, description)
            VALUES (?, ?, ?, ?)
        """, (name, data.get('color', '#6366f1'), data.get('icon', '🏷️'), data.get('description', '')))
        tag = db.fetch_one("SELECT * FROM custom_tags WHERE name = ?", (name,))
        db.disconnect()
        return jsonify({'success': True, 'tag': dict(tag) if tag else None})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tags/custom/<int:tag_id>', methods=['PUT'])
def tags_custom_update(tag_id):
    """Edytuj tag"""
    try:
        data = request.json or {}
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        updates = []
        params = []
        for field in ['name', 'color', 'icon', 'description']:
            if field in data:
                updates.append(f"{field} = ?")
                params.append(data[field])

        if updates:
            params.append(tag_id)
            db.execute(f"UPDATE custom_tags SET {', '.join(updates)} WHERE id = ?", params)

        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/tags/custom/<int:tag_id>', methods=['DELETE'])
def tags_custom_delete(tag_id):
    """Usuń tag"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("DELETE FROM file_custom_tags WHERE tag_id = ?", (tag_id,))
        db.execute("DELETE FROM custom_tags WHERE id = ?", (tag_id,))
        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/<int:file_id>/tags', methods=['GET'])
def file_tags_get(file_id):
    """Tagi przypisane do pliku"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        rows = db.fetch_all("""
            SELECT ct.* FROM custom_tags ct
            JOIN file_custom_tags fct ON ct.id = fct.tag_id
            WHERE fct.file_id = ?
        """, (file_id,))
        db.disconnect()
        return jsonify({'success': True, 'tags': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/<int:file_id>/tags/<int:tag_id>', methods=['POST'])
def file_tag_add(file_id, tag_id):
    """Dodaj tag do pliku"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute(
            "INSERT OR IGNORE INTO file_custom_tags (file_id, tag_id) VALUES (?, ?)",
            (file_id, tag_id)
        )
        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/files/<int:file_id>/tags/<int:tag_id>', methods=['DELETE'])
def file_tag_remove(file_id, tag_id):
    """Usuń tag z pliku"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute(
            "DELETE FROM file_custom_tags WHERE file_id = ? AND tag_id = ?",
            (file_id, tag_id)
        )
        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# SPOTLIGHT SEARCH - uniwersalne wyszukiwanie
# ============================================

@app.route('/api/spotlight', methods=['GET'])
def spotlight_search():
    """
    Uniwersalne wyszukiwanie wszędzie naraz:
    osoby, dokumenty, segregatory, KW, tagi, akcje
    """
    try:
        query = (request.args.get('q') or '').strip()
        if len(query) < 2:
            return jsonify({'success': True, 'results': {}})

        query_lower = query.lower()
        like_pattern = f'%{query_lower}%'

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        # OSOBY
        persons = db.fetch_all("""
            SELECT e.id, e.entity_value as name,
                (SELECT COUNT(*) FROM entity_occurrences WHERE entity_id = e.id) as occurrences,
                (SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = e.id) as alias_of
            FROM entities e
            WHERE e.entity_type = 'person' AND LOWER(e.entity_value) LIKE ?
            ORDER BY occurrences DESC
            LIMIT 8
        """, (like_pattern,))

        # PLIKI
        files = db.fetch_all("""
            SELECT sf.id, sf.filename, sf.page_count, b.name as binder_name
            FROM source_files sf
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE LOWER(sf.filename) LIKE ?
            ORDER BY sf.id DESC
            LIMIT 8
        """, (like_pattern,))

        # SEGREGATORY
        binders = db.fetch_all("""
            SELECT b.id, b.name,
                (SELECT COUNT(*) FROM source_files WHERE binder_id = b.id) as files_count
            FROM binders b
            WHERE LOWER(b.name) LIKE ?
            ORDER BY files_count DESC
            LIMIT 5
        """, (like_pattern,))

        # KW
        kws = db.fetch_all("""
            SELECT id, kw_full, kw_district FROM land_registers
            WHERE LOWER(kw_full) LIKE ? OR LOWER(kw_district) LIKE ?
            LIMIT 5
        """, (like_pattern, like_pattern))

        # TAGI
        tags = db.fetch_all("""
            SELECT * FROM custom_tags WHERE LOWER(name) LIKE ?
            LIMIT 5
        """, (like_pattern,))

        # TEKST (jeśli zapytanie dłuższe niż 3 znaki)
        text_matches = []
        if len(query) >= 3:
            text_matches = db.fetch_all("""
                SELECT DISTINCT sf.id as file_id, sf.filename, p.page_number,
                    SUBSTR(COALESCE(p.fixed_text, p.text_content), MAX(1, INSTR(LOWER(COALESCE(p.fixed_text, p.text_content)), ?) - 50), 200) as context
                FROM pages p
                JOIN source_files sf ON p.file_id = sf.id
                WHERE LOWER(COALESCE(p.fixed_text, p.text_content)) LIKE ?
                LIMIT 5
            """, (query_lower, like_pattern))

        # AKCJE (statyczne sugestie)
        actions = []
        actions_db = [
            ('Przejdź do: Wyszukiwarka', 'search-results', '🔍'),
            ('Przejdź do: Import plików', 'upload', '📤'),
            ('Przejdź do: Mapa adresów', 'map', '🗺️'),
            ('Przejdź do: Duplikaty', 'duplicates', '🔁'),
            ('Przejdź do: Typy dokumentów', 'doc-types', '📑'),
            ('Przejdź do: Ulubione', 'favorites', '⭐'),
            ('Przejdź do: Operacje (kolejka)', 'bulk-ops', '⚙️'),
            ('Przejdź do: Statystyki', 'statistics', '📊'),
            ('Akcja: Skanuj duplikaty', 'action:scan-duplicates', '🔍'),
            ('Akcja: Wykryj typy dokumentów', 'action:detect-types', '🤖'),
            ('Akcja: Połącz osoby automatycznie', 'action:auto-link', '👥'),
        ]
        for label, target, icon in actions_db:
            if query_lower in label.lower():
                actions.append({'label': label, 'target': target, 'icon': icon})

        db.disconnect()

        return jsonify({
            'success': True,
            'query': query,
            'results': {
                'persons': [dict(p) for p in persons],
                'files': [dict(f) for f in files],
                'binders': [dict(b) for b in binders],
                'kws': [dict(k) for k in kws],
                'tags': [dict(t) for t in tags],
                'text_matches': [dict(t) for t in text_matches],
                'actions': actions[:5]
            }
        })
    except Exception as e:
        import traceback
        print(f"❌ spotlight: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ============================================
# PDF REPORTS - drukowalne raporty
# ============================================

@app.route('/report/person/<int:entity_id>')
def report_person(entity_id):
    """HTML report dla osoby - można wydrukować jako PDF (Ctrl+P)"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        person = db.fetch_one("SELECT * FROM entities WHERE id = ?", (entity_id,))
        if not person:
            db.disconnect()
            return "Osoba nie znaleziona", 404

        canonical_row = db.fetch_one(
            "SELECT canonical_entity_id FROM person_aliases WHERE alias_entity_id = ?",
            (entity_id,)
        )
        canonical_id = canonical_row['canonical_entity_id'] if canonical_row else entity_id
        canonical = db.fetch_one("SELECT * FROM entities WHERE id = ?", (canonical_id,))

        aliases = db.fetch_all("SELECT * FROM person_aliases WHERE canonical_entity_id = ?", (canonical_id,))
        all_ids = [canonical_id] + [a['alias_entity_id'] for a in aliases]
        placeholders = ','.join('?' * len(all_ids))

        # Wszystkie nazwy
        all_names = db.fetch_all(f"SELECT entity_value FROM entities WHERE id IN ({placeholders})", all_ids)

        # Atrybuty
        attrs = db.fetch_all(
            f"SELECT DISTINCT attr_type, attr_value FROM person_attributes WHERE entity_id IN ({placeholders})",
            all_ids
        )

        # Pliki
        files = db.fetch_all(f"""
            SELECT DISTINCT sf.id, sf.filename, sf.page_count, b.name as binder_name,
                COUNT(eo.id) as occurrences,
                MIN(eo.page_id) as first_page_id
            FROM entity_occurrences eo
            JOIN source_files sf ON eo.file_id = sf.id
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE eo.entity_id IN ({placeholders})
            GROUP BY sf.id
            ORDER BY sf.created_at DESC
        """, all_ids)

        # Timeline z datami
        timeline = []
        for f in files:
            doc_date = None
            doc_type = None
            doc_summary = None
            first_page = db.fetch_one(
                "SELECT text_content, fixed_text FROM pages WHERE id = ?",
                (f['first_page_id'],)
            )
            if first_page:
                text = first_page['fixed_text'] or first_page['text_content'] or ''
                doc_date = _extract_date_from_text(text[:1000])
            dt_row = db.fetch_one(
                "SELECT doc_type FROM document_types WHERE file_id = ? ORDER BY confidence DESC LIMIT 1",
                (f['id'],)
            )
            if dt_row:
                doc_type = dt_row['doc_type']
            summ_row = db.fetch_one("SELECT short_summary FROM file_summaries WHERE file_id = ?", (f['id'],))
            if summ_row:
                doc_summary = summ_row['short_summary']
            timeline.append({
                'file_id': f['id'],
                'filename': f['filename'],
                'binder': f['binder_name'],
                'pages': f['page_count'],
                'occurrences': f['occurrences'],
                'date': doc_date or '—',
                'doc_type': doc_type,
                'summary': doc_summary
            })

        timeline.sort(key=lambda t: t['date'] if t['date'] != '—' else '0000', reverse=True)

        # KW
        kws = db.fetch_all(f"""
            SELECT DISTINCT lr.kw_full, lr.kw_district
            FROM entity_occurrences eo
            JOIN land_register_occurrences lro ON eo.page_id = lro.page_id
            JOIN land_registers lr ON lro.kw_id = lr.id
            WHERE eo.entity_id IN ({placeholders})
        """, all_ids)

        db.disconnect()

        return render_template('report_person.html',
            person=dict(canonical),
            all_names=[n['entity_value'] for n in all_names],
            attrs=[dict(a) for a in attrs],
            files=[dict(f) for f in files],
            timeline=timeline,
            kws=[dict(k) for k in kws],
            generated_at=datetime.now().strftime('%Y-%m-%d %H:%M')
        )
    except Exception as e:
        import traceback
        return f"Błąd: {e}<pre>{traceback.format_exc()}</pre>", 500


@app.route('/report/kw/<int:kw_id>')
def report_kw_pdf(kw_id):
    """HTML report dla KW (drukowalny PDF)"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        kw = db.fetch_one("SELECT * FROM land_registers WHERE id = ?", (kw_id,))
        if not kw:
            db.disconnect()
            return "KW nie znaleziona", 404

        # Pliki z tym KW
        files = db.fetch_all("""
            SELECT DISTINCT sf.id, sf.filename, sf.page_count, b.name as binder_name,
                COUNT(DISTINCT lro.page_id) as pages_with_kw
            FROM land_register_occurrences lro
            JOIN source_files sf ON lro.file_id = sf.id
            LEFT JOIN binders b ON sf.binder_id = b.id
            WHERE lro.kw_id = ?
            GROUP BY sf.id
        """, (kw_id,))

        # Osoby związane z KW
        persons = db.fetch_all("""
            SELECT DISTINCT e.id, e.entity_value, COUNT(*) as occurrences
            FROM land_register_occurrences lro
            JOIN entity_occurrences eo ON lro.page_id = eo.page_id
            JOIN entities e ON eo.entity_id = e.id
            WHERE lro.kw_id = ? AND e.entity_type = 'person'
            GROUP BY e.id
            ORDER BY occurrences DESC
            LIMIT 20
        """, (kw_id,))

        db.disconnect()

        return render_template('report_kw.html',
            kw=dict(kw),
            files=[dict(f) for f in files],
            persons=[dict(p) for p in persons],
            generated_at=datetime.now().strftime('%Y-%m-%d %H:%M')
        )
    except Exception as e:
        import traceback
        return f"Błąd: {e}<pre>{traceback.format_exc()}</pre>", 500


# ============================================
# BACKGROUND PROCESSOR - kolejka w tle
# ============================================

class BackgroundProcessor:
    """
    Worker w tle który przetwarza pliki z kolejki processing_queue.
    Może być uruchamiany on-demand lub w harmonogramie (np. w nocy).
    """

    def __init__(self):
        self.running = False
        self.current_task = None
        self.worker_thread = None
        self.stop_event = threading.Event()
        self.wake_event = threading.Event()

        # Harmonogram nocny (domyślnie 22:00 - 06:00)
        self.schedule_enabled = False
        self.schedule_start = dtime(22, 0)  # 22:00
        self.schedule_end = dtime(6, 0)     # 06:00

        # Stats
        self.stats = {
            'total_processed': 0,
            'total_failed': 0,
            'started_at': None,
            'last_activity': None
        }

    def is_in_working_hours(self):
        """Sprawdź czy aktualnie jesteśmy w godzinach pracy schedulera"""
        if not self.schedule_enabled:
            return True  # Bez schedulera - zawsze działa

        now = datetime.now().time()
        start = self.schedule_start
        end = self.schedule_end

        # Jeśli end jest po start (np. 09:00 - 17:00) - tego samego dnia
        if start <= end:
            return start <= now <= end
        else:
            # Przekracza północ (np. 22:00 - 06:00)
            return now >= start or now <= end

    def add_to_queue(self, file_path, binder_name='Default', task_type='full_process', priority=5, section_name=''):
        """Dodaj plik do kolejki przetwarzania"""
        try:
            db = DatabaseManager(CONFIG['db_path'])
            db.connect()
            filename = Path(file_path).name
            db.execute("""
                INSERT INTO processing_queue
                (file_path, filename, binder_name, section_name, task_type, status, priority)
                VALUES (?, ?, ?, ?, ?, 'pending', ?)
            """, (str(file_path), filename, binder_name, section_name, task_type, priority))
            db.disconnect()
            self.wake_event.set()
            return True
        except Exception as e:
            print(f"❌ Błąd dodawania do kolejki: {e}")
            return False

    def get_next_task(self):
        """Pobierz następne zadanie z kolejki"""
        try:
            db = DatabaseManager(CONFIG['db_path'])
            db.connect()

            task = db.fetch_one("""
                SELECT * FROM processing_queue
                WHERE status = 'pending'
                ORDER BY priority ASC, created_at ASC
                LIMIT 1
            """)

            if task:
                # Mark as processing
                db.execute("""
                    UPDATE processing_queue
                    SET status = 'processing', started_at = CURRENT_TIMESTAMP, progress_percent = 0
                    WHERE id = ?
                """, (task['id'],))
                task = dict(task)

            db.disconnect()
            return task
        except Exception as e:
            print(f"❌ Błąd pobierania zadania: {e}")
            return None

    def mark_completed(self, task_id, file_id=None):
        """Oznacz zadanie jako zakończone"""
        try:
            db = DatabaseManager(CONFIG['db_path'])
            db.connect()
            db.execute("""
                UPDATE processing_queue
                SET status = 'completed', completed_at = CURRENT_TIMESTAMP,
                    progress_percent = 100, file_id = ?
                WHERE id = ?
            """, (file_id, task_id))
            db.disconnect()
        except Exception as e:
            print(f"❌ Błąd mark_completed: {e}")

    def mark_failed(self, task_id, error_msg):
        """Oznacz zadanie jako nieudane (z retry)"""
        try:
            db = DatabaseManager(CONFIG['db_path'])
            db.connect()
            task = db.fetch_one("SELECT retry_count, max_retries FROM processing_queue WHERE id = ?", (task_id,))
            if task and task['retry_count'] < task['max_retries']:
                # Retry - przywróć do pending
                db.execute("""
                    UPDATE processing_queue
                    SET status = 'pending', retry_count = retry_count + 1, error_message = ?,
                        started_at = NULL
                    WHERE id = ?
                """, (error_msg[:500], task_id))
            else:
                # Final fail
                db.execute("""
                    UPDATE processing_queue
                    SET status = 'failed', completed_at = CURRENT_TIMESTAMP, error_message = ?
                    WHERE id = ?
                """, (error_msg[:500], task_id))
            db.disconnect()
        except Exception as e:
            print(f"❌ Błąd mark_failed: {e}")

    def update_progress(self, task_id, percent, text=""):
        """Aktualizuj progress zadania"""
        try:
            db = DatabaseManager(CONFIG['db_path'])
            db.connect()
            db.execute("""
                UPDATE processing_queue
                SET progress_percent = ?, progress_text = ?
                WHERE id = ?
            """, (int(percent), text[:200], task_id))
            db.disconnect()
        except Exception:
            pass

    def process_task(self, task):
        """Przetwarzaj pojedyncze zadanie"""
        task_id = task['id']
        file_path = task['file_path']
        binder_name = task.get('binder_name', 'Default')
        section_name = normalize_section_name(task.get('section_name'))
        task_type = task.get('task_type', 'full_process')

        try:
            print(f"\n🔄 [QUEUE] Rozpoczynam: {task['filename']}")
            self.update_progress(task_id, 10, "Otwieranie pliku...")

            if task_type == 'full_process':
                # Pełen pipeline: OCR + ekstrakcja
                self.update_progress(task_id, 20, "OCR + ekstrakcja...")
                result = auto_process_pdf(Path(file_path), binder_name)

                if result and result.get('success'):
                    file_id = result.get('file_id')
                    if section_name and file_id:
                        db = DatabaseManager(CONFIG['db_path'])
                        db.connect()
                        db.execute("UPDATE source_files SET section_name = ? WHERE id = ?", (section_name, file_id))
                        db.disconnect()
                    self.mark_completed(task_id, file_id=file_id)
                    print(f"✅ [QUEUE] Zakończone: {task['filename']} (file_id={file_id})")

                    # Po pełnym procesie - dodaj task na llama_fix (jeśli włączone)
                    if AUTO_LLAMA_FIX and file_id and _llama_check_available():
                        self.add_to_queue(file_path, binder_name, task_type='llama_fix', priority=7)
                        print(f"   📝 Dodano do kolejki llama_fix")

                    return True
                else:
                    self.mark_failed(task_id, "auto_process_pdf zwrócił None")
                    return False

            elif task_type == 'summary':
                # Wygeneruj streszczenie pliku przez Llamę
                db = DatabaseManager(CONFIG['db_path'])
                db.connect()
                file_row = db.fetch_one("SELECT id FROM source_files WHERE filepath = ?", (str(file_path),))
                if not file_row:
                    file_row = db.fetch_one("SELECT id FROM source_files WHERE filename = ?", (task['filename'],))
                db.disconnect()

                if not file_row:
                    self.mark_failed(task_id, "Plik nie znaleziony")
                    return False

                self.update_progress(task_id, 50, "Llama generuje streszczenie...")
                success, result = _generate_file_summary(file_row['id'])
                if success:
                    self.mark_completed(task_id, file_id=file_row['id'])
                    return True
                else:
                    self.mark_failed(task_id, str(result))
                    return False

            elif task_type == 'llama_fix':
                # Tylko poprawa OCR przez Llamę
                self.update_progress(task_id, 30, "Llama poprawia OCR...")

                # Znajdź file_id po ścieżce
                db = DatabaseManager(CONFIG['db_path'])
                db.connect()
                file_row = db.fetch_one("SELECT id FROM source_files WHERE filepath = ?", (str(file_path),))
                if not file_row:
                    db.disconnect()
                    self.mark_failed(task_id, "Plik nie znaleziony w bazie")
                    return False

                file_id = file_row['id']
                pages = db.fetch_all("""
                    SELECT id, page_number, text_content
                    FROM pages
                    WHERE file_id = ? AND fixed_text IS NULL AND text_content IS NOT NULL
                    ORDER BY page_number
                """, (file_id,))
                db.disconnect()

                if not pages:
                    self.mark_completed(task_id, file_id=file_id)
                    return True

                total = len(pages)
                fixed = 0
                for i, page in enumerate(pages):
                    if self.stop_event.is_set() or not self.is_in_working_hours():
                        # Przerwij i wróć do pending
                        self.update_progress(task_id, int(i / total * 100), f"Przerwano na stronie {page['page_number']}")
                        db = DatabaseManager(CONFIG['db_path'])
                        db.connect()
                        db.execute(
                            "UPDATE processing_queue SET status = 'pending', started_at = NULL WHERE id = ?",
                            (task_id,)
                        )
                        db.disconnect()
                        return False

                    text = page['text_content']
                    if not text or len(text.strip()) < 20:
                        continue

                    self.update_progress(task_id, int((i / total) * 100), f"Llama: strona {page['page_number']}/{total}")

                    try:
                        success, fixed_text = _llama_correct_text(text, context=f"Plik {task['filename']}, str. {page['page_number']}")
                        if success and fixed_text:
                            db = DatabaseManager(CONFIG['db_path'])
                            db.connect()
                            db.execute("""
                                UPDATE pages SET fixed_text = ?, ocr_fixed_at = CURRENT_TIMESTAMP, ocr_confidence = 'auto'
                                WHERE id = ?
                            """, (fixed_text, page['id']))
                            db.disconnect()
                            fixed += 1
                    except Exception as e:
                        print(f"   ⚠ Strona {page['page_number']}: {e}")

                self.mark_completed(task_id, file_id=file_id)
                print(f"✅ [QUEUE] Llama poprawił {fixed}/{total} stron w {task['filename']}")
                return True

            else:
                self.mark_failed(task_id, f"Nieznany task_type: {task_type}")
                return False

        except Exception as e:
            import traceback
            print(f"❌ [QUEUE] Błąd w {task['filename']}: {e}")
            print(traceback.format_exc())
            self.mark_failed(task_id, str(e))
            return False

    def worker_loop(self):
        """Główna pętla workera"""
        print(f"🚀 BackgroundProcessor: worker uruchomiony")
        self.stats['started_at'] = datetime.now().isoformat()

        while not self.stop_event.is_set():
            try:
                # Sprawdź czy jesteśmy w godzinach pracy
                if not self.is_in_working_hours():
                    self.current_task = None
                    # Sprawdzaj co minutę
                    self.wake_event.wait(timeout=60)
                    self.wake_event.clear()
                    continue

                # Pobierz zadanie
                task = self.get_next_task()
                if not task:
                    # Brak zadań - czekaj 10s
                    self.current_task = None
                    self.wake_event.wait(timeout=1)
                    self.wake_event.clear()
                    continue

                self.current_task = {
                    'id': task['id'],
                    'filename': task['filename'],
                    'task_type': task['task_type'],
                    'started_at': datetime.now().isoformat()
                }

                success = self.process_task(task)
                if success:
                    self.stats['total_processed'] += 1
                else:
                    self.stats['total_failed'] += 1

                self.stats['last_activity'] = datetime.now().isoformat()
                self.current_task = None

                # Krótka pauza między taskami
                self.wake_event.wait(timeout=0.2)
                self.wake_event.clear()

            except Exception as e:
                print(f"❌ Worker exception: {e}")
                time.sleep(5)

        print(f"⏹ BackgroundProcessor: worker zatrzymany")

    def start(self):
        """Uruchom worker w tle (daemon thread)"""
        if self.running:
            return False
        self.running = True
        self.stop_event.clear()
        self.wake_event.set()
        self.worker_thread = threading.Thread(target=self.worker_loop, daemon=True)
        self.worker_thread.start()
        return True

    def stop(self):
        """Zatrzymaj worker"""
        if not self.running:
            return False
        self.running = False
        self.stop_event.set()
        self.wake_event.set()
        if self.worker_thread:
            self.worker_thread.join(timeout=5)
        return True

    def get_status(self):
        """Pobierz status processora"""
        try:
            db = DatabaseManager(CONFIG['db_path'])
            db.connect()
            counts = db.fetch_one("""
                SELECT
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN status = 'processing' THEN 1 ELSE 0 END) as processing,
                    SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed,
                    COUNT(*) as total
                FROM processing_queue
            """)
            db.disconnect()

            queue_stats = dict(counts) if counts else {}
            # Convert None -> 0
            for k in queue_stats:
                queue_stats[k] = queue_stats[k] or 0

            return {
                'running': self.running,
                'in_working_hours': self.is_in_working_hours(),
                'current_task': self.current_task,
                'schedule': {
                    'enabled': self.schedule_enabled,
                    'start': self.schedule_start.strftime('%H:%M'),
                    'end': self.schedule_end.strftime('%H:%M')
                },
                'queue': queue_stats,
                'stats': self.stats
            }
        except Exception as e:
            return {
                'running': self.running,
                'error': str(e)
            }


# Inicjalizacja globalnego processora
AUTO_LLAMA_FIX = False  # Czy po OCR od razu dodawać Llama fix do kolejki
background_processor = BackgroundProcessor()


# ============================================
# BACKGROUND PROCESSOR ENDPOINTS
# ============================================

@app.route('/api/background/status', methods=['GET'])
def bg_status():
    """Pobierz status background processora"""
    return jsonify(background_processor.get_status())


@app.route('/api/background/start', methods=['POST'])
def bg_start():
    """Uruchom background worker"""
    started = background_processor.start()
    return jsonify({
        'success': True,
        'message': 'Worker uruchomiony' if started else 'Worker już działa',
        'status': background_processor.get_status()
    })


@app.route('/api/background/stop', methods=['POST'])
def bg_stop():
    """Zatrzymaj background worker"""
    stopped = background_processor.stop()
    return jsonify({
        'success': True,
        'message': 'Worker zatrzymany' if stopped else 'Worker nie był uruchomiony'
    })


@app.route('/api/background/schedule', methods=['POST'])
def bg_set_schedule():
    """Ustaw harmonogram pracy (godziny nocne)"""
    try:
        data = request.json
        if 'enabled' in data:
            background_processor.schedule_enabled = bool(data['enabled'])
        if 'start' in data:
            h, m = data['start'].split(':')
            background_processor.schedule_start = dtime(int(h), int(m))
        if 'end' in data:
            h, m = data['end'].split(':')
            background_processor.schedule_end = dtime(int(h), int(m))
        if 'auto_llama_fix' in data:
            global AUTO_LLAMA_FIX
            AUTO_LLAMA_FIX = bool(data['auto_llama_fix'])

        return jsonify({
            'success': True,
            'schedule': {
                'enabled': background_processor.schedule_enabled,
                'start': background_processor.schedule_start.strftime('%H:%M'),
                'end': background_processor.schedule_end.strftime('%H:%M')
            },
            'auto_llama_fix': AUTO_LLAMA_FIX
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/background/queue', methods=['GET'])
def bg_get_queue():
    """Pobierz listę zadań w kolejce"""
    try:
        status_filter = request.args.get('status', 'all')
        limit = int(request.args.get('limit', 50))

        db = DatabaseManager(CONFIG['db_path'])
        db.connect()

        if status_filter == 'all':
            tasks = db.fetch_all("""
                SELECT * FROM processing_queue
                ORDER BY
                    CASE status
                        WHEN 'processing' THEN 1
                        WHEN 'pending' THEN 2
                        WHEN 'failed' THEN 3
                        WHEN 'completed' THEN 4
                    END,
                    created_at DESC
                LIMIT ?
            """, (limit,))
        else:
            tasks = db.fetch_all(
                "SELECT * FROM processing_queue WHERE status = ? ORDER BY created_at DESC LIMIT ?",
                (status_filter, limit)
            )

        db.disconnect()

        return jsonify({
            'success': True,
            'tasks': [dict(t) for t in tasks]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/background/clear-completed', methods=['POST'])
def bg_clear_completed():
    """Wyczyść zakończone zadania z kolejki"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        cursor = db.execute("DELETE FROM processing_queue WHERE status IN ('completed', 'failed')")
        db.disconnect()
        return jsonify({'success': True, 'message': 'Wyczyszczono zakończone zadania'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/background/retry-failed', methods=['POST'])
def bg_retry_failed():
    """Resetuj failed zadania do pending"""
    try:
        db = DatabaseManager(CONFIG['db_path'])
        db.connect()
        db.execute("""
            UPDATE processing_queue
            SET status = 'pending', retry_count = 0, error_message = NULL, started_at = NULL
            WHERE status = 'failed'
        """)
        db.disconnect()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================
# MULTI-FILE UPLOAD
# ============================================

@app.route('/api/upload-multiple', methods=['POST'])
def upload_multiple():
    """
    Upload wielu plików naraz - tylko zapisuje + dodaje do kolejki.
    Pełna analiza dzieje się w tle (worker przetwarza kolejkę).
    """
    try:
        files = request.files.getlist('files')
        binder_name = request.form.get('binder_name', 'Default')
        priority = int(request.form.get('priority', 5))
        auto_start = request.form.get('auto_start', 'true').lower() == 'true'

        if not files:
            return jsonify({'error': 'Brak plików w żądaniu'}), 400

        imports_dir = Path(CONFIG['imports_dir'])
        imports_dir.mkdir(parents=True, exist_ok=True)

        added = []
        skipped = []
        errors = []

        for file in files:
            try:
                if not file.filename:
                    continue

                if not file.filename.lower().endswith('.pdf'):
                    skipped.append(f"{file.filename} (nie PDF)")
                    continue

                # Zapisz plik
                filepath = imports_dir / file.filename
                file.save(str(filepath))

                # Dodaj do kolejki
                background_processor.add_to_queue(
                    filepath,
                    binder_name=binder_name,
                    task_type='full_process',
                    priority=priority
                )

                added.append(file.filename)
                print(f"📥 Dodano do kolejki: {file.filename}")

            except Exception as e:
                errors.append(f"{file.filename}: {str(e)}")
                print(f"❌ Błąd zapisu {file.filename}: {e}")

        # Auto-start worker jeśli nie uruchomiony
        if auto_start and not background_processor.running:
            background_processor.start()

        return jsonify({
            'success': True,
            'added': added,
            'added_count': len(added),
            'skipped': skipped,
            'errors': errors,
            'queue_size': len(added),
            'worker_running': background_processor.running,
            'message': f'Dodano {len(added)} plików do kolejki'
        })

    except Exception as e:
        import traceback
        print(f"❌ upload_multiple error: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("\n🌐 Virtual Segregatory Web UI")
    print("=" * 50)
    print("✅ Otwórz przeglądarkę: http://localhost:5001")
    print("=" * 50 + "\n")

    # Auto-uruchom background worker
    print("🚀 Uruchamiam background worker...")
    background_processor.start()
    print("✅ Worker działa - dodaj pliki przez UI")

    app.run(debug=False, host='localhost', port=5001, use_reloader=False)
